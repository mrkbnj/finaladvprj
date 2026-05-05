import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
import uuid
import qrcode
import re
import hashlib
from datetime import datetime, timezone, timedelta

_PST = timezone(timedelta(hours=8))   # Philippine Standard Time (UTC+8)
def _now():
    """Return the current datetime in Philippine Standard Time."""
    return datetime.now(_PST).replace(tzinfo=None)

import sys

if getattr(sys, 'frozen', False):
    # Running as a PyInstaller EXE
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Running as a normal Python script
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE = os.path.join(BASE_DIR, "warehouse.xlsx")
LOG_FILE = FILE  # Logs stored inside warehouse.xlsx as a very hidden sheet
LOG_SHEET = "activity_logs"
USERS_FILE = os.path.join(BASE_DIR, "users.xlsx")
QR_FOLDER = os.path.join(BASE_DIR, "qr_codes")
QR_FOLDER_W1 = os.path.join(QR_FOLDER, "warehouse_1")
QR_FOLDER_W2 = os.path.join(QR_FOLDER, "warehouse_2")
QR_LABELS_FOLDER = os.path.join(BASE_DIR, "qr_labels")
QR_LABELS_FOLDER_W1 = os.path.join(QR_LABELS_FOLDER, "warehouse_1")
QR_LABELS_FOLDER_W2 = os.path.join(QR_LABELS_FOLDER, "warehouse_2")
EXCEL_FOLDER = os.path.join(BASE_DIR, "excel_exports")
EXCEL_FOLDER_W1 = os.path.join(EXCEL_FOLDER, "warehouse_1")
EXCEL_FOLDER_W2 = os.path.join(EXCEL_FOLDER, "warehouse_2")
DUMP_FOLDER       = os.path.join(BASE_DIR, "dump")
DUMP_EXCEL_W1     = os.path.join(DUMP_FOLDER, "excel_exports", "warehouse_1")
DUMP_EXCEL_W2     = os.path.join(DUMP_FOLDER, "excel_exports", "warehouse_2")
DUMP_LABELS_W1    = os.path.join(DUMP_FOLDER, "qr_labels", "warehouse_1")
DUMP_LABELS_W2    = os.path.join(DUMP_FOLDER, "qr_labels", "warehouse_2")
PULL_QR_FOLDER    = os.path.join(BASE_DIR, "pull_qrs")
PULL_QR_FOLDER_W1 = os.path.join(PULL_QR_FOLDER, "warehouse_1")
PULL_QR_FOLDER_W2 = os.path.join(PULL_QR_FOLDER, "warehouse_2")
PULL_EXCEL_FOLDER    = os.path.join(BASE_DIR, "pull_excel")
PULL_EXCEL_FOLDER_W1 = os.path.join(PULL_EXCEL_FOLDER, "warehouse_1")
PULL_EXCEL_FOLDER_W2 = os.path.join(PULL_EXCEL_FOLDER, "warehouse_2")

QR_FOLDER_YK          = os.path.join(QR_FOLDER,          "yubikey")
QR_LABELS_FOLDER_YK   = os.path.join(QR_LABELS_FOLDER,   "yubikey")
EXCEL_FOLDER_YK        = os.path.join(EXCEL_FOLDER,       "yubikey")
PULL_QR_FOLDER_YK      = os.path.join(PULL_QR_FOLDER,     "yubikey")
PULL_EXCEL_FOLDER_YK   = os.path.join(PULL_EXCEL_FOLDER,  "yubikey")
DUMP_EXCEL_YK          = os.path.join(DUMP_FOLDER, "excel_exports", "yubikey")
DUMP_LABELS_YK         = os.path.join(DUMP_FOLDER, "qr_labels",     "yubikey")

SHELVES = [
    "Area A", "Area B", "Area C",
    "Rack 1 - Bay 1", "Rack 1 - Bay 2", "Rack 1 - Bay 3",
    "Rack 2 - Bay 1", "Rack 2 - Bay 2", "Rack 2 - Bay 3",
]
SHELVES_W1 = SHELVES_W2 = SHELVES

EQUIPMENT_TYPES = ["Monitor", "Keyboard", "Mouse", "Headset"]
STATUS_CHOICES  = ["No Issue", "Minimal", "Defective", "Missing"]

# ========== TOOLTIP ==========

class Tooltip:
    """Show a single tooltip at a time when the user hovers over a widget."""
    _active = None  # class-level: only one tooltip visible at a time

    def __init__(self, widget, text):
        self.widget   = widget
        self.text     = text
        self._tip_win = None
        widget.bind("<Enter>",   self._show)
        widget.bind("<Leave>",   self._hide)
        widget.bind("<Destroy>", self._hide)

    def _show(self, event=None):
        # Close any currently open tooltip first
        if Tooltip._active and Tooltip._active is not self:
            Tooltip._active._hide()
        if self._tip_win or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tip_win = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(tw, text=self.text, justify="left",
                 background="#ffffe0", relief="solid", borderwidth=1,
                 font=("Helvetica", 8), wraplength=220,
                 padx=4, pady=3).pack()
        Tooltip._active = self

    def _hide(self, event=None):
        tw, self._tip_win = self._tip_win, None
        if tw:
            tw.destroy()
        if Tooltip._active is self:
            Tooltip._active = None

def tip(widget, text):
    """Attach a tooltip to widget. Returns the widget for inline use."""
    Tooltip(widget, text)
    return widget

staged_items = []
selected_staged_index = None
staged_sets = []
selected_set_index = None
current_user = ""
current_is_admin = False
session_start = ""

# Track last generated Excel paths per warehouse (for VIEW EXCEL button)
_last_excel_path = {1: None, 2: None, 3: None}

# ── Checkbox selection state (warehouse table rows) ──────────
# Maps tree iid -> bool; populated/cleared on every table reload.
w1_row_checks: dict = {}   # W1 warehouse table checkboxes
w2_row_checks: dict = {}   # W2 warehouse table checkboxes
w1_pull_row_checks: dict = {}  # W1 pull history table checkboxes
w2_pull_row_checks: dict = {}  # W2 pull history table checkboxes
w1_persistent_checks: set = set()  # Hostnames persistently checked in W1 (survives search/filter)
w2_persistent_checks: set = set()  # (set_id, equip_type) tuples persistently checked in W2
w1_pull_persistent_checks: set = set()  # Hostnames persistently checked in W1 pull history
w2_pull_persistent_checks: set = set()  # (set_id, equip_type) tuples persistently checked in W2 pull history

# ── Yubikey state ─────────────────────────────────────────
staged_yk_items = []
selected_staged_yk_index = None
yk_row_checks:        dict = {}
yk_pull_row_checks:   dict = {}
yk_persistent_checks:      set = set()
yk_pull_persistent_checks: set = set()

# ========== INITIALIZATION ==========

def initialize_file():
    # Always ensure all folders exist
    for folder in (QR_FOLDER_W1, QR_FOLDER_W2, QR_FOLDER_YK,
                   QR_LABELS_FOLDER_W1, QR_LABELS_FOLDER_W2, QR_LABELS_FOLDER_YK,
                   EXCEL_FOLDER_W1, EXCEL_FOLDER_W2, EXCEL_FOLDER_YK,
                   PULL_QR_FOLDER_W1, PULL_QR_FOLDER_W2, PULL_QR_FOLDER_YK,
                   PULL_EXCEL_FOLDER_W1, PULL_EXCEL_FOLDER_W2, PULL_EXCEL_FOLDER_YK,
                   DUMP_EXCEL_W1, DUMP_EXCEL_W2, DUMP_EXCEL_YK,
                   DUMP_LABELS_W1, DUMP_LABELS_W2, DUMP_LABELS_YK):
        os.makedirs(folder, exist_ok=True)

    sheets_to_create = {}
    if not os.path.exists(FILE):
        sheets_to_create = {"items": None, "shelves": None, "pullouts": None,
                            "items_w2": None, "shelves_w2": None, "pullouts_w2": None,
                            "items_yk": None, "shelves_yk": None, "pullouts_yk": None}
        mode = 'w'
    else:
        try:
            with pd.ExcelFile(FILE) as xls:
                existing = xls.sheet_names
        except Exception as e:
            messagebox.showerror("File Error", f"Could not read '{FILE}':\n{e}")
            return
        needed = ["items", "shelves", "pullouts", "items_w2", "shelves_w2", "pullouts_w2",
                  "items_yk", "shelves_yk", "pullouts_yk"]
        sheets_to_create = {s: None for s in needed if s not in existing}
        mode = 'a'

    if not sheets_to_create:
        return

    default_dfs = {
        # W1 Sheets
        "items": pd.DataFrame(columns=["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]),
        "shelves": pd.DataFrame({"Shelf": SHELVES_W1, "Status": ["AVAILABLE"] * len(SHELVES_W1), "Date_Full": [None] * len(SHELVES_W1)}),
        "pullouts": pd.DataFrame(columns=["Hostname", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]),
        # W2 Sheets
        "items_w2": pd.DataFrame(columns=["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]),
        "shelves_w2": pd.DataFrame({"Shelf": SHELVES_W2, "Status": ["AVAILABLE"] * len(SHELVES_W2), "Date_Full": [None] * len(SHELVES_W2)}),
        "pullouts_w2": pd.DataFrame(columns=["Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]),
        # YK Sheets
        "items_yk":   pd.DataFrame(columns=["QR", "Hostname", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]),
        "shelves_yk": pd.DataFrame({"Shelf": SHELVES, "Status": ["AVAILABLE"] * len(SHELVES), "Date_Full": [None] * len(SHELVES)}),
        "pullouts_yk": pd.DataFrame(columns=["QR", "Hostname", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]),
    }
    try:
        with pd.ExcelWriter(FILE, engine='openpyxl', mode=mode) as writer:
            for sheet in sheets_to_create:
                default_dfs[sheet].to_excel(writer, sheet_name=sheet, index=False)
            for ws in writer.book.worksheets:
                ws.protection.sheet = True
                ws.protection.enable()
    except Exception as e:
        messagebox.showerror("File Error", f"Could not create '{FILE}':\n{e}\n\nMake sure the file is not open in Excel.")

def initialize_log():
    from openpyxl import load_workbook
    if not os.path.exists(FILE):
        initialize_file()
    try:
        wb = load_workbook(FILE)
        if LOG_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(LOG_SHEET)
            ws.append(["Timestamp", "User", "Action", "Details"])
            ws.sheet_state = "veryHidden"
            ws.protection.sheet = True
            ws.protection.enable()
            wb.save(FILE)
    except Exception as e:
        print(f"[initialize_log error] {e}")

# ========== USERS DATABASE ==========

def _hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()

def _is_admin_password(pw: str) -> bool:
    """Admin passwords must contain all of: ! @ # $"""
    return all(c in pw for c in '!@#$')

def initialize_users():
    """Create users.xlsx with a default admin account if it doesn't exist."""
    if not os.path.exists(USERS_FILE):
        default_pw = "Admin@123"          # contains @  → admin role
        df = pd.DataFrame([{
            "Username": "admin",
            "Password": _hash_password(default_pw),
            "Role":     "admin",
            "Created":  _now().strftime("%Y-%m-%d %H:%M:%S"),
        }])
        with pd.ExcelWriter(USERS_FILE, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="users", index=False)
            for ws in writer.book.worksheets:
                ws.protection.sheet = True
                ws.protection.enable()

def load_users() -> pd.DataFrame:
    """Return the users DataFrame (always with expected columns)."""
    initialize_users()
    from openpyxl import load_workbook
    wb = load_workbook(USERS_FILE, data_only=True)
    ws = wb["users"]
    ws.protection.sheet = False
    import tempfile
    tmp = USERS_FILE + ".~tmp.xlsx"
    wb.save(tmp)
    df = pd.read_excel(tmp, sheet_name="users")
    try:
        os.remove(tmp)
    except Exception:
        pass
    for col in ["Username", "Password", "Role", "Created"]:
        if col not in df.columns:
            df[col] = ""
    return df

def _save_users(df: pd.DataFrame):
    with pd.ExcelWriter(USERS_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="users", index=False)
        for ws in writer.book.worksheets:
            ws.protection.sheet = True
            ws.protection.enable()

def authenticate_user(username: str, password: str):
    """Return (ok, role) or (False, None)."""
    df = load_users()
    row = df[df["Username"].str.lower() == username.strip().lower()]
    if row.empty:
        return False, None
    stored_hash = str(row.iloc[0]["Password"])
    if stored_hash != _hash_password(password):
        return False, None
    role = str(row.iloc[0].get("Role", "user")).lower()
    return True, role

def create_account(username: str, password: str) -> str:
    """Create a new account. Returns error string or '' on success."""
    username = username.strip()
    if not username:
        return "Username cannot be empty."
    if len(username) < 3:
        return "Username must be at least 3 characters."
    if not re.match(r'^[A-Za-z0-9_]+$', username):
        return "Username may only contain letters, digits, and underscores."
    if len(password) < 6:
        return "Password must be at least 6 characters."
    df = load_users()
    if username.lower() in df["Username"].str.lower().tolist():
        return f"Username '{username}' already exists."
    role = "admin" if _is_admin_password(password) else "user"
    new_row = pd.DataFrame([{
        "Username": username,
        "Password": _hash_password(password),
        "Role":     role,
        "Created":  _now().strftime("%Y-%m-%d %H:%M:%S"),
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    _save_users(df)
    return ""

def delete_account(username: str) -> str:
    """Delete account by username. Returns error string or '' on success."""
    df = load_users()
    match = df[df["Username"].str.lower() == username.strip().lower()]
    if match.empty:
        return f"Account '{username}' not found."
    if str(match.iloc[0]["Role"]).lower() == "admin":
        # prevent deleting the last admin
        remaining_admins = df[df["Role"].str.lower() == "admin"]
        if len(remaining_admins) <= 1:
            return "Cannot delete the only admin account."
    df = df[df["Username"].str.lower() != username.strip().lower()].reset_index(drop=True)
    _save_users(df)
    return ""

def change_password(username: str, new_password: str) -> str:
    """Change password and update role. Returns error string or '' on success."""
    if len(new_password) < 6:
        return "Password must be at least 6 characters."
    df = load_users()
    idx = df[df["Username"].str.lower() == username.strip().lower()].index
    if idx.empty:
        return f"Account '{username}' not found."
    i = idx[0]
    df.at[i, "Password"] = _hash_password(new_password)
    df.at[i, "Role"]     = "admin" if _is_admin_password(new_password) else "user"
    _save_users(df)
    return ""



def _load_sheet(file, sheet, init_fn):
    def _read():
        # Read with openpyxl and strip sheet protection so data loads correctly
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        ws = wb[sheet]
        ws.protection.sheet = False   # temporarily lift for pandas
        import tempfile, os
        tmp = file + ".~tmp.xlsx"
        wb.save(tmp)
        df = pd.read_excel(tmp, sheet_name=sheet)
        try:
            os.remove(tmp)
        except Exception:
            pass
        return df

    try:
        return _read()
    except Exception:
        try:
            init_fn()
            return _read()
        except Exception as e:
            messagebox.showerror("File Error",
                f"Could not load sheet '{sheet}' from '{file}':\n{e}\n\n"
                "Make sure the file is not open in Excel and the folder is writable.")
            return pd.DataFrame()

def load_items():
    df = _load_sheet(FILE, "items", initialize_file)
    expected = ["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]
    for col in expected:
        if col not in df.columns:
            df[col] = pd.Series(dtype=str)
    return df
def load_shelves():     return _load_sheet(FILE, "shelves", initialize_file) # W1 Only
def load_shelves_w2():  return _load_sheet(FILE, "shelves_w2", initialize_file) # W2 Only
def load_pullouts():    return _load_sheet(FILE, "pullouts", initialize_file)
def load_items_w2():
    df = _load_sheet(FILE, "items_w2", initialize_file)
    expected = ["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]
    for col in expected:
        if col not in df.columns:
            df[col] = pd.Series(dtype=str)
    return df
def load_pullouts_w2(): return _load_sheet(FILE, "pullouts_w2", initialize_file)
def load_logs():
    from openpyxl import load_workbook
    try:
        initialize_log()
        wb = load_workbook(FILE, data_only=True)
        if LOG_SHEET not in wb.sheetnames:
            return pd.DataFrame(columns=["Timestamp", "User", "Action", "Details"])
        ws = wb[LOG_SHEET]
        ws.sheet_state = "visible"
        ws.protection.sheet = False
        data = [row for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row)]
        if not data:
            return pd.DataFrame(columns=["Timestamp", "User", "Action", "Details"])
        headers = data[0]
        rows = data[1:]
        return pd.DataFrame(rows, columns=list(headers))
    except Exception as e:
        return pd.DataFrame(columns=["Timestamp", "User", "Action", "Details"])

def _excel_locked_error():
    messagebox.showerror(
        "File Locked",
        f"Cannot save — '{FILE}' is open in Excel or another program.\n\n"
        "Please close the file and try again."
    )

def _write_all_sheets(df_items, df_shelves, df_pullouts, df_items_w2, df_shelves_w2, df_po2,
                      df_items_yk=None, df_shelves_yk=None, df_pullouts_yk=None):
    """Single write point for the warehouse Excel file. Raises on failure."""
    # Load YK data if not supplied so it is never accidentally wiped by W1/W2 saves
    if df_items_yk is None:
        try:    df_items_yk = load_items_yk()
        except: df_items_yk = pd.DataFrame(columns=["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"])
    if df_shelves_yk is None:
        try:    df_shelves_yk = load_shelves_yk()
        except: df_shelves_yk = pd.DataFrame({"Shelf": SHELVES, "Status": ["AVAILABLE"]*len(SHELVES), "Date_Full": [None]*len(SHELVES)})
    if df_pullouts_yk is None:
        try:    df_pullouts_yk = load_pullouts_yk()
        except: df_pullouts_yk = pd.DataFrame(columns=["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"])
    from openpyxl import load_workbook
    try:
        # Preserve the activity_logs sheet before overwriting
        existing_log_ws = None
        if os.path.exists(FILE):
            try:
                wb_old = load_workbook(FILE, data_only=True)
                if LOG_SHEET in wb_old.sheetnames:
                    ws_old = wb_old[LOG_SHEET]
                    ws_old.sheet_state = "visible"
                    ws_old.protection.sheet = False
                    log_data = [
                        tuple(cell.value for cell in row)
                        for row in ws_old.iter_rows()
                        if any(cell.value is not None for cell in row)
                    ]
                    if not log_data:
                        log_data = [("Timestamp", "User", "Action", "Details")]
                else:
                    log_data = [("Timestamp", "User", "Action", "Details")]
            except Exception:
                log_data = [("Timestamp", "User", "Action", "Details")]
        else:
            log_data = [("Timestamp", "User", "Action", "Details")]

        with pd.ExcelWriter(FILE, engine='openpyxl') as writer:
            df_items.to_excel(writer,     sheet_name="items",       index=False)
            df_shelves.to_excel(writer,   sheet_name="shelves",     index=False)
            df_pullouts.to_excel(writer,  sheet_name="pullouts",    index=False)
            df_items_w2.to_excel(writer,  sheet_name="items_w2",    index=False)
            df_shelves_w2.to_excel(writer,sheet_name="shelves_w2",  index=False)
            df_po2.to_excel(writer,       sheet_name="pullouts_w2", index=False)
            df_items_yk.to_excel(writer,  sheet_name="items_yk",    index=False)
            df_shelves_yk.to_excel(writer, sheet_name="shelves_yk",  index=False)
            df_pullouts_yk.to_excel(writer,sheet_name="pullouts_yk",index=False)
            for ws in writer.book.worksheets:
                ws.protection.sheet = True
                ws.protection.enable()
            # Re-add the log sheet as veryHidden
            log_ws = writer.book.create_sheet(LOG_SHEET)
            for row in log_data:
                log_ws.append(list(row))
            log_ws.sheet_state = "veryHidden"
            log_ws.protection.sheet = True
            log_ws.protection.enable()
    except PermissionError:
        _excel_locked_error()
        raise

def save_warehouse_1(df_items, df_shelves, df_pullouts=None):
    """Saves Warehouse 1 sheets while preserving Warehouse 2."""
    if df_pullouts is None: df_pullouts = load_pullouts()
    _write_all_sheets(df_items, df_shelves, df_pullouts,
                      load_items_w2(), load_shelves_w2(), load_pullouts_w2())

def save_warehouse_2(df_items_w2, df_shelves_w2, df_pullouts_w2=None):
    """Saves Warehouse 2 sheets while preserving Warehouse 1."""
    if df_pullouts_w2 is None: df_pullouts_w2 = load_pullouts_w2()
    _write_all_sheets(load_items(), load_shelves(), load_pullouts(),
                      df_items_w2, df_shelves_w2, df_pullouts_w2)

def load_items_yk():
    df = _load_sheet(FILE, "items_yk", initialize_file)
    for col in ["QR", "Hostname", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]:
        if col not in df.columns:
            df[col] = pd.Series(dtype=str)
    return df

def load_shelves_yk():
    df = _load_sheet(FILE, "shelves_yk", initialize_file)
    for col in ["Shelf", "Status", "Date_Full"]:
        if col not in df.columns:
            df[col] = pd.Series(dtype=str)
    return df

def load_pullouts_yk():
    df = _load_sheet(FILE, "pullouts_yk", initialize_file)
    for col in ["QR", "Hostname", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]:
        if col not in df.columns:
            df[col] = pd.Series(dtype=str)
    return df

def save_warehouse_yk(df_items_yk, df_shelves_yk=None, df_pullouts_yk=None):
    """Saves Yubikey sheets while preserving Warehouse 1 & 2."""
    if df_shelves_yk is None: df_shelves_yk = load_shelves_yk()
    if df_pullouts_yk is None: df_pullouts_yk = load_pullouts_yk()
    _write_all_sheets(load_items(), load_shelves(), load_pullouts(),
                      load_items_w2(), load_shelves_w2(), load_pullouts_w2(),
                      df_items_yk, df_shelves_yk, df_pullouts_yk)

def _unhide_file(path):
    """Remove the hidden attribute on Windows so it can be written to."""
    try:
        import ctypes
        ctypes.windll.kernel32.SetFileAttributesW(str(path), 1)  # FILE_ATTRIBUTE_NORMAL
    except Exception:
        pass

def save_log(action, details=""):
    from openpyxl import load_workbook
    # Ensure warehouse.xlsx exists before writing logs
    if not os.path.exists(FILE):
        initialize_file()
    try:
        initialize_log()
        wb = load_workbook(FILE)
        if LOG_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(LOG_SHEET)
            ws.append(["Timestamp", "User", "Action", "Details"])
        else:
            ws = wb[LOG_SHEET]
            ws.protection.sheet = False
        ws.sheet_state = "veryHidden"
        ws.append([
            _now().strftime("%Y-%m-%d %H:%M:%S"),
            current_user,
            action,
            details
        ])
        ws.protection.sheet = True
        ws.protection.enable()
        wb.save(FILE)
    except Exception as e:
        print(f"[save_log error] {e}")

# ========== QR HELPERS ==========

def qr_path_for(hostname, warehouse=1):
    if warehouse == 1:   folder = QR_FOLDER_W1
    elif warehouse == 2: folder = QR_FOLDER_W2
    else:                folder = QR_FOLDER_YK
    return os.path.join(folder, f"{hostname.replace(' ', '_')}.png")

def generate_qr(hostname, data, warehouse=1):
    if warehouse == 1:   folder = QR_FOLDER_W1
    elif warehouse == 2: folder = QR_FOLDER_W2
    else:                folder = QR_FOLDER_YK
    os.makedirs(folder, exist_ok=True)
    qr_img = qrcode.make(data)
    qr_img.save(qr_path_for(hostname, warehouse))

def delete_qr(hostname, warehouse=1):
    """Move QR PNG to the dump folder (used only when an item is permanently deleted)."""
    import shutil
    path = qr_path_for(hostname, warehouse)
    if os.path.exists(path):
        try:
            if warehouse == 1:   sub = "warehouse_1"
            elif warehouse == 2: sub = "warehouse_2"
            else:                sub = "yubikey"
            dump_qr_folder = os.path.join(DUMP_FOLDER, "qr_codes", sub)
            os.makedirs(dump_qr_folder, exist_ok=True)
            dest = os.path.join(dump_qr_folder, os.path.basename(path))
            if os.path.exists(dest):
                base, ext = os.path.splitext(os.path.basename(path))
                dest = os.path.join(dump_qr_folder,
                                    f"{base}_{_now().strftime('%Y%m%d%H%M%S')}{ext}")
            shutil.move(path, dest)
        except Exception as e:
            messagebox.showwarning("Warning", f"QR file not moved to dump: {e}")

def remove_qr(hostname, warehouse=1):
    """Directly remove a QR PNG without archiving (used when unstaging back to staging)."""
    path = qr_path_for(hostname, warehouse)
    if os.path.exists(path):
        try:
            os.remove(path)
        except Exception as e:
            messagebox.showwarning("Warning", f"QR file could not be removed: {e}")

def pull_qr_path_for(hostname, warehouse=1):
    """Return the expected pull_qrs path for a hostname/qr_label."""
    if warehouse == 1:   folder = PULL_QR_FOLDER_W1
    elif warehouse == 2: folder = PULL_QR_FOLDER_W2
    else:                folder = PULL_QR_FOLDER_YK
    return os.path.join(folder, f"{hostname.replace(' ', '_')}.png")

def pull_qr(hostname, warehouse=1):
    """Move QR PNG to the pull_qrs folder when an item is pulled out."""
    import shutil
    path = qr_path_for(hostname, warehouse)
    if os.path.exists(path):
        try:
            if warehouse == 1:   dest_folder = PULL_QR_FOLDER_W1
            elif warehouse == 2: dest_folder = PULL_QR_FOLDER_W2
            else:                dest_folder = PULL_QR_FOLDER_YK
            os.makedirs(dest_folder, exist_ok=True)
            dest = pull_qr_path_for(hostname, warehouse)
            shutil.move(path, dest)
        except Exception as e:
            messagebox.showwarning("Warning", f"QR file not moved to pull_qrs: {e}")

def delete_pull_qr(hostname, warehouse=1):
    """Delete a QR PNG from the pull_qrs folder (used when item is removed from pull history)."""
    path = pull_qr_path_for(hostname, warehouse)
    if os.path.exists(path):
        try:
            os.remove(path)
        except Exception as e:
            messagebox.showwarning("Warning", f"Pull QR file could not be removed: {e}")

# ========== MINI CALENDAR PICKER ==========

def pick_date(parent, target_var, title="Select Date"):
    """Pop up a simple month calendar; sets target_var to 'YYYY-MM-DD' on pick."""
    from calendar import monthcalendar, setfirstweekday, SUNDAY
    setfirstweekday(SUNDAY)
    cal_win = tk.Toplevel(parent)
    cal_win.title(title)
    cal_win.resizable(False, False)
    cal_win.transient(parent)
    cal_win.grab_set()

    today = _now()
    # mutable state inside closure
    state = {"year": today.year, "month": today.month}

    header = tk.Frame(cal_win, bg="#2c3e50")
    header.pack(fill="x")
    prev_btn = tk.Button(header, text="◀", bg="#2c3e50", fg="white", bd=0, font=("Helvetica", 10),
                         command=lambda: _change_month(-1))
    prev_btn.pack(side="left", padx=8, pady=4)
    month_lbl = tk.Label(header, text="", bg="#2c3e50", fg="white", font=("Helvetica", 10, "bold"), width=16)
    month_lbl.pack(side="left", expand=True)
    next_btn = tk.Button(header, text="▶", bg="#2c3e50", fg="white", bd=0, font=("Helvetica", 10),
                         command=lambda: _change_month(1))
    next_btn.pack(side="right", padx=8, pady=4)

    day_names = tk.Frame(cal_win, bg="#dce3f0")
    day_names.pack(fill="x")
    for i, d in enumerate(["Su","Mo","Tu","We","Th","Fr","Sa"]):
        tk.Label(day_names, text=d, width=4, font=("Helvetica", 8, "bold"),
                 bg="#dce3f0", fg="#2c3e50").grid(row=0, column=i, padx=2, pady=3)

    grid_frame = tk.Frame(cal_win, bg="white", padx=4, pady=4)
    grid_frame.pack()

    def _change_month(delta):
        m = state["month"] + delta
        y = state["year"]
        if m > 12: m, y = 1, y + 1
        if m < 1:  m, y = 12, y - 1
        state["month"], state["year"] = m, y
        _draw()

    def _draw():
        for w in grid_frame.winfo_children():
            w.destroy()
        y, m = state["year"], state["month"]
        month_lbl.config(text=datetime(y, m, 1).strftime("%B %Y"))
        weeks = monthcalendar(y, m)
        for r, week in enumerate(weeks):
            for c, day in enumerate(week):
                if day == 0:
                    tk.Label(grid_frame, text="", width=4, bg="white").grid(row=r, column=c, padx=1, pady=1)
                else:
                    is_today = (day == today.day and m == today.month and y == today.year)
                    btn = tk.Button(
                        grid_frame, text=str(day), width=4,
                        font=("Helvetica", 9, "bold" if is_today else "normal"),
                        bg="#2980b9" if is_today else "white",
                        fg="white" if is_today else "#2c3e50",
                        relief="flat", cursor="hand2",
                        command=lambda d=day: _select(d)
                    )
                    btn.grid(row=r, column=c, padx=1, pady=1)

    def _select(day):
        chosen = f"{state['year']}-{state['month']:02d}-{day:02d}"
        target_var.set(chosen)
        cal_win.destroy()

    # "Clear" button
    clear_frame = tk.Frame(cal_win)
    clear_frame.pack(fill="x", pady=(0, 4))
    tk.Button(clear_frame, text="Clear date", fg="gray", bd=0, font=("Helvetica", 8),
              command=lambda: [target_var.set(""), cal_win.destroy()]).pack()

    _draw()
    cal_win.update_idletasks()
    # centre over parent
    px, py = parent.winfo_rootx(), parent.winfo_rooty()
    pw, ph = parent.winfo_width(), parent.winfo_height()
    cw, ch = cal_win.winfo_reqwidth(), cal_win.winfo_reqheight()
    cal_win.geometry(f"+{px + (pw - cw)//2}+{py + (ph - ch)//2}")
    cal_win.focus_force()


def _date_picker_widget(parent, var, label_text):
    """Return a frame with a calendar-button + date label. var is a StringVar."""
    frame = tk.Frame(parent)
    tk.Label(frame, text=label_text).pack(side="left", padx=(0, 2))
    date_lbl = tk.Label(frame, textvariable=var, width=11, relief="sunken",
                        bg="white", font=("Helvetica", 9), anchor="w", padx=4)
    date_lbl.pack(side="left", padx=(0, 2))
    tk.Button(frame, text="📅", width=2,
              command=lambda: pick_date(parent.winfo_toplevel(), var)).pack(side="left")
    return frame

# ========== COLUMN SORT ==========

def attach_sort_headers(tree):
    """No-op: sorting removed, headings are display-only."""
    pass


def next_set_id():
    df = load_items_w2()
    # Also include staged sets
    existing_ids = set(df["Set ID"].dropna().tolist()) if "Set ID" in df.columns else set()
    for s in staged_sets:
        existing_ids.add(s["set_id"])
    n = 1
    while True:
        sid = f"SET-{n:03d}"
        if sid not in existing_ids:
            return sid
        n += 1

# ========== W1 STAGING ==========

def update_staged_display():
    staged_listbox.delete(0, tk.END)
    if not staged_items:
        staged_listbox.insert(tk.END, "No staged items")
        return
    for item in staged_items:
        staged_listbox.insert(tk.END, f"{item['Hostname']} → {item['Shelf']} → {item.get('Status', '')}")

def select_staged_item(event):
    global selected_staged_index
    selection = staged_listbox.curselection()
    if not selection:
        return
    index = selection[0]
    selected_staged_index = index
    item = staged_items[index]
    _fill_input_fields(item["Hostname"], item.get("Checked By", ""), item["Shelf"], item.get("Status", ""), item.get("Remarks", ""))

# ========== W1 INPUT HELPERS ==========

def _fill_input_fields(hostname="", checked_by="", shelf="", status="", remarks=""):
    hostname_entry.delete(0, tk.END);   hostname_entry.insert(0, hostname)
    checked_by_entry.delete(0, tk.END); checked_by_entry.insert(0, checked_by)
    shelf_var.set(shelf)
    remarks_var.set(status)
    remarks_text_var.set(remarks)

def _clear_input_fields():
    _fill_input_fields()

# ========== W1 CORE ==========

def remove_from_staging():
    global selected_staged_index
    sel = staged_listbox.curselection()
    if sel:
        index = sel[0]
        if index >= len(staged_items):
            messagebox.showerror("Error", "Invalid staged selection")
            selected_staged_index = None
            return
        removed = staged_items.pop(index)
        selected_staged_index = None
        _clear_input_fields()
        update_staged_display()
        messagebox.showinfo("Removed", f"'{removed['Hostname']}' removed from staging")
    else:
        if not staged_items:
            messagebox.showinfo("Info", "No staged items to clear")
            return
        if not messagebox.askyesno("Confirm", f"Clear all {len(staged_items)} staged item(s)?"):
            return
        staged_items.clear()
        selected_staged_index = None
        _clear_input_fields()
        update_staged_display()
        messagebox.showinfo("Cleared", "All staged items cleared")

def put_item():
    hostname   = hostname_entry.get().strip()
    shelf      = shelf_var.get()
    serial     = ""
    checked_by = checked_by_entry.get().strip()
    status     = remarks_var.get()
    remarks    = remarks_text_var.get().strip()

    if not hostname:
        messagebox.showerror("Error", "Please enter a Hostname"); return
    if not checked_by:
        messagebox.showerror("Error", "Please enter Checked By"); return
    if not shelf:
        messagebox.showerror("Error", "Please select a Shelf"); return
    if not status:
        messagebox.showerror("Error", "Please select a Status"); return

    df_items = load_items()
    df_shelves = load_shelves()

    if hostname in df_items["Hostname"].values:
        messagebox.showerror("Error", "Hostname already exists in warehouse"); return
    if any(item['Hostname'] == hostname for item in staged_items):
        messagebox.showerror("Error", "Hostname already staged"); return
    
    shelf_status = df_shelves[df_shelves["Shelf"] == shelf]["Status"].values
    if len(shelf_status) > 0 and shelf_status[0] == "FULL":
        messagebox.showerror("Error", "Shelf is marked FULL"); return

    staged_items.append({
        "Hostname":      hostname,
        "Checked By":    checked_by,
        "Shelf":         shelf,
        "Status":        status,
        "Remarks":       remarks,
    })
    _clear_input_fields()
    messagebox.showinfo("Staged", f"'{hostname}' added to staging queue")
    update_staged_display()

def import_excel_to_staging():
    """Import rows from an Excel file into W1 staging list."""
    filepath = filedialog.askopenfilename(
        title="Import Excel — Warehouse 1",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    if not filepath: return
    try:
        df_imp = pd.read_excel(filepath, dtype=str).fillna("")
    except Exception as e:
        messagebox.showerror("Import Error", f"Could not read file:\n{e}"); return

    required = ["Hostname", "Shelf", "Checked By", "Status"]
    missing_cols = [c for c in required if c not in df_imp.columns]
    if missing_cols:
        messagebox.showerror("Import Error",
            f"Missing required column(s): {', '.join(missing_cols)}\n\n"
            f"Required: Hostname, Shelf, Checked By, Status\n"
            f"Optional: Remarks"); return

    df_items   = load_items()
    df_shelves = load_shelves()
    valid_shelves  = df_shelves["Shelf"].tolist()
    existing_hosts = df_items["Hostname"].values.tolist()
    staged_hosts   = [i["Hostname"] for i in staged_items]

    added = skipped = 0
    skip_log = []
    for i, row in df_imp.iterrows():
        hostname   = str(row.get("Hostname",   "")).strip()
        shelf      = str(row.get("Shelf",      "")).strip()
        checked_by = str(row.get("Checked By", "")).strip()
        status     = str(row.get("Status",     "")).strip()
        remarks    = str(row.get("Remarks",    "")).strip()
        row_num    = i + 2  # Excel row number (1-indexed + header)

        if not hostname:   skip_log.append(f"Row {row_num}: Missing Hostname"); skipped += 1; continue
        if not shelf:      skip_log.append(f"Row {row_num}: Missing Shelf ({hostname})"); skipped += 1; continue
        if not checked_by: skip_log.append(f"Row {row_num}: Missing Checked By ({hostname})"); skipped += 1; continue
        if not status:     skip_log.append(f"Row {row_num}: Missing Status ({hostname})"); skipped += 1; continue
        if status not in STATUS_CHOICES:
            skip_log.append(f"Row {row_num}: Invalid Status '{status}' ({hostname})"); skipped += 1; continue
        if shelf not in valid_shelves:
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' not found ({hostname})"); skipped += 1; continue
        sh_status = df_shelves[df_shelves["Shelf"] == shelf]["Status"].values
        if len(sh_status) > 0 and sh_status[0] == "FULL":
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' is FULL ({hostname})"); skipped += 1; continue
        if hostname in existing_hosts:
            skip_log.append(f"Row {row_num}: Hostname already in warehouse ({hostname})"); skipped += 1; continue
        if hostname in staged_hosts:
            skip_log.append(f"Row {row_num}: Hostname already staged ({hostname})"); skipped += 1; continue

        staged_items.append({"Hostname": hostname, "Checked By": checked_by,
                              "Shelf": shelf, "Status": status, "Remarks": remarks})
        staged_hosts.append(hostname)
        added += 1

    update_staged_display()
    msg = f"Import complete.\n\n✅ Added: {added}\n⛔ Skipped: {skipped}"
    if skip_log: msg += "\n\nSkipped details:\n" + "\n".join(skip_log[:20])
    if len(skip_log) > 20: msg += f"\n...and {len(skip_log)-20} more."
    (messagebox.showinfo if skipped == 0 else messagebox.showwarning)("Import Result", msg)

def put_warehouse():
    if not staged_items:
        messagebox.showerror("Error", "No staged items to put"); return
    if not messagebox.askyesno("Confirm", f"Put {len(staged_items)} item(s) to warehouse?"):
        return

    try:
        df_items = load_items()
        df_shelves = load_shelves()
        for col in ["Checked By"]:
            if col not in df_items.columns:
                df_items[col] = ""

        # Re-check shelf FULL status at commit time (shelf may have changed since staging)
        for _staged_item in staged_items:
            _shelf_val = _staged_item["Shelf"]
            _s_stat = df_shelves[df_shelves["Shelf"] == _shelf_val]["Status"].values
            if len(_s_stat) > 0 and _s_stat[0] == "FULL":
                messagebox.showerror("Error",
                    f"Shelf '{_shelf_val}' is marked FULL.\n"
                    f"Item '{_staged_item['Hostname']}' cannot be placed there.\n"
                    "Edit or remove it from staging first.")
                return

        now_str = _now().strftime("%Y-%m-%d %H:%M:%S")
        for item in staged_items:
            qr_code = str(uuid.uuid4())
            generate_qr(item['Hostname'], item['Hostname'], warehouse=1)
            df_items = pd.concat([df_items, pd.DataFrame([{
                "QR": qr_code,
                "Hostname": item['Hostname'],
                "Checked By": item.get('Checked By', ''),
                "Shelf": item['Shelf'],
                "Status": item.get('Status', ''),
                "Remarks": item.get('Remarks', ''),
                "Date": now_str
            }])], ignore_index=True)

        save_warehouse_1(df_items, df_shelves)

        count = len(staged_items)
        for item in staged_items:
            save_log("PUT WAREHOUSE", f"[W1] Hostname: {item['Hostname']} | Shelf: {item['Shelf']}")

        staged_items.clear()
        messagebox.showinfo("Success", f"{count} item(s) added to Warehouse 1.\nQR codes generated.\nUse 'GENERATE FILES' to create PDF labels and export to Excel.")
        update_staged_display()
        w1_refresh_all()

    except Exception as e:
        messagebox.showerror("Save Error",
            f"Failed to save to Excel:\n{str(e)}\n\n"
            "Common causes:\n• Excel file is open → close it\n• Wrong folder")
        


def update_item():
    global selected_staged_index
    new_hostname   = hostname_entry.get().strip()
    new_serial     = ""
    new_checked_by = checked_by_entry.get().strip()
    new_shelf      = shelf_var.get()
    new_status     = remarks_var.get()
    new_remarks    = remarks_text_var.get().strip()

    if not new_hostname:
        messagebox.showerror("Error", "Hostname cannot be empty"); return
    
    if not new_checked_by:
        messagebox.showerror("Error", "Checked By cannot be empty"); return
    if not new_shelf:
        messagebox.showerror("Error", "Please select a Shelf"); return
    if not new_status:
        messagebox.showerror("Error", "Please select a Status"); return

    if selected_staged_index is not None:
        index = selected_staged_index
        if index >= len(staged_items):
            messagebox.showerror("Error", "Invalid staged selection")
            selected_staged_index = None
            return
        if any(i != index and item['Hostname'] == new_hostname for i, item in enumerate(staged_items)):
            messagebox.showerror("Error", "Hostname already exists in staging"); return
        staged_items[index].update({
            "Hostname":   new_hostname,
            "Checked By": new_checked_by,
            "Shelf":      new_shelf,
            "Status":     new_status,
            "Remarks":    new_remarks,
        })
        
        save_log("UPDATE ITEM", f"[W1] Hostname: {new_hostname} | Shelf: {new_shelf}")
        messagebox.showinfo("Updated", "Staged item updated")
        update_staged_display()
        selected_staged_index = None
        return

    messagebox.showerror(
        "Update Not Allowed",
        "Items cannot be updated directly from the warehouse.\n\n"
        "Double-click the row to move it back to staging,\n"
        "then select it in the staged list and click UPDATE.")

def delete_item():
    selected = tree_warehouse.selection()
    if not selected:
        messagebox.showerror("Error", "Select item"); return

    df_items = load_items()
    df_shelves = load_shelves()
    hostname = tree_warehouse.item(selected[0], "values")[2]

    if not messagebox.askyesno("Confirm Delete",
            f"Move '{hostname}' to the dump folder?\n"
            "It will no longer appear in the warehouse but can be recovered manually."):
        return

    # Move QR to dump
    delete_qr(hostname, warehouse=1)

    # Move the item row to dump/dumped_records.xlsx
    import shutil
    row = df_items[df_items["Hostname"] == hostname]
    if not row.empty:
        dump_file  = os.path.join(DUMP_FOLDER, "dumped_records.xlsx")
        dump_sheet = "dumped_w1"
        try:
            if os.path.exists(dump_file):
                existing = pd.read_excel(dump_file, sheet_name=None)
                df_dump  = existing.get(dump_sheet, pd.DataFrame())
            else:
                df_dump = pd.DataFrame()
            row_copy = row.copy()
            row_copy["Dumped At"] = _now().strftime("%Y-%m-%d %H:%M:%S")
            df_dump = pd.concat([df_dump, row_copy], ignore_index=True)
            with pd.ExcelWriter(dump_file, engine="openpyxl",
                                mode="a" if os.path.exists(dump_file) else "w",
                                if_sheet_exists="replace") as writer:
                df_dump.to_excel(writer, sheet_name=dump_sheet, index=False)
        except Exception as e:
            messagebox.showwarning("Dump Warning", f"Record moved but dump log failed:\n{e}")

    df_items = df_items[df_items["Hostname"] != hostname].reset_index(drop=True)
    save_warehouse_1(df_items, df_shelves)
    save_log("DELETE ITEM", f"[W1] Hostname: {hostname} | Dumped")
    messagebox.showinfo("Moved to Dump", f"'{hostname}' has been moved to the dump folder.")
    w1_refresh_all()

def pull_search_live(event=None):
    """Filter whichever W1 view is currently active based on the search box."""
    keyword = pull_item_entry.get().strip().lower()

    if tree_available.winfo_ismapped():
        # Shelf status view is active
        df = load_shelves().sort_values("Shelf")
        df_items_all = load_items()
        if keyword:
            mask = False
            for col in ["Shelf", "Status", "Date_Full"]:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        tree_available.delete(*tree_available.get_children())
        for _, row in df.iterrows():
            date_full  = row.get("Date_Full", "")
            shelf_name = row["Shelf"]
            item_count = int((df_items_all["Shelf"] == shelf_name).sum()) if "Shelf" in df_items_all.columns else 0
            tree_available.insert("", "end", values=(shelf_name, row["Status"], item_count, date_full if pd.notna(date_full) else ""))
        w1_search_label.config(text=f"{len(df)} match(es)" if keyword else "")

    elif tree_pullouts.winfo_ismapped():
        # Pull history view is active — respect all active filters
        df = load_pullouts()
        shelf_filter   = pull_shelf_var.get()
        remarks_filter = pull_remarks_var.get()
        date_from      = w1_date_from_var.get().strip()
        date_to        = w1_date_to_var.get().strip()
        if keyword:
            search_cols = ["Hostname", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_filter:   df = df[df["Shelf"] == shelf_filter]
        if remarks_filter: df = df[df["Status"] == remarks_filter]
        df = _filter_by_date(df, date_from, date_to)
        tree_pullouts.delete(*tree_pullouts.get_children())
        w1_pull_row_checks.clear()
        for _, row in df.iterrows():
            hostname = str(row.get("Hostname", ""))
            checked  = hostname in w1_pull_persistent_checks
            iid = tree_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c, "") for c in ["Hostname", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
            ))
            w1_pull_row_checks[iid] = checked
        active = bool(keyword or shelf_filter or remarks_filter or date_from or date_to)
        w1_search_label.config(text=f"{len(df)} match(es)" if active else "", fg="darkorange" if active else "blue")

    else:
        # Warehouse view (default)
        show_warehouse()
        if keyword:
            df = load_items()
            search_cols = ["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
            _populate_warehouse_tree(df)
            w1_search_label.config(text=f"{len(df)} match(es)")
        else:
            w1_search_label.config(text="")

def pull_item():
    reason = pull_reason_filter_var.get().strip()
    if not reason:
        messagebox.showerror("Error", "Please enter a pull reason"); return

    # Collect checked rows; fall back to treeview selection; then fall back to search entry
    checked = [iid for iid, state in w1_row_checks.items() if state]
    if checked:
        target_iids = checked
    elif tree_warehouse.selection():
        target_iids = list(tree_warehouse.selection())
    else:
        # Legacy: single item from search entry
        hostname_input = pull_item_entry.get().strip()
        if not hostname_input:
            messagebox.showerror("Error", "Select or check item(s) to pull, or type a hostname"); return
        df_items = load_items()
        match = df_items[df_items["Hostname"] == hostname_input]
        if match.empty:
            match = df_items[df_items["Hostname"].astype(str).str.lower().str.contains(hostname_input.lower(), na=False)]
        if match.empty:
            messagebox.showerror("Error", f"'{hostname_input}' not found in warehouse"); return
        if len(match) > 1:
            names = "\n".join(match["Hostname"].tolist())
            messagebox.showerror("Error", f"Multiple matches. Be more specific:\n{names}"); return
        target_iids = None  # signal to use match directly below
        hostname = match.iloc[0]["Hostname"]
        item_row = match.iloc[0]
        if not messagebox.askyesno("Confirm Pull Out",
                f"Pull out '{hostname}'?\nReason: {reason}"):
            return
        df_items = load_items(); df_shelves = load_shelves(); df_pullouts = load_pullouts()
        shelf = str(item_row.get("Shelf", ""))
        pull_qr(hostname, warehouse=1)
        df_items = df_items[df_items["Hostname"] != hostname].reset_index(drop=True)
        df_pullouts = pd.concat([df_pullouts, pd.DataFrame([{
            "QR":            str(item_row.get("QR", "")),
            "Hostname":      hostname,
            "Checked By":    str(item_row.get("Checked By", "")),
            "Shelf":         shelf,
            "Status":        str(item_row.get("Status", "")),
            "Remarks":       str(item_row.get("Remarks", "")),
            "Pull Reason":   reason,
            "Date":          _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_warehouse_1(df_items, df_shelves, df_pullouts)
        save_log("WAREHOUSE PULL", f"[W1] Hostname: {hostname} | Shelf: {shelf} | Reason: {reason}")
        try:
            all_reasons = sorted(load_pullouts()["Pull Reason"].dropna().unique().tolist())
            pull_reason_filter_entry["values"] = [""] + all_reasons
        except Exception:
            pass
        messagebox.showinfo("Success", f"'{hostname}' pulled out successfully")
        pull_item_entry.delete(0, tk.END)
        pull_reason_filter_var.set("")
        w1_refresh_all()
        return

    # ── Bulk pull from checked/selected rows ──────────────────
    hostnames = [tree_warehouse.item(iid, "values")[2] for iid in target_iids]
    if not messagebox.askyesno("Confirm Pull Out",
            f"Pull out {len(hostnames)} item(s)?\n" +
            "\n".join(hostnames) +
            f"\n\nReason: {reason}"):
        return

    df_items   = load_items()
    df_shelves = load_shelves()
    df_pullouts = load_pullouts()
    pulled = 0
    for hostname in hostnames:
        match = df_items[df_items["Hostname"] == hostname]
        if match.empty:
            continue
        item_row = match.iloc[0]
        shelf    = str(item_row.get("Shelf", ""))
        pull_qr(hostname, warehouse=1)
        df_items = df_items[df_items["Hostname"] != hostname].reset_index(drop=True)
        df_pullouts = pd.concat([df_pullouts, pd.DataFrame([{
            "QR":            str(item_row.get("QR", "")),
            "Hostname":      hostname,
            "Checked By":    str(item_row.get("Checked By", "")),
            "Shelf":         shelf,
            "Status":        str(item_row.get("Status", "")),
            "Remarks":       str(item_row.get("Remarks", "")),
            "Pull Reason":   reason,
            "Date":          _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_log("WAREHOUSE PULL", f"[W1] Hostname: {hostname} | Shelf: {shelf} | Reason: {reason}")
        pulled += 1

    save_warehouse_1(df_items, df_shelves, df_pullouts)
    try:
        all_reasons = sorted(load_pullouts()["Pull Reason"].dropna().unique().tolist())
        pull_reason_filter_entry["values"] = [""] + all_reasons
    except Exception:
        pass
    messagebox.showinfo("Success", f"{pulled} item(s) pulled out successfully")
    pull_item_entry.delete(0, tk.END)
    pull_reason_filter_var.set("")
    w1_refresh_all()

def undo_pull(event=None):
    # Collect checked rows; fall back to treeview selection if nothing checked
    checked = [iid for iid, state in w1_pull_row_checks.items() if state]
    if not checked:
        sel = tree_pullouts.selection()
        if sel:
            checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Warehouse", "Check at least one row in the Pull History table.")
        return

    # Single confirmation for all selected items
    preview = []
    for item_id in checked:
        v = tree_pullouts.item(item_id, "values")
        if v:
            preview.append(f"  • {v[1]}  (Shelf: {v[2]})")
    if not messagebox.askyesno("Undo Pull",
            f"Restore {len(preview)} item(s) back to Warehouse 1?\n\n" + "\n".join(preview)):
        return

    # ── Pre-validate: block entire restore if any target shelf is FULL ──
    _df_sh_chk = load_shelves()
    for item_id in checked:
        _v = tree_pullouts.item(item_id, "values")
        if not _v:
            continue
        _shelf = _v[2]
        _s_stat = _df_sh_chk[_df_sh_chk["Shelf"] == _shelf]["Status"].values
        if len(_s_stat) > 0 and _s_stat[0] == "FULL":
            messagebox.showerror("Shelf Full",
                f"Cannot restore — shelf '{_shelf}' is marked FULL.\n"
                "Set it to AVAILABLE first, then retry.")
            return

    restored = 0
    for item_id in checked:
        values = tree_pullouts.item(item_id, "values")
        if not values:
            continue
        # values: ☐(0), Hostname(1), Shelf(2), Status(3), Remarks(4), PullReason(5), Date(6)
        hostname, shelf, status, remarks = values[1], values[2], values[3], values[4]

        df_items = load_items()
        df_shelves = load_shelves()
        df_pullouts = load_pullouts()

        if hostname in df_items["Hostname"].values:
            messagebox.showerror("Error", f"'{hostname}' already exists in warehouse")
            continue

        match = df_pullouts[df_pullouts["Hostname"] == hostname]
        if match.empty:
            messagebox.showerror("Error", f"'{hostname}' not found in pull history")
            continue

        pull_row = match.iloc[0]

        # ── Restore QR: move back from pull_qrs/ instead of regenerating ──
        import shutil
        pull_qr_file = pull_qr_path_for(hostname, warehouse=1)
        wh_qr_file   = qr_path_for(hostname, warehouse=1)
        qr_code = str(pull_row.get("QR", ""))
        if os.path.exists(pull_qr_file):
            try:
                os.makedirs(QR_FOLDER_W1, exist_ok=True)
                shutil.move(pull_qr_file, wh_qr_file)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR file could not be moved back: {e}")
        elif not os.path.exists(wh_qr_file):
            # Fallback: regenerate only if truly missing from both locations
            try:
                qr_code = str(uuid.uuid4())
                generate_qr(hostname, qr_code, warehouse=1)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR code not regenerated: {e}")

        for col in ["Checked By"]:
            if col not in df_items.columns:
                df_items[col] = ""

        df_items = pd.concat([df_items, pd.DataFrame([{
            "QR": qr_code,
            "Hostname": hostname,
            "Checked By": str(pull_row.get("Checked By", "")),
            "Shelf": shelf,
            "Status": status,
            "Remarks": remarks,
            "Date": _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)

        df_pullouts = df_pullouts[df_pullouts["Hostname"] != hostname].reset_index(drop=True)
        save_warehouse_1(df_items, df_shelves, df_pullouts)
        save_log("UNDO PULL", f"[W1] Hostname: {hostname} | Shelf: {shelf}")
        restored += 1

    if restored:
        messagebox.showinfo("Restored", f"{restored} item(s) restored to Warehouse 1.")
    show_pullouts()

def unstage_from_warehouse(event=None):
    # Collect checked rows; fall back to treeview selection if nothing checked
    checked = [iid for iid, state in w1_row_checks.items() if state]
    if not checked:
        sel = tree_warehouse.selection()
        if sel:
            checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Stage", "Check at least one row in the Warehouse table.")
        return

    # Single confirmation for all selected items
    preview = []
    for item_id in checked:
        v = tree_warehouse.item(item_id, "values")
        if v:
            preview.append(f"  • {v[2]}  (Shelf: {v[4]})")
    if not messagebox.askyesno("Move to Staging",
            f"Move {len(preview)} item(s) back to staging?\n\n" + "\n".join(preview)):
        return

    moved = 0
    for item_id in checked:
        values = tree_warehouse.item(item_id, "values")
        if not values:
            continue
        # values: ☐(0), QR(1), Hostname(2), Checked By(3), Shelf(4), Status(5), Remarks(6), Date(7)
        hostname, checked_by, shelf, status, remarks = values[2], values[3], values[4], values[5], values[6]
        if any(item['Hostname'] == hostname for item in staged_items):
            messagebox.showerror("Error", f"'{hostname}' is already in staging")
            continue

        df_items = load_items()
        df_shelves = load_shelves()
        remove_qr(hostname, warehouse=1)
        df_items = df_items[df_items["Hostname"] != hostname].reset_index(drop=True)
        save_warehouse_1(df_items, df_shelves)
        staged_items.append({"Hostname": hostname, "Checked By": checked_by, "Shelf": shelf, "Status": status, "Remarks": remarks})
        save_log("UNSTAGE", f"[W1] Hostname: {hostname} | Shelf: {shelf}")
        moved += 1

    if moved:
        messagebox.showinfo("Moved", f"{moved} item(s) moved back to staging.")
        update_staged_display()
        w1_refresh_all()

# ========== W1 SHELF MANAGEMENT ==========

def add_shelf():
    new_shelf = remove_shelf_var.get().strip()
    if not new_shelf:
        messagebox.showerror("Error", "Enter shelf name"); return
    df_shelves = load_shelves()
    if new_shelf in df_shelves["Shelf"].values:
        messagebox.showerror("Error", "Shelf already exists"); return
    df_shelves = pd.concat([df_shelves, pd.DataFrame([{"Shelf": new_shelf, "Status": "AVAILABLE"}])], ignore_index=True)
    df_shelves = df_shelves.sort_values("Shelf", ignore_index=True)
    save_warehouse_1(load_items(), df_shelves)
    save_log("ADD SHELF", f"[W1] Shelf: {new_shelf}")
    messagebox.showinfo("Success", f"Shelf '{new_shelf}' added")
    remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def remove_shelf():
    shelf_name = remove_shelf_var.get().strip()
    if not shelf_name:
        messagebox.showerror("Error", "Select a shelf to remove"); return
    df_items = load_items()
    df_shelves = load_shelves()
    items_in_shelf = df_items[df_items["Shelf"] == shelf_name]
    if not items_in_shelf.empty:
        messagebox.showerror("Error", f"Cannot remove shelf '{shelf_name}' - it has {len(items_in_shelf)} item(s)"); return
    if shelf_name not in df_shelves["Shelf"].values:
        messagebox.showerror("Error", f"Shelf '{shelf_name}' does not exist"); return
    df_shelves = df_shelves[df_shelves["Shelf"] != shelf_name].sort_values("Shelf", ignore_index=True)
    save_warehouse_1(df_items, df_shelves)
    save_log("REMOVE SHELF", f"[W1] Shelf: {shelf_name}")
    messagebox.showinfo("Success", f"Shelf '{shelf_name}' removed")
    remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def set_shelf_status(new_status):
    shelf = shelf_control_var.get()
    if not shelf:
        messagebox.showerror("Error", "Select a shelf from Shelf Control"); return
    df_items = load_items()
    df_shelves = load_shelves()
    idx = df_shelves[df_shelves["Shelf"] == shelf].index
    if len(idx) == 0:
        return
    df_shelves.at[idx[0], "Status"] = new_status
    df_shelves.at[idx[0], "Date_Full"] = _now().strftime("%Y-%m-%d %H:%M:%S") if new_status == "FULL" else None
    save_warehouse_1(df_items, df_shelves)
    save_log("SHELF STATUS", f"[W1] Shelf: {shelf} → {new_status}")
    w1_status_label.config(text=f"{shelf} → {new_status}")
    w1_refresh_all()

# ========== SHARED HELPERS ==========

def _filter_by_date(df, date_from, date_to, col="Date"):
    """Filter a DataFrame by date range. Both args are optional strings 'YYYY-MM-DD'.
    col: the column to filter on (default 'Date'; use 'Timestamp' for activity log)."""
    if not date_from and not date_to:
        return df
    try:
        df = df.copy()
        df[col] = pd.to_datetime(df[col], errors="coerce")
        if date_from:
            df = df[df[col] >= pd.to_datetime(date_from)]
        if date_to:
            df = df[df[col] <= pd.to_datetime(date_to) + pd.Timedelta(days=1)]
    except Exception:
        pass
    return df

# ========== W1 DISPLAY ==========

def _show_tree(tree):
    for t in (tree_warehouse, tree_available, tree_pullouts, tree_qr):
        if t is not tree:
            t.pack_forget()
    tree.pack(fill="both", expand=True)

def _open_qr_gallery(warehouse, filter_keys=None):
    """Shared QR gallery window for both warehouses.
    filter_keys: if provided, only show items whose QR key is in this list.
    """
    from PIL import Image, ImageTk
    wh_label = f"Warehouse {warehouse}"
    bg_color = "#2c3e50" if warehouse == 1 else "#1a5276"
    btn_color = "#1a252f" if warehouse == 1 else "#154360"

    qr_win = tk.Toplevel(root)
    qr_win.title(f"Stored QR Codes — {wh_label}"
                 + (f"  [{len(filter_keys)} selected]" if filter_keys else ""))
    qr_win.geometry("860x560")

    toolbar = tk.Frame(qr_win, bg=bg_color)
    toolbar.pack(fill="x")
    tk.Label(toolbar, text=f"Stored QR Codes — {wh_label}",
             bg=bg_color, fg="white", font=("Helvetica", 10, "bold")).pack(side="left", padx=10, pady=6)
    search_var = tk.StringVar()
    tk.Label(toolbar, text="Search:", bg=bg_color, fg="white").pack(side="left", padx=(20, 2))
    tk.Entry(toolbar, textvariable=search_var, width=18).pack(side="left", pady=4)
    count_lbl = tk.Label(toolbar, text="", bg=bg_color, fg="#aed6f1")
    count_lbl.pack(side="left", padx=10)
    tk.Button(toolbar, text="↻", command=lambda: _load_gallery(search_var.get()),
              bg=btn_color, fg="white", relief="flat", padx=8).pack(side="right", padx=8, pady=4)

    container = tk.Frame(qr_win)
    container.pack(fill="both", expand=True)
    canvas_qr = tk.Canvas(container, bg="#f4f6f7", highlightthickness=0)
    scrollbar_qr = ttk.Scrollbar(container, orient="vertical", command=canvas_qr.yview)
    canvas_qr.configure(yscrollcommand=scrollbar_qr.set)
    scrollbar_qr.pack(side="right", fill="y")
    canvas_qr.pack(side="left", fill="both", expand=True)
    canvas_qr.bind("<MouseWheel>", lambda e: canvas_qr.yview_scroll(int(-1*(e.delta/120)), "units"))

    inner = tk.Frame(canvas_qr, bg="#f4f6f7")
    canvas_window_id = canvas_qr.create_window((0, 0), window=inner, anchor="nw")
    _img_refs = []

    def _load_gallery(keyword=""):
        for w in inner.winfo_children():
            w.destroy()
        _img_refs.clear()
        COLS, THUMB, PAD = 4, 120, 14
        row_f = col_f = shown = 0
        df = load_items() if warehouse == 1 else load_items_w2()

        for _, row in df.iterrows():
            if warehouse == 1:
                key   = str(row.get("Hostname", ""))
                shelf = str(row.get("Shelf", ""))
                path  = qr_path_for(key, warehouse=1)
                kw_fields = [key, shelf]
                cell_labels = [(key, ("Helvetica", 8, "bold"), "#2c3e50", 0),
                               (f"Shelf: {shelf}", ("Helvetica", 7), "gray", 0)]
            else:
                set_id  = str(row.get("Set ID", ""))
                eq_type = str(row.get("Equipment Type", ""))
                shelf   = str(row.get("Shelf", ""))
                key     = f"{set_id}-{eq_type}"
                sub     = str(row.get("Serial Number", ""))
                host    = str(row.get("Hostname", ""))
                path    = qr_path_for(key, warehouse=2)
                kw_fields = [set_id, eq_type, shelf]
                cell_labels = [(key, ("Helvetica", 8, "bold"), "#2c3e50", 0),
                               (host, ("Helvetica", 7, "italic"), "#2c3e50", 0),
                               (f"S/N: {sub}", ("Helvetica", 7), "#555", 0)]

            # Filter by selection keys (from Stored QR button) first,
            # then by the gallery's own search box
            if filter_keys is not None and key not in filter_keys:
                continue
            if keyword and not any(keyword.lower() in f.lower() for f in kw_fields):
                continue

            cell = tk.Frame(inner, bg="white", bd=1, relief="solid", padx=PAD, pady=PAD)
            cell.grid(row=row_f, column=col_f, padx=8, pady=8, sticky="n")
            if os.path.exists(path):
                try:
                    img = Image.open(path).resize((THUMB, THUMB), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    _img_refs.append(photo)
                    tk.Label(cell, image=photo, bg="white").pack()
                except Exception:
                    tk.Label(cell, text="[Error]", bg="white", fg="red", width=14).pack()
            else:
                tk.Label(cell, text="[No QR File]", bg="#fdf2f2", fg="#c0392b",
                         width=14, height=6, font=("Helvetica", 8)).pack()
            for text, font, fg, _ in cell_labels:
                tk.Label(cell, text=text, bg="white", font=font, fg=fg, wraplength=130).pack(pady=(4,0) if _ == 0 else 0)

            col_f += 1; shown += 1
            if col_f >= COLS:
                col_f = 0; row_f += 1

        if shown == 0:
            tk.Label(inner, text="No QR codes found.", bg="#f4f6f7",
                     font=("Helvetica", 10), fg="gray").grid(row=0, column=0, padx=20, pady=40)
        count_lbl.config(text=f"{shown} QR code(s)")
        inner.update_idletasks()
        canvas_qr.configure(scrollregion=canvas_qr.bbox("all"))
        canvas_qr.itemconfigure(canvas_window_id, width=canvas_qr.winfo_width())

    canvas_qr.bind("<Configure>", lambda e: canvas_qr.itemconfigure(canvas_window_id, width=e.width))
    search_var.trace_add("write", lambda *_: _load_gallery(search_var.get()))
    _load_gallery()

def show_qr_codes():    _open_qr_gallery(warehouse=1)
def w2_show_qr_codes(): _open_qr_gallery(warehouse=2)

# ── Checkbox helpers ──────────────────────────────────────────

def _get_w1_selected_rows():
    """Return list of row-value tuples for checked (or all) W1 warehouse rows.
    w1_row_checks maps iid -> bool (plain bool, not BooleanVar)."""
    checked = [iid for iid, state in w1_row_checks.items() if state]
    if checked:
        return [tree_warehouse.item(iid, "values") for iid in checked]
    # Fall back to all visible rows when nothing is explicitly checked
    return [tree_warehouse.item(iid, "values") for iid in tree_warehouse.get_children()]

def _get_w2_selected_rows():
    """Return list of row-value tuples for checked (or all) W2 warehouse rows.
    w2_row_checks maps iid -> bool (plain bool, not BooleanVar)."""
    checked = [iid for iid, state in w2_row_checks.items() if state]
    if checked:
        return [tree_w2_warehouse.item(iid, "values") for iid in checked]
    return [tree_w2_warehouse.item(iid, "values") for iid in tree_w2_warehouse.get_children()]

def generate_stored_qr(warehouse=1):
    """Generate QR PNGs for selected/filtered items, open the QR gallery filtered
    to those items, and write a 'qr_selection_w1' / 'qr_selection_w2' sheet to
    the warehouse Excel with the details of those items.

    W1 column layout: ☐(0), QR(1), Hostname(2), Serial(3), Checked By(4), Shelf(5), Status(6), Remarks(7), Date(8)
    W2 column layout: ☐(0), QR(1), Set ID(2), Hostname(3), Equip Type(4), Serial(5), Checked By(6), Shelf(7), Status(8), Remarks(9), Date(10)
    """
    if warehouse == 1:
        rows = _get_w1_selected_rows()
    else:
        rows = _get_w2_selected_rows()

    if not rows:
        messagebox.showinfo("Stored QR", "No items to generate QR codes for.")
        return

    # ── 0. Preview selected items before proceeding ───────────
    if warehouse == 1:
        preview_lines = [f"  • {v[2]}" for v in rows]
    else:
        preview_lines = [f"  • {v[3]}" for v in rows]
    preview_msg = (
        f"You are about to generate files for {len(rows)} item(s):\n\n"
        + "\n".join(preview_lines)
        + "\n\nProceed?"
    )
    if not messagebox.askyesno("Confirm — Generate Files", preview_msg):
        return

    # ── 1. Generate QR PNGs ───────────────────────────────────
    count_ok = count_skip = 0
    qr_keys = []   # track keys generated so the gallery can filter to them

    for values in rows:
        try:
            if warehouse == 1:
                hostname = str(values[2])
                generate_qr(hostname, hostname, warehouse=1)
                qr_keys.append(hostname)
            else:
                set_id   = str(values[2])
                hostname = str(values[3])
                eq_type  = str(values[4])
                qr_key   = f"{set_id}-{eq_type}"
                generate_qr(qr_key, hostname, warehouse=2)
                qr_keys.append(qr_key)
            count_ok += 1
        except Exception:
            count_skip += 1

    if count_skip:
        messagebox.showwarning(
            "Generate Files",
            f"{count_ok} QR code(s) generated.\n{count_skip} item(s) skipped due to errors."
        )

    # ── 2. Write selection sheet to Excel ─────────────────────
    sheet_name = "qr_selection_w1" if warehouse == 1 else "qr_selection_w2"
    try:
        if warehouse == 1:
            cols = ["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]
            # values indices:  1       2          3       4       5        6        7
            records = [
                {c: v for c, v in zip(cols, [values[1], values[2], values[3],
                                              values[4], values[5], values[6], values[7]])}
                for values in rows
            ]
        else:
            cols = ["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number",
                    "Checked By", "Shelf", "Status", "Remarks", "Date"]
            # values indices: 1     2        3          4                 5
            #                 6          7       8         9          10
            records = [
                {c: v for c, v in zip(cols, [values[1], values[2], values[3], values[4],
                                              values[5], values[6], values[7], values[8],
                                              values[9], values[10]])}
                for values in rows
            ]

        df_sel = pd.DataFrame(records, columns=cols)

        # Append / replace the selection sheet without touching other sheets
        try:
            from openpyxl import load_workbook
            wb = load_workbook(FILE)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            # Write header
            ws.append(cols)
            for rec in df_sel.itertuples(index=False):
                ws.append(list(rec))
            ws.protection.sheet = True
            ws.protection.enable()
            wb.save(FILE)
        except PermissionError:
            _excel_locked_error()

    except Exception as e:
        messagebox.showwarning("Stored QR", f"QR codes generated but Excel sheet could not be saved:\n{e}")

    # ── 3. Prompt user for PDF + Excel export ────────────────
    # ── 3. Prompt user for PDF + Excel export ────────────────
    def _do_generate_files():
        # ── Scan existing PDFs for this warehouse ─────────────
        pdf_folder = QR_LABELS_FOLDER_W1 if warehouse == 1 else QR_LABELS_FOLDER_W2
        existing_pdfs = []
        if os.path.exists(pdf_folder):
            existing_pdfs = sorted(
                [os.path.splitext(f)[0] for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")],
                reverse=True
            )

        # ── Scan existing Excel files in warehouse-specific export folder ──
        excel_folder = EXCEL_FOLDER_W1 if warehouse == 1 else EXCEL_FOLDER_W2
        os.makedirs(excel_folder, exist_ok=True)
        existing_excels = sorted(
            [os.path.splitext(f)[0] for f in os.listdir(excel_folder)
             if f.lower().endswith(".xlsx")],
            reverse=True
        )

        # ── Build dialog ───────────────────────────────────────
        name_win = tk.Toplevel(root)
        name_win.title("Generate Files — Name Your Export")
        name_win.resizable(False, False)
        name_win.transient(root)
        name_win.grab_set()

        tk.Label(name_win, text="Generate PDF Labels & Excel Export",
                 font=("Helvetica", 10, "bold"), bg="#6c3483", fg="white",
                 padx=10, pady=6).pack(fill="x")

        form = tk.Frame(name_win, padx=16, pady=12)
        form.pack()

        # ── Section header helper ──────────────────────────────
        def _section(parent, text, row):
            tk.Label(parent, text=text, font=("Helvetica", 8, "bold"), fg="#6c3483",
                     anchor="w").grid(row=row, column=0, columnspan=3, sticky="w", pady=(10, 2))

        # ── File type checkboxes ───────────────────────────────
        tk.Label(form, text="Generate:", anchor="w", width=14).grid(row=0, column=0, sticky="w", pady=(0, 4))
        gen_pdf_var   = tk.BooleanVar(value=True)
        gen_excel_var = tk.BooleanVar(value=True)
        chk_frame = tk.Frame(form)
        chk_frame.grid(row=0, column=1, columnspan=2, sticky="w", pady=(0, 4))
        pdf_chk   = tk.Checkbutton(chk_frame, text="PDF Labels",   variable=gen_pdf_var)
        excel_chk = tk.Checkbutton(chk_frame, text="Excel Export", variable=gen_excel_var)
        pdf_chk.pack(side="left", padx=(0, 12))
        excel_chk.pack(side="left")

        # ── PDF row ────────────────────────────────────────────
        _section(form, "▸ PDF Label File", 1)
        pdf_name_lbl = tk.Label(form, text="File Name:", anchor="w", width=14)
        pdf_name_lbl.grid(row=2, column=0, sticky="w", pady=3)
        pdf_name_var = tk.StringVar(value="")
        pdf_name_cb  = ttk.Combobox(form, textvariable=pdf_name_var, width=28,
                                     values=existing_pdfs)
        pdf_name_cb.grid(row=2, column=1, pady=3, padx=(4, 0))
        tk.Label(form, text=".pdf", fg="gray", font=("Helvetica", 8)).grid(row=2, column=2, sticky="w", padx=(3, 0))
        pdf_hint = tk.Label(form, text="  ↳ Select existing PDF to append pages to, or type a new name",
                 fg="gray", font=("Helvetica", 7))
        pdf_hint.grid(row=3, column=1, columnspan=2, sticky="w")

        # ── Excel file row ─────────────────────────────────────
        _section(form, "▸ Excel File", 4)
        excel_name_lbl = tk.Label(form, text="File Name:", anchor="w", width=14)
        excel_name_lbl.grid(row=5, column=0, sticky="w", pady=3)
        file_name_var = tk.StringVar(value="")
        file_name_cb  = ttk.Combobox(form, textvariable=file_name_var, width=28,
                                      values=existing_excels)
        file_name_cb.grid(row=5, column=1, pady=3, padx=(4, 0))
        tk.Label(form, text=".xlsx", fg="gray", font=("Helvetica", 8)).grid(row=5, column=2, sticky="w", padx=(3, 0))
        excel_hint = tk.Label(form, text="  ↳ Select existing Excel to append a sheet to, or type a new name",
                 fg="gray", font=("Helvetica", 7))
        excel_hint.grid(row=6, column=1, columnspan=2, sticky="w")

        # ── Sheet name row — dynamically lists sheets when an existing Excel is chosen ──
        _section(form, "▸ Excel Sheet", 7)
        sheet_name_lbl = tk.Label(form, text="Sheet Name:", anchor="w", width=14)
        sheet_name_lbl.grid(row=8, column=0, sticky="w", pady=3)
        sheet_name_var = tk.StringVar(value="")
        sheet_name_cb  = ttk.Combobox(form, textvariable=sheet_name_var, width=28)
        sheet_name_cb.grid(row=8, column=1, pady=3, padx=(4, 0))
        sheet_hint = tk.Label(form, text="  ↳ New sheet name to add (existing sheet of same name will be replaced)",
                 fg="gray", font=("Helvetica", 7))
        sheet_hint.grid(row=9, column=1, columnspan=2, sticky="w")

        def _refresh_sheet_list(*_):
            """Populate sheet dropdown whenever the chosen Excel file changes."""
            chosen = file_name_var.get().strip()
            if not chosen:
                sheet_name_cb["values"] = []
                return
            safe = chosen if chosen.lower().endswith(".xlsx") else chosen + ".xlsx"
            xl_peek_folder = EXCEL_FOLDER_W1 if warehouse == 1 else EXCEL_FOLDER_W2
            path = os.path.join(xl_peek_folder, safe.replace("/", "-").replace("\\", "-"))
            if os.path.exists(path):
                try:
                    from openpyxl import load_workbook
                    wb_peek = load_workbook(path, read_only=True)
                    sheet_name_cb["values"] = wb_peek.sheetnames
                    wb_peek.close()
                    return
                except Exception:
                    pass
            sheet_name_cb["values"] = []

        file_name_cb.bind("<<ComboboxSelected>>", _refresh_sheet_list)
        file_name_var.trace_add("write", _refresh_sheet_list)

        def _toggle_pdf_widgets(*_):
            state = "normal" if gen_pdf_var.get() else "disabled"
            pdf_name_cb.config(state=state)
            pdf_name_lbl.config(fg="black" if gen_pdf_var.get() else "gray")
            pdf_hint.config(fg="gray")

        def _toggle_excel_widgets(*_):
            state = "normal" if gen_excel_var.get() else "disabled"
            file_name_cb.config(state=state)
            sheet_name_cb.config(state=state)
            excel_name_lbl.config(fg="black" if gen_excel_var.get() else "gray")
            sheet_name_lbl.config(fg="black" if gen_excel_var.get() else "gray")
            excel_hint.config(fg="gray")
            sheet_hint.config(fg="gray")

        gen_pdf_var.trace_add("write", _toggle_pdf_widgets)
        gen_excel_var.trace_add("write", _toggle_excel_widgets)

        _wh_label = "Warehouse 1" if warehouse == 1 else "Warehouse 2"
        tk.Label(form, text=f"(Excel files saved to excel_exports/{_wh_label.lower().replace(' ', '_')}/)",
                 fg="gray", font=("Helvetica", 8)).grid(row=10, column=0, columnspan=3, sticky="w", pady=(8, 0))

        error_lbl = tk.Label(form, text="", fg="red", font=("Helvetica", 8))
        error_lbl.grid(row=11, column=0, columnspan=3, sticky="w", pady=(4, 0))

        confirmed = [False]

        def on_confirm():
            want_pdf   = gen_pdf_var.get()
            want_excel = gen_excel_var.get()
            if not want_pdf and not want_excel:
                error_lbl.config(text="Please select at least one file type to generate."); return
            if want_pdf and not pdf_name_var.get().strip():
                error_lbl.config(text="Please enter a PDF file name."); return
            if want_excel and not file_name_var.get().strip():
                error_lbl.config(text="Please enter an Excel file name."); return
            if want_excel and not sheet_name_var.get().strip():
                error_lbl.config(text="Please enter a sheet name."); return
            confirmed[0] = True
            name_win.destroy()

        btn_row = tk.Frame(name_win, pady=10)
        btn_row.pack()
        tk.Button(btn_row, text="GENERATE", command=on_confirm,
                  bg="#6c3483", fg="white", width=12).pack(side="left", padx=6)
        tk.Button(btn_row, text="Cancel", command=name_win.destroy,
                  width=10).pack(side="left", padx=6)

        name_win.update_idletasks()
        px, py = root.winfo_rootx(), root.winfo_rooty()
        pw, ph = root.winfo_width(), root.winfo_height()
        nw, nh = name_win.winfo_reqwidth(), name_win.winfo_reqheight()
        name_win.geometry(f"+{px+(pw-nw)//2}+{py+(ph-nh)//2}")
        name_win.focus_force()
        root.wait_window(name_win)

        if not confirmed[0]:
            return
        save_log("GENERATE FILES", f"[W{'1' if warehouse == 1 else '2'}] {len(rows)} item(s)")
        pdf_name_str   = pdf_name_var.get().strip()
        file_name_str  = file_name_var.get().strip()
        sheet_name_str = sheet_name_var.get().strip()

        # ── Check for already-generated items ─────────────────
        already_in_pdf   = []
        already_in_excel = []

        # Check PDF sidecar
        import json
        pdf_folder  = QR_LABELS_FOLDER_W1 if warehouse == 1 else QR_LABELS_FOLDER_W2
        safe_pdf    = pdf_name_str.replace(" ", "_").replace("/", "-").replace("\\", "-")
        if not safe_pdf.lower().endswith(".pdf"):
            safe_pdf += ".pdf"
        sidecar_path = os.path.join(pdf_folder, safe_pdf + ".keys.json")
        if os.path.exists(sidecar_path):
            try:
                with open(sidecar_path, "r") as kf:
                    existing_keys = set(json.load(kf).get("keys", []))
                for values in rows:
                    if warehouse == 1:
                        key = str(values[2])
                        label = f"  • {values[2]}"
                    else:
                        key   = f"{values[2]}-{values[4]}"
                        label = f"  • {values[2]} — {values[4]}  (Serial: {values[5]})"
                    if key in existing_keys:
                        already_in_pdf.append(label)
            except Exception:
                pass

        # Check Excel sheet
        safe_xl = file_name_str.replace("/", "-").replace("\\", "-")
        if not safe_xl.lower().endswith(".xlsx"):
            safe_xl += ".xlsx"
        xl_check_folder = EXCEL_FOLDER_W1 if warehouse == 1 else EXCEL_FOLDER_W2
        excel_check_path = os.path.join(xl_check_folder, safe_xl)
        if os.path.exists(excel_check_path):
            try:
                from openpyxl import load_workbook
                wb_chk = load_workbook(excel_check_path, data_only=True)
                for ws_p in wb_chk.worksheets:
                    ws_p.protection.sheet = False
                if sheet_name_str in wb_chk.sheetnames:
                    ws_chk = wb_chk[sheet_name_str]
                    existing_xl_keys = set()
                    for xl_row in ws_chk.iter_rows(min_row=2, values_only=True):
                        if xl_row and xl_row[0] is not None:
                            if warehouse == 1:
                                existing_xl_keys.add(str(xl_row[0]))
                            else:
                                existing_xl_keys.add((str(xl_row[1]), str(xl_row[3]), str(xl_row[4])))
                    for values in rows:
                        if warehouse == 1:
                            key   = str(values[2])
                            label = f"  • {values[2]}"
                        else:
                            key   = (str(values[2]), str(values[4]), str(values[5]))
                            label = f"  • {values[2]} — {values[4]}  (Serial: {values[5]})"
                        if key in existing_xl_keys:
                            already_in_excel.append(label)
                wb_chk.close()
            except Exception:
                pass

        if already_in_pdf or already_in_excel:
            parts = []
            if already_in_pdf:
                parts.append(
                    f"Already in PDF '{pdf_name_str}'  ({len(already_in_pdf)} item(s)):\n"
                    + "\n".join(already_in_pdf)
                )
            if already_in_excel:
                parts.append(
                    f"Already in Excel '{file_name_str}' / sheet '{sheet_name_str}'  ({len(already_in_excel)} item(s)):\n"
                    + "\n".join(already_in_excel)
                )
            msg = (
                "Some selected items were already generated before:\n\n"
                + "\n\n".join(parts)
                + "\n\nDo you want to continue? (duplicates will be skipped)"
            )
            if not messagebox.askyesno("Already Generated", msg):
                return

        # --- Generate PDF ---
        pdf_msg = ""
        if gen_pdf_var.get():
            try:
                if warehouse == 1:
                    pdf_items = [
                        {
                            "Hostname":   str(values[2]),
                            "Checked By": str(values[3]),
                            "Shelf":      str(values[4]),
                            "Status":     str(values[5]),
                            "Remarks":    str(values[6]),
                            "_warehouse": 1,
                        }
                        for values in rows
                    ]
                else:
                    pdf_items = [
                        {
                            "Set ID":         str(values[2]),
                            "Hostname":       str(values[3]),
                            "Equipment Type": str(values[4]),
                            "Serial Number":  str(values[5]),
                            "Checked By":     str(values[6]),
                            "Shelf":          str(values[7]),
                            "Status":         str(values[8]),
                            "Remarks":        str(values[9]),
                            "_warehouse":     2,
                        }
                        for values in rows
                    ]
                _pdf_existed = os.path.exists(os.path.join(
                    QR_LABELS_FOLDER_W1 if warehouse == 1 else QR_LABELS_FOLDER_W2,
                    (pdf_name_str.replace(" ", "_") + ".pdf") if not pdf_name_str.lower().endswith(".pdf") else pdf_name_str.replace(" ", "_")
                ))
                pdf_path = generate_qr_pdf(pdf_items, custom_name=pdf_name_str)
                pdf_msg = f"PDF saved to:\n{pdf_path}"
                _wh_tag = f"[W{'1' if warehouse == 1 else '2'}]"
                save_log("FILE UPDATED" if _pdf_existed else "FILE CREATED",
                         f"{_wh_tag} PDF: {os.path.basename(pdf_path)} | Path: {pdf_path}")
            except Exception as pdf_err:
                pdf_msg = f"PDF generation failed: {pdf_err}"
        else:
            pdf_msg = "PDF skipped."

        # --- Generate Excel ---
        excel_msg = ""
        if gen_excel_var.get():
            try:
                import stat
                from openpyxl import load_workbook
                from openpyxl import Workbook as _Workbook

                safe_fname = file_name_str.replace("/", "-").replace("\\", "-")
                if not safe_fname.lower().endswith(".xlsx"):
                    safe_fname += ".xlsx"
                xl_save_folder = EXCEL_FOLDER_W1 if warehouse == 1 else EXCEL_FOLDER_W2
                os.makedirs(xl_save_folder, exist_ok=True)
                excel_path = os.path.join(xl_save_folder, safe_fname)

                if warehouse == 1:
                    cols_xl = ["Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]
                    records_xl = [
                        [values[2], values[3], values[4], values[5], values[6], values[7]]
                        for values in rows
                    ]
                else:
                    cols_xl = ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]
                    records_xl = [
                        [values[2], values[3], values[4], values[5], values[6], values[7], values[8], values[9], values[10]]
                        for values in rows
                    ]

                gen_at = _now().strftime("%Y-%m-%d %H:%M:%S")
                header = cols_xl

                # ── Lift read-only if file exists ─────────────────────
                if os.path.exists(excel_path):
                    try:
                        import ctypes
                        ctypes.windll.kernel32.SetFileAttributesW(excel_path, 0x80)
                    except Exception:
                        pass
                    os.chmod(excel_path, stat.S_IWRITE | stat.S_IREAD)

                    # Read all existing sheets into plain Python lists (read_only=True is safe)
                    wb_read = load_workbook(excel_path, read_only=True)
                    existing_sheets = {}
                    for sname in wb_read.sheetnames:
                        existing_sheets[sname] = [
                            list(row) for row in wb_read[sname].iter_rows(values_only=True)
                        ]
                    wb_read.close()

                    # Build a fresh workbook with all existing data preserved
                    wb_xl = _Workbook()
                    wb_xl.remove(wb_xl.active)  # remove blank default sheet
                    for sname, srows in existing_sheets.items():
                        ws_p = wb_xl.create_sheet(sname)
                        for row in srows:
                            ws_p.append([v if v is not None else "" for v in row])

                    # Append new rows into target sheet, skipping duplicates
                    if sheet_name_str in existing_sheets:
                        ws_xl = wb_xl[sheet_name_str]
                        existing_xl_keys = set()
                        for row in existing_sheets[sheet_name_str][1:]:  # skip header row
                            if row and len(row) > 2 and row[1] is not None:
                                if warehouse == 1:
                                    existing_xl_keys.add((str(row[1]), str(row[2])))
                                else:
                                    existing_xl_keys.add((str(row[1]), str(row[3]), str(row[4])))
                        for rec in records_xl:
                            if warehouse == 1:
                                key = (str(rec[0]), str(rec[1]))
                            else:
                                key = (str(rec[0]), str(rec[2]), str(rec[3]))
                            if key not in existing_xl_keys:
                                ws_xl.append([str(v) for v in rec])
                    else:
                        ws_xl = wb_xl.create_sheet(sheet_name_str)
                        ws_xl.append(header)
                        for rec in records_xl:
                            ws_xl.append([str(v) for v in rec])

                    for ws_p in wb_xl.worksheets:
                        ws_p.protection.sheet = True
                        ws_p.protection.enable()
                    wb_xl.save(excel_path)
                    wb_xl.close()

                else:
                    # Brand new file
                    wb_xl = _Workbook()
                    ws_xl = wb_xl.active
                    ws_xl.title = sheet_name_str
                    ws_xl.append(header)
                    for rec in records_xl:
                        ws_xl.append([str(v) for v in rec])
                    for ws_p in wb_xl.worksheets:
                        ws_p.protection.sheet = True
                        ws_p.protection.enable()
                    wb_xl.save(excel_path)
                    wb_xl.close()

                # ── Lock read-only via Windows API + chmod ────────────
                try:
                    import ctypes
                    ctypes.windll.kernel32.SetFileAttributesW(excel_path, 0x01)
                except Exception:
                    pass
                os.chmod(excel_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)

                excel_msg = f"Excel saved to:\n{excel_path}"
                _last_excel_path[warehouse] = excel_path
                _wh_tag = f"[W{'1' if warehouse == 1 else '2'}]"
                _xl_action = "FILE UPDATED" if os.path.exists(excel_path) else "FILE CREATED"
                save_log(_xl_action, f"{_wh_tag} Excel: {os.path.basename(excel_path)} | Sheet: {sheet_name_str} | Path: {excel_path}")
            except Exception as xl_err:
                excel_msg = f"Excel export failed: {xl_err}"
        else:
            excel_msg = "Excel skipped."

        messagebox.showinfo("Generate Files",
            f"{count_ok} QR code(s) processed.\n\n{pdf_msg}\n\n{excel_msg}")

    _do_generate_files()

    # ── 4. Open gallery filtered to the generated items ───────
    _open_qr_gallery(warehouse=warehouse, filter_keys=qr_keys)

def _export_pull_history(warehouse=1):
    """Export pull history rows to pull_excel/warehouse_X/.
    Dialog mirrors Generate Files: file name dropdown + sheet name, supports appending."""
    if warehouse == 1:
        tree           = tree_pullouts
        checks         = w1_pull_row_checks
        pull_xl_folder = PULL_EXCEL_FOLDER_W1
        wh_label       = "Warehouse_1"
        cols           = ["Hostname", "Shelf", "Status",
                          "Remarks", "Pull Reason", "Date"]
        col_indices    = [1, 2, 3, 4, 5, 6]
    else:
        tree           = tree_w2_pullouts
        checks         = w2_pull_row_checks
        pull_xl_folder = PULL_EXCEL_FOLDER_W2
        wh_label       = "Warehouse_2"
        cols           = ["Set ID", "Hostname", "Equipment Type", "Serial Number",
                          "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
        col_indices    = [1, 2, 3, 4, 5, 6, 7, 8, 9]

    # Respect checkboxes first, fall back to all visible rows
    checked = [iid for iid, state in checks.items() if state]
    iids    = checked if checked else list(tree.get_children())

    if not iids:
        messagebox.showinfo("Export Pull History", "No pull history rows to export.")
        return

    rows = [tree.item(iid, "values") for iid in iids]

    # ── Preview selected rows before proceeding ───────────────
    if warehouse == 1:
        preview_lines = [f"  • {v[1]}" for v in rows]
    else:
        preview_lines = [f"  • {v[2]}" for v in rows]
    preview_msg = (
        f"You are about to export {len(rows)} pull history row(s):\n\n"
        + "\n".join(preview_lines)
        + "\n\nProceed?"
    )
    if not messagebox.askyesno("Confirm — Export Pull History", preview_msg):
        return

    records   = [{c: values[i] for c, i in zip(cols, col_indices)} for values in rows]
    df_export = pd.DataFrame(records, columns=cols)
    df_export.insert(0, "Exported At", _now().strftime("%Y-%m-%d %H:%M:%S"))

    # Scan pull_excel folder for existing files to populate dropdown
    os.makedirs(pull_xl_folder, exist_ok=True)
    existing_excels = sorted(
        [f for f in os.listdir(pull_xl_folder) if f.lower().endswith(".xlsx")],
        reverse=True
    )

    export_win = tk.Toplevel(root)
    export_win.title(f"Export Pull History — {wh_label.replace('_', ' ')}")
    export_win.resizable(False, False)
    export_win.transient(root)
    export_win.grab_set()

    tk.Label(export_win,
             text=f"Export {len(rows)} Pull History Row(s) to Excel",
             font=("Helvetica", 10, "bold"), bg="#922b21", fg="white",
             padx=10, pady=6).pack(fill="x")

    form = tk.Frame(export_win, padx=16, pady=12)
    form.pack()

    def _section(text, row):
        tk.Label(form, text=text, font=("Helvetica", 8, "bold"), fg="#922b21",
                 anchor="w").grid(row=row, column=0, columnspan=3, sticky="w", pady=(10, 2))

    # ── Excel file row ─────────────────────────────────────────────
    _section("▸ Pull Excel File", 0)
    tk.Label(form, text="File Name:", anchor="w", width=14).grid(row=1, column=0, sticky="w", pady=3)
    file_name_var = tk.StringVar(value="")
    file_name_cb  = ttk.Combobox(form, textvariable=file_name_var, width=28,
                                  values=existing_excels)
    file_name_cb.grid(row=1, column=1, pady=3, padx=(4, 0))
    tk.Label(form, text=".xlsx", fg="gray", font=("Helvetica", 8)).grid(
        row=1, column=2, sticky="w", padx=(3, 0))
    tk.Label(form, text="  ↳ Select existing file to append a sheet to, or type a new name",
             fg="gray", font=("Helvetica", 7)).grid(row=2, column=1, columnspan=2, sticky="w")

    # ── Sheet name row ─────────────────────────────────────────────
    _section("▸ Excel Sheet", 3)
    tk.Label(form, text="Sheet Name:", anchor="w", width=14).grid(row=4, column=0, sticky="w", pady=3)
    sheet_name_var = tk.StringVar(value="")
    sheet_name_cb  = ttk.Combobox(form, textvariable=sheet_name_var, width=28)
    sheet_name_cb.grid(row=4, column=1, pady=3, padx=(4, 0))
    tk.Label(form, text="  ↳ New sheet name to add (existing sheet of same name will be replaced)",
             fg="gray", font=("Helvetica", 7)).grid(row=5, column=1, columnspan=2, sticky="w")

    def _refresh_sheet_list(*_):
        chosen = file_name_var.get().strip()
        if not chosen:
            sheet_name_cb["values"] = []; return
        safe = chosen if chosen.lower().endswith(".xlsx") else chosen + ".xlsx"
        path = os.path.join(pull_xl_folder, safe.replace("/", "-").replace("\\", "-"))
        if os.path.exists(path):
            try:
                from openpyxl import load_workbook
                wb_peek = load_workbook(path, read_only=True)
                sheet_name_cb["values"] = wb_peek.sheetnames
                wb_peek.close()
                return
            except Exception:
                pass
        sheet_name_cb["values"] = []

    file_name_cb.bind("<<ComboboxSelected>>", _refresh_sheet_list)
    file_name_var.trace_add("write", _refresh_sheet_list)

    _wh_label_display = "Warehouse 1" if warehouse == 1 else "Warehouse 2"
    tk.Label(form,
             text=f"(Files saved to pull_excel/{_wh_label_display.lower().replace(' ', '_')}/)",
             fg="gray", font=("Helvetica", 8)).grid(
        row=6, column=0, columnspan=3, sticky="w", pady=(8, 0))

    error_lbl = tk.Label(form, text="", fg="red", font=("Helvetica", 8))
    error_lbl.grid(row=7, column=0, columnspan=3, sticky="w", pady=(4, 0))

    confirmed = [False]

    def on_confirm():
        fn = file_name_var.get().strip()
        sn = sheet_name_var.get().strip()
        if not fn:
            error_lbl.config(text="Please enter an Excel file name."); return
        if not sn:
            error_lbl.config(text="Please enter a sheet name."); return
        confirmed[0] = True
        export_win.destroy()

    btn_row = tk.Frame(export_win, pady=10)
    btn_row.pack()
    tk.Button(btn_row, text="EXPORT", command=on_confirm,
              bg="#922b21", fg="white", width=12).pack(side="left", padx=6)
    tk.Button(btn_row, text="Cancel", command=export_win.destroy,
              width=10).pack(side="left", padx=6)

    export_win.update_idletasks()
    px, py = root.winfo_rootx(), root.winfo_rooty()
    pw, ph = root.winfo_width(), root.winfo_height()
    nw, nh = export_win.winfo_reqwidth(), export_win.winfo_reqheight()
    export_win.geometry(f"+{px+(pw-nw)//2}+{py+(ph-nh)//2}")
    export_win.focus_force()
    root.wait_window(export_win)

    if not confirmed[0]:
        return

    file_name_str  = file_name_var.get().strip()
    sheet_name_str = sheet_name_var.get().strip()

    safe_xl = file_name_str.replace("/", "-").replace("\\", "-")
    if not safe_xl.lower().endswith(".xlsx"):
        safe_xl += ".xlsx"
    out_path = os.path.join(pull_xl_folder, safe_xl)

    # ── Check for already-exported rows ───────────────────────
    already_in_excel = []
    if os.path.exists(out_path):
        try:
            from openpyxl import load_workbook
            wb_chk = load_workbook(out_path, read_only=True)
            if sheet_name_str in wb_chk.sheetnames:
                ws_chk = wb_chk[sheet_name_str]
                existing_keys = set()
                for xl_row in ws_chk.iter_rows(min_row=2, values_only=True):
                    if xl_row and xl_row[1] is not None:
                        if warehouse == 1:
                            # col layout: Exported At(0), Hostname(1), Serial(2)...
                            existing_keys.add((str(xl_row[1]), str(xl_row[2])))
                        else:
                            # col layout: Exported At(0), Set ID(1), Hostname(2), Equip Type(3), Serial(4)...
                            existing_keys.add((str(xl_row[1]), str(xl_row[3]), str(xl_row[4])))
                for values in rows:
                    if warehouse == 1:
                        key   = (str(values[1]), str(values[2]))
                        label = f"  • {values[1]}  (Serial: {values[2]})"
                    else:
                        key   = (str(values[1]), str(values[3]), str(values[4]))
                        label = f"  • {values[1]} — {values[3]}  (Serial: {values[4]})"
                    if key in existing_keys:
                        already_in_excel.append(label)
            wb_chk.close()
        except Exception:
            pass

    if already_in_excel:
        msg = (
            f"Some selected rows were already exported to '{file_name_str}' / sheet '{sheet_name_str}'  "
            f"({len(already_in_excel)} item(s)):\n\n"
            + "\n".join(already_in_excel)
            + "\n\nDo you want to continue? (duplicates will be skipped)"
        )
        if not messagebox.askyesno("Already Exported", msg):
            return

    try:
        import openpyxl
        from openpyxl import load_workbook

        # ── Build records, skipping duplicates ────────────────
        existing_keys = set()
        if os.path.exists(out_path):
            try:
                wb_read = load_workbook(out_path, read_only=True)
                if sheet_name_str in wb_read.sheetnames:
                    ws_read = wb_read[sheet_name_str]
                    for xl_row in ws_read.iter_rows(min_row=2, values_only=True):
                        if xl_row and xl_row[1] is not None:
                            if warehouse == 1:
                                existing_keys.add((str(xl_row[1]), str(xl_row[2])))
                            else:
                                existing_keys.add((str(xl_row[1]), str(xl_row[3]), str(xl_row[4])))
                wb_read.close()
            except Exception:
                pass

        new_records = []
        for rec, values in zip(records, rows):
            if warehouse == 1:
                key = (str(values[1]), str(values[2]))
            else:
                key = (str(values[1]), str(values[3]), str(values[4]))
            if key not in existing_keys:
                new_records.append(rec)

        if not new_records:
            messagebox.showinfo("Export Complete", "No new rows to export — all selected items already exist in that sheet.")
            return

        df_new = pd.DataFrame(new_records, columns=cols)
        df_new.insert(0, "Exported At", _now().strftime("%Y-%m-%d %H:%M:%S"))

        if os.path.exists(out_path):
            wb = load_workbook(out_path)
            for ws_p in wb.worksheets:
                ws_p.protection.sheet = False
            if sheet_name_str in wb.sheetnames:
                # Append new rows to existing sheet instead of replacing
                ws = wb[sheet_name_str]
                for row_data in df_new.itertuples(index=False, name=None):
                    ws.append(list(row_data))
            else:
                ws = wb.create_sheet(title=sheet_name_str)
                ws.append(list(df_new.columns))
                for row_data in df_new.itertuples(index=False, name=None):
                    ws.append(list(row_data))
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name_str
            ws.append(list(df_new.columns))
            for row_data in df_new.itertuples(index=False, name=None):
                ws.append(list(row_data))

        for ws_p in wb.worksheets:
            ws_p.protection.sheet = True
            ws_p.protection.enable()
        wb.save(out_path)
        _last_excel_path[warehouse] = out_path
        skipped = len(rows) - len(new_records)
        skip_msg = f"\n{skipped} duplicate(s) skipped." if skipped else ""
        messagebox.showinfo("Export Complete",
            f"{len(new_records)} row(s) exported to:\n{out_path}"
            f"\nSheet: '{sheet_name_str}'{skip_msg}\n\nUse VIEW EXCEL to open it.")
    except PermissionError:
        _excel_locked_error()
    except Exception as e:
        messagebox.showerror("Export Failed", f"Export failed:\n{e}")

def w1_generate_stored_qr():
    if tree_pullouts.winfo_ismapped():
        _export_pull_history(warehouse=1)
    else:
        generate_stored_qr(warehouse=1)

def w2_generate_stored_qr():
    if tree_w2_pullouts.winfo_ismapped():
        _export_pull_history(warehouse=2)
    else:
        generate_stored_qr(warehouse=2)

def view_excel(warehouse=None):
    """Open an Excel file manager dialog similar to the QR Label Manager."""
    manager = tk.Toplevel(root)
    manager.title(f"Excel File Manager — Warehouse {warehouse}" if warehouse else "Excel File Manager")
    manager.geometry("700x500")
    manager.resizable(False, False)

    # ── Header ──────────────────────────────────────────────
    hdr = tk.Frame(manager, bg="#1e8449")
    hdr.pack(fill="x")
    tk.Label(hdr, text="Generated Excel Files", font=("Helvetica", 10, "bold"),
             bg="#1e8449", fg="white").pack(side="left", padx=10, pady=6)
    sel_count_lbl = tk.Label(hdr, text="", font=("Helvetica", 9),
                              bg="#1e8449", fg="#f0b429")
    sel_count_lbl.pack(side="right", padx=10)

    # ── Checklist canvas area ────────────────────────────────
    list_frame = tk.Frame(manager, bd=1, relief="sunken")
    list_frame.pack(fill="both", expand=True, padx=10, pady=(8, 4))

    canvas_cl = tk.Canvas(list_frame, bg="white", highlightthickness=0)
    sb_cl = ttk.Scrollbar(list_frame, orient="vertical", command=canvas_cl.yview)
    canvas_cl.configure(yscrollcommand=sb_cl.set)
    sb_cl.pack(side="right", fill="y")
    canvas_cl.pack(side="left", fill="both", expand=True)
    canvas_cl.bind("<MouseWheel>", lambda e: canvas_cl.yview_scroll(int(-1*(e.delta/120)), "units"))

    inner_cl = tk.Frame(canvas_cl, bg="white")
    cw_id = canvas_cl.create_window((0, 0), window=inner_cl, anchor="nw")
    canvas_cl.bind("<Configure>", lambda e: canvas_cl.itemconfigure(cw_id, width=e.width))

    row_data   = []   # (iid, full_path, wh_label, filename, date_str, size_str, var)
    check_vars = {}   # iid -> BooleanVar

    def _refresh_sel_count():
        n = sum(1 for v in check_vars.values() if v.get())
        sel_count_lbl.config(text=f"{n} selected" if n else "")
        clear_btn.config(state="normal" if n else "disabled")

    def _toggle_all():
        want = not all(v.get() for v in check_vars.values())
        for v in check_vars.values():
            v.set(want)
        _refresh_sel_count()
        _repaint_rows()

    def _repaint_rows():
        for iid, _, _, _, _, _, var in row_data:
            try:
                fr = inner_cl.nametowidget(f"row_{iid}")
                fr.config(bg="#e8f5e9" if var.get() else ("white" if row_data.index(
                    next(r for r in row_data if r[0] == iid)) % 2 == 0 else "#f7f9fc"))
            except Exception:
                pass

    EVEN_BG, ODD_BG, SEL_BG = "white", "#f7f9fc", "#e8f5e9"

    def load_excel_files():
        for w in inner_cl.winfo_children():
            w.destroy()
        row_data.clear()
        check_vars.clear()

        now = _now()
        all_warehouses = [("Warehouse 1", 1), ("Warehouse 2", 2)]
        filtered = [(wl, wn) for wl, wn in all_warehouses if warehouse is None or wn == warehouse]

        hdr_row = tk.Frame(inner_cl, bg="#dce3f0")
        hdr_row.pack(fill="x")
        tk.Label(hdr_row, text="✔", width=3, bg="#dce3f0", font=("Helvetica", 9, "bold")).pack(side="left", padx=(6,0))
        for txt, w in [("Warehouse", 14), ("Type", 12), ("Filename", 24), ("Created", 22), ("Size", 10)]:
            tk.Label(hdr_row, text=txt, width=w, bg="#dce3f0",
                     font=("Helvetica", 9, "bold"), anchor="w").pack(side="left", padx=4, pady=5)

        idx = 0
        for warehouse_label, wh_num in filtered:
            # Scan both regular excel_exports AND pull_excel folders
            folders_to_scan = [
                (EXCEL_FOLDER_W1 if wh_num == 1 else EXCEL_FOLDER_W2, "Generated"),
                (PULL_EXCEL_FOLDER_W1 if wh_num == 1 else PULL_EXCEL_FOLDER_W2, "Pull History"),
            ]
            for xl_folder, folder_type in folders_to_scan:
                if not os.path.exists(xl_folder):
                    continue
                files = sorted(
                    [f for f in os.listdir(xl_folder) if f.lower().endswith(".xlsx")],
                    reverse=True
                )
                for f in files:
                    full_path = os.path.join(xl_folder, f)
                    size_kb = round(os.path.getsize(full_path) / 1024, 1)
                    try:
                        mtime = os.path.getmtime(full_path)
                        file_dt = datetime.fromtimestamp(mtime)
                        delta = now - file_dt
                        age = "Today" if delta.days == 0 else ("1 day ago" if delta.days == 1 else f"{delta.days} days ago")
                        date_str = f"{file_dt.strftime('%Y-%m-%d')}  ({age})"
                    except Exception:
                        date_str = "Unknown"

                    iid = f"row{idx}"
                    var = tk.BooleanVar(value=False)
                    check_vars[iid] = var
                    bg = EVEN_BG if idx % 2 == 0 else ODD_BG

                    row_fr = tk.Frame(inner_cl, bg=bg, name=f"row_{iid}", cursor="hand2")
                    row_fr.pack(fill="x")

                    def _make_toggle(v=var):
                        def _toggle(e=None):
                            v.set(not v.get())
                            _refresh_sel_count()
                            _repaint_rows()
                        return _toggle

                    cb = tk.Checkbutton(row_fr, variable=var, bg=bg,
                                        command=lambda: [_refresh_sel_count(), _repaint_rows()])
                    cb.pack(side="left", padx=(6, 0), pady=4)
                    size_str = f"{size_kb} KB" if size_kb < 1024 else f"{round(size_kb/1024, 2)} MB"
                    size_str = f"{size_kb} KB" if size_kb < 1024 else f"{round(size_kb / 1024, 2)} MB"
                    size_str = f"{size_kb} KB" if size_kb < 1024 else f"{round(size_kb / 1024, 2)} MB"
                    for txt, w in [(warehouse_label, 14), (folder_type, 12), (f, 24), (date_str, 22), (size_str, 10)]:
                        lbl = tk.Label(row_fr, text=txt, bg=bg, anchor="w", width=w,
                                       font=("Helvetica", 9))
                        lbl.pack(side="left", padx=4, pady=4)
                        lbl.bind("<Button-1>", _make_toggle(var))
                    row_fr.bind("<Button-1>", _make_toggle(var))

                    # Store full_path and folder_type so clear_selected can dump correctly
                    row_data.append((iid, full_path, warehouse_label, f, date_str, f"{size_kb} kb", var, folder_type, wh_num))
                    idx += 1

        if idx == 0:
            tk.Label(inner_cl, text="No generated Excel files found.", fg="gray",
                     font=("Helvetica", 10), bg="white").pack(pady=30)

        inner_cl.update_idletasks()
        canvas_cl.configure(scrollregion=canvas_cl.bbox("all"))
        _refresh_sel_count()

    def open_selected():
        chosen = [rd for rd in row_data if rd[6].get()]
        if not chosen:
            messagebox.showwarning("Warning", "Check a file to open.", parent=manager); return
        if len(chosen) > 1:
            messagebox.showerror("Error", "Only 1 file can be opened at a time.\nPlease check only one file.", parent=manager); return
        full_path = chosen[0][1]
        if os.path.exists(full_path):
            os.startfile(full_path)
        else:
            messagebox.showerror("Error", "File not found.", parent=manager)

    def clear_selected():
        chosen = [rd for rd in row_data if rd[6].get()]
        if not chosen:
            messagebox.showwarning("Warning", "Check at least one file to move to dump.", parent=manager); return
        count = len(chosen)
        prompt = (f"Move '{chosen[0][3]}' to dump?" if count == 1
                  else f"Move {count} checked file(s) to dump?\nThey can be recovered from the dump folder.")
        if not messagebox.askyesno("Confirm Move to Dump", prompt, parent=manager):
            return
        import shutil, stat, ctypes
        failed = []
        for rd in chosen:
            _, full_path, warehouse_label, fname, _, _, _, folder_type, wh_num = rd
            # Route to the correct dump subfolder based on file type and warehouse
            if folder_type == "Pull History":
                dump_folder = os.path.join(DUMP_FOLDER, "pull_excel",
                                           "warehouse_1" if wh_num == 1 else "warehouse_2")
            else:
                dump_folder = DUMP_EXCEL_W1 if wh_num == 1 else DUMP_EXCEL_W2
            os.makedirs(dump_folder, exist_ok=True)
            try:
                if os.path.exists(full_path):
                    try:
                        ctypes.windll.kernel32.SetFileAttributesW(full_path, 0x80)
                    except Exception:
                        pass
                    os.chmod(full_path, stat.S_IWRITE | stat.S_IREAD)
                    dest = os.path.join(dump_folder, fname)
                    if os.path.exists(dest):
                        base, ext = os.path.splitext(fname)
                        dest = os.path.join(dump_folder,
                                            f"{base}_{_now().strftime('%Y%m%d%H%M%S')}{ext}")
                    shutil.move(full_path, dest)
                    save_log("FILE DELETED", f"File: {fname} | Warehouse: {warehouse_label} | Moved to dump: {dest}")
            except Exception as e:
                failed.append(f"{fname}: {e}")
        load_excel_files()
        if failed:
            messagebox.showerror("Error", "Some files could not be moved:\n" + "\n".join(failed), parent=manager)
        else:
            messagebox.showinfo("Moved to Dump", f"{count} file(s) moved to dump folder.", parent=manager)

    # ── Bottom toolbar ───────────────────────────────────────
    btn_frame_m = tk.Frame(manager)
    btn_frame_m.pack(pady=8)
    tk.Button(btn_frame_m, text="☑", command=_toggle_all, width=4).pack(side="left", padx=4)
    tk.Button(btn_frame_m, text="OPEN", command=open_selected, width=10).pack(side="left", padx=4)
    clear_btn = tk.Button(btn_frame_m, text="✕", command=clear_selected,
                           width=10, bg="#922b21", fg="white", state="disabled")
    clear_btn.pack(side="left", padx=4)

    load_excel_files()

def w1_view_excel(): view_excel(warehouse=1)
def w2_view_excel(): view_excel(warehouse=2)

def view_stored_qr(warehouse=1):
    """Open the QR gallery.
    - Pull History table active → reads from pull_qrs/ folder directly.
    - Warehouse table active   → reads from qr_codes/ via _open_qr_gallery."""
    pull_active = (tree_pullouts.winfo_ismapped() if warehouse == 1
                   else tree_w2_pullouts.winfo_ismapped())

    if pull_active:
        # ── Pull QR gallery — reads PNGs straight from pull_qrs/ ──
        folder = PULL_QR_FOLDER_W1 if warehouse == 1 else PULL_QR_FOLDER_W2
        if not os.path.exists(folder):
            messagebox.showinfo("View QR", "No pull QR codes folder found."); return
        files = [f for f in os.listdir(folder) if f.lower().endswith(".png")]
        if not files:
            messagebox.showinfo("View QR", "No QR codes for pulled items yet."); return

        # Build filter_keys from visible pull history rows
        if warehouse == 1:
            visible_iids = list(tree_pullouts.get_children())
            # pull tree W1: CP0(0), Hostname(1), Serial(2), Shelf(3)...
            filter_keys = (
                [str(tree_pullouts.item(iid, "values")[1]) for iid in visible_iids]
                if visible_iids else None
            )
        else:
            visible_iids = list(tree_w2_pullouts.get_children())
            # pull tree W2: CP0(0), Set ID(1), Hostname(2), Equip Type(3)...
            filter_keys = (
                [f"{tree_w2_pullouts.item(iid, 'values')[1]}-{tree_w2_pullouts.item(iid, 'values')[3]}"
                 for iid in visible_iids]
                if visible_iids else None
            )

        from PIL import Image, ImageTk
        bg_color  = "#2c3e50" if warehouse == 1 else "#1a5276"
        btn_color = "#1a252f" if warehouse == 1 else "#154360"

        qr_win = tk.Toplevel(root)
        qr_win.title(f"Pull QR Codes — Warehouse {warehouse}")
        qr_win.geometry("860x560")

        toolbar = tk.Frame(qr_win, bg=bg_color)
        toolbar.pack(fill="x")
        tk.Label(toolbar, text=f"Pull QR Codes — Warehouse {warehouse}",
                 bg=bg_color, fg="white", font=("Helvetica", 10, "bold")).pack(side="left", padx=10, pady=6)
        search_var = tk.StringVar()
        tk.Label(toolbar, text="Search:", bg=bg_color, fg="white").pack(side="left", padx=(20, 2))
        tk.Entry(toolbar, textvariable=search_var, width=18).pack(side="left", pady=4)
        count_lbl = tk.Label(toolbar, text="", bg=bg_color, fg="#aed6f1")
        count_lbl.pack(side="left", padx=10)

        container = tk.Frame(qr_win)
        container.pack(fill="both", expand=True)
        canvas_qr = tk.Canvas(container, bg="#f4f6f7", highlightthickness=0)
        sb_qr = ttk.Scrollbar(container, orient="vertical", command=canvas_qr.yview)
        canvas_qr.configure(yscrollcommand=sb_qr.set)
        sb_qr.pack(side="right", fill="y")
        canvas_qr.pack(side="left", fill="both", expand=True)
        canvas_qr.bind("<MouseWheel>",
                       lambda e: canvas_qr.yview_scroll(int(-1*(e.delta/120)), "units"))

        inner = tk.Frame(canvas_qr, bg="#f4f6f7")
        cw_id = canvas_qr.create_window((0, 0), window=inner, anchor="nw")
        _img_refs = []

        def _load_pull_gallery(keyword=""):
            for w in inner.winfo_children():
                w.destroy()
            _img_refs.clear()
            COLS, THUMB, PAD = 4, 120, 14
            row_f = col_f = shown = 0

            for fname in sorted(files):
                # Restore the key the same way qr_path_for stored it
                key = os.path.splitext(fname)[0].replace("_", " ")
                if filter_keys is not None and key not in filter_keys:
                    continue
                if keyword and keyword.lower() not in key.lower():
                    continue

                path = os.path.join(folder, fname)
                cell = tk.Frame(inner, bg="white", bd=1, relief="solid", padx=PAD, pady=PAD)
                cell.grid(row=row_f, column=col_f, padx=8, pady=8, sticky="n")
                try:
                    img   = Image.open(path).resize((THUMB, THUMB), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    _img_refs.append(photo)
                    tk.Label(cell, image=photo, bg="white").pack()
                except Exception:
                    tk.Label(cell, text="[Error]", bg="white", fg="red", width=14).pack()
                tk.Label(cell, text=key, bg="white", font=("Helvetica", 8, "bold"),
                         fg="#2c3e50", wraplength=130).pack(pady=(4, 0))

                col_f += 1; shown += 1
                if col_f >= COLS:
                    col_f = 0; row_f += 1

            if shown == 0:
                tk.Label(inner, text="No QR codes found.", bg="#f4f6f7",
                         font=("Helvetica", 10), fg="gray").grid(row=0, column=0, padx=20, pady=40)
            count_lbl.config(text=f"{shown} QR code(s)")
            inner.update_idletasks()
            canvas_qr.configure(scrollregion=canvas_qr.bbox("all"))
            canvas_qr.itemconfigure(cw_id, width=canvas_qr.winfo_width())

        tk.Button(toolbar, text="↻", command=lambda: _load_pull_gallery(search_var.get()),
                  bg=btn_color, fg="white", relief="flat", padx=8).pack(side="right", padx=8, pady=4)
        canvas_qr.bind("<Configure>",
                       lambda e: canvas_qr.itemconfigure(cw_id, width=e.width))
        search_var.trace_add("write", lambda *_: _load_pull_gallery(search_var.get()))
        _load_pull_gallery()

    else:
        # ── Warehouse QR gallery — original behaviour ──────────────
        folder = QR_FOLDER_W1 if warehouse == 1 else QR_FOLDER_W2
        if not os.path.exists(folder):
            messagebox.showinfo("View QR", "No QR codes folder found."); return
        files = [f for f in os.listdir(folder) if f.lower().endswith(".png")]
        if not files:
            messagebox.showinfo("View QR", "No QR codes have been generated yet."); return
        if warehouse == 1:
            visible_iids = list(tree_warehouse.get_children())
            if visible_iids:
                filter_keys = [str(tree_warehouse.item(iid, "values")[2])
                               for iid in visible_iids]
            else:
                filter_keys = [os.path.splitext(f)[0].replace("_", " ") for f in files]
        else:
            visible_iids = list(tree_w2_warehouse.get_children())
            if visible_iids:
                filter_keys = [
                    f"{tree_w2_warehouse.item(iid, 'values')[2]}-{tree_w2_warehouse.item(iid, 'values')[4]}"
                    for iid in visible_iids
                ]
            else:
                filter_keys = [os.path.splitext(f)[0].replace("_", " ") for f in files]
        _open_qr_gallery(warehouse=warehouse, filter_keys=filter_keys)

def w1_view_stored_qr(): view_stored_qr(warehouse=1)
def w2_view_stored_qr(): view_stored_qr(warehouse=2)

def _w1_refresh_select_all_label():
    """Update the Select All / Deselect All button label for W1.
    Safe to call before UI exists — errors are silently swallowed."""
    try:
        all_iids = list(tree_warehouse.get_children())
        checked  = [iid for iid in all_iids if w1_row_checks.get(iid)]
        w1_select_all_btn.config(
            text="DESELECT ALL" if all_iids and len(checked) == len(all_iids) else "SELECT ALL")
    except (NameError, tk.TclError, AttributeError):
        pass

def _w2_refresh_select_all_label():
    """Update the Select All / Deselect All button label for W2.
    Safe to call before UI exists — errors are silently swallowed."""
    try:
        all_iids = list(tree_w2_warehouse.get_children())
        checked  = [iid for iid in all_iids if w2_row_checks.get(iid)]
        w2_select_all_btn.config(
            text="DESELECT ALL" if all_iids and len(checked) == len(all_iids) else "SELECT ALL")
    except (NameError, tk.TclError, AttributeError):
        pass


def show_warehouse():
    try: w1_back_to_wh_btn.pack_forget()
    except Exception: pass
    try: w1_back_to_stage_btn.pack(side="left", padx=(0, 6))
    except Exception: pass
    w1_update_full_shelves_display()
    try: w1_back_to_wh_btn.pack_forget()
    except Exception: pass
    df_items = load_items()
    if "Date" not in df_items.columns:
        df_items["Date"] = ""
    # Apply active date filters if set
    try:
        date_from = w1_date_from_var.get().strip()
        date_to   = w1_date_to_var.get().strip()
        df_items = _filter_by_date(df_items, date_from, date_to)
    except (NameError, Exception):
        pass  # vars not yet created during startup
    _populate_warehouse_tree(df_items)

def show_available():
    try: w1_back_to_wh_btn.pack_forget()
    except Exception: pass
    try: w1_back_to_stage_btn.pack_forget()
    except Exception: pass
    _show_tree(tree_available)
    try: w1_back_to_wh_btn.pack_forget()
    except Exception: pass
    tree_available.delete(*tree_available.get_children())
    try:
        keyword = search_entry.get().strip().lower()
    except Exception:
        keyword = ""
    df = load_shelves().sort_values("Shelf")
    df_items = load_items()
    if keyword:
        mask = False
        for col in ["Shelf", "Status", "Date_Full"]:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
        df = df[mask]
    for _, row in df.iterrows():
        date_full = row.get("Date_Full", "")
        shelf_name = row["Shelf"]
        item_count = int((df_items["Shelf"] == shelf_name).sum()) if "Shelf" in df_items.columns else 0
        tree_available.insert("", "end", values=(shelf_name, row["Status"], item_count, date_full if pd.notna(date_full) else ""))

def show_pullouts():
    try: w1_back_to_stage_btn.pack_forget()
    except Exception: pass
    try: w1_back_to_wh_btn.pack(side="left", padx=(0, 6))
    except Exception: pass
    _show_tree(tree_pullouts)
    w1_back_to_wh_btn.pack(side="left", padx=(0, 6))
    tree_pullouts.delete(*tree_pullouts.get_children())
    df_po = load_pullouts()
    # Only apply dropdown/date filters — keyword is NOT applied here so that
    # clicking PULL HISTORY always shows the full pull history, not a
    # search-box-filtered subset.
    try:
        shelf_filter   = pull_shelf_var.get()
        remarks_filter = pull_remarks_var.get()
        date_from      = w1_date_from_var.get().strip()
        date_to        = w1_date_to_var.get().strip()
    except (NameError, Exception):
        shelf_filter = remarks_filter = date_from = date_to = ""
    if shelf_filter:   df_po = df_po[df_po["Shelf"] == shelf_filter]
    if remarks_filter: df_po = df_po[df_po["Status"] == remarks_filter]
    df_po = _filter_by_date(df_po, date_from, date_to)
    w1_pull_row_checks.clear()
    for _, row in df_po.iterrows():
        hostname = str(row.get("Hostname", ""))
        checked  = hostname in w1_pull_persistent_checks
        iid = tree_pullouts.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c, "") for c in ["Hostname", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
        ))
        w1_pull_row_checks[iid] = checked
    try:
        all_reasons = sorted(load_pullouts()["Pull Reason"].dropna().unique().tolist())
        pull_reason_filter_entry["values"] = [""] + all_reasons
    except Exception:
        pass
    w1_search_label.config(text="", fg="blue")

def _populate_warehouse_tree(df):
    _show_tree(tree_warehouse)
    tree_warehouse.delete(*tree_warehouse.get_children())
    w1_row_checks.clear()
    for _, row in df.iterrows():
        hostname = str(row.get("Hostname", ""))
        checked  = hostname in w1_persistent_checks
        iid = tree_warehouse.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c, "") for c in ["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"])
        ))
        w1_row_checks[iid] = checked
    _w1_refresh_select_all_label()

def search_item():
    keyword        = search_entry.get().strip().lower()
    shelf_filter   = pull_shelf_var.get()
    remarks_filter = pull_remarks_var.get()
    date_from      = w1_date_from_var.get().strip()
    date_to        = w1_date_to_var.get().strip()

    # ── Pull history view is active: filter it instead of warehouse ──
    if tree_pullouts.winfo_ismapped():
        df = load_pullouts()
        if keyword:
            search_cols = ["Hostname", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_filter:   df = df[df["Shelf"] == shelf_filter]
        if remarks_filter: df = df[df["Status"] == remarks_filter]
        df = _filter_by_date(df, date_from, date_to)
        tree_pullouts.delete(*tree_pullouts.get_children())
        w1_pull_row_checks.clear()
        for _, row in df.iterrows():
            hostname = str(row.get("Hostname", ""))
            checked  = hostname in w1_pull_persistent_checks
            iid = tree_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c, "") for c in ["Hostname", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
            ))
            w1_pull_row_checks[iid] = checked
        parts = []
        if keyword:        parts.append(f"Search: \"{keyword}\"")
        if shelf_filter:   parts.append(f"Shelf: {shelf_filter}")
        if remarks_filter: parts.append(f"Status: {remarks_filter}")
        if date_from:      parts.append(f"From: {date_from}")
        if date_to:        parts.append(f"To: {date_to}")
        label = f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")
        w1_search_label.config(text=label if parts else "", fg="darkorange" if parts else "blue")
        return

    # ── Default: filter warehouse view ───────────────────────────────
    df = load_items()
    if keyword:
        search_cols = ["QR", "Hostname", "Checked By", "Shelf", "Status", "Remarks", "Date"]
        mask = False
        for col in search_cols:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
        df = df[mask]
    if shelf_filter:   df = df[df["Shelf"] == shelf_filter]
    if remarks_filter: df = df[df["Status"] == remarks_filter]
    df = _filter_by_date(df, date_from, date_to)
    _populate_warehouse_tree(df)
    parts = []
    if keyword:        parts.append(f"Search: \"{keyword}\"")
    if shelf_filter:   parts.append(f"Shelf: {shelf_filter}")
    if remarks_filter: parts.append(f"Status: {remarks_filter}")
    if date_from:      parts.append(f"From: {date_from}")
    if date_to:        parts.append(f"To: {date_to}")
    label = f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")
    w1_search_label.config(text=label if parts else "")
    if parts:
        save_log("SEARCH", f"[W1] {' | '.join(parts)} → {len(df)} result(s)")

def filter_pull_history():
    """Filter and display the W1 pull history table only. Completely separate from warehouse search."""
    reason = pull_reason_filter_var.get().strip().lower()
    date_from = w1_pull_date_from_var.get().strip()
    date_to = w1_pull_date_to_var.get().strip()
    show_pullouts()  # always switch to pull history view first
    if not reason and not date_from and not date_to:
        return
    tree_pullouts.delete(*tree_pullouts.get_children())
    df = load_pullouts()
    if reason:
        search_cols = ["Hostname", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
        mask = False
        for col in search_cols:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(reason, na=False)
        df = df[mask]
    df = _filter_by_date(df, date_from, date_to)
    # Refresh pull reason dropdown
    all_reasons = sorted(load_pullouts()["Pull Reason"].dropna().unique().tolist())
    pull_reason_filter_entry["values"] = [""] + all_reasons
    w1_pull_row_checks.clear()
    for _, row in df.iterrows():
        hostname = str(row.get("Hostname", ""))
        checked  = hostname in w1_pull_persistent_checks
        iid = tree_pullouts.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c, "") for c in ["Hostname", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
        ))

        w1_pull_row_checks[iid] = checked
    w1_search_label.config(text=f"Pull History Filtered: {len(df)} result(s)")

def w1_update_full_shelves_display():
    df_shelves = load_shelves()
    full_shelves = df_shelves[df_shelves["Status"] == "FULL"]["Shelf"].tolist()
    w1_full_label.config(text="FULL Shelves:\n" + "\n".join(full_shelves) if full_shelves else "FULL Shelves: None")

def w1_refresh_all():
    show_warehouse()
    update_all_shelf_dropdowns()

def update_all_shelf_dropdowns():
    # Load separate lists
    w1_shelf_list = sorted(load_shelves()["Shelf"].tolist())
    w2_shelf_list = sorted(load_shelves_w2()["Shelf"].tolist())
    try:
        yk_shelf_list = sorted(load_shelves_yk()["Shelf"].tolist())
    except Exception:
        yk_shelf_list = []

    # Update W1 specific dropdowns
    for dropdown in (shelf_dropdown, shelf_control_dropdown, remove_shelf_dropdown, pull_shelf_dropdown):
        try:
            dropdown["values"] = w1_shelf_list
        except NameError:
            pass # Ignore if UI isn't initialized yet

    # Update W2 specific dropdowns
    for dropdown in (w2_shelf_control_dropdown, w2_remove_shelf_dropdown, w2_pull_shelf_dropdown):
        try:
            dropdown["values"] = w2_shelf_list
        except NameError:
            pass

    # Update YK specific dropdowns
    for dropdown in (yk_shelf_control_dropdown, yk_remove_shelf_dropdown, yk_shelf_dropdown, yk_shelf_filter_dropdown):
        try:
            dropdown["values"] = yk_shelf_list
        except NameError:
            pass

def select_pull_item(event):
    """Toggle checkbox on click in W1 pull history table."""
    selected = tree_pullouts.selection()
    if selected:
        iid = selected[0]
        if iid in w1_pull_row_checks:
            w1_pull_row_checks[iid] = not w1_pull_row_checks[iid]
            values = tree_pullouts.item(iid, "values")
            hostname = str(values[1])  # Hostname is col index 1 in pull table
            if w1_pull_row_checks[iid]:
                w1_pull_persistent_checks.add(hostname)
            else:
                w1_pull_persistent_checks.discard(hostname)
            tree_pullouts.set(iid, "CP0", "☑" if w1_pull_row_checks[iid] else "☐")

def w2_select_pull_item(event):
    """Toggle checkbox on click in W2 pull history table."""
    selected = tree_w2_pullouts.selection()
    if selected:
        iid = selected[0]
        if iid in w2_pull_row_checks:
            w2_pull_row_checks[iid] = not w2_pull_row_checks[iid]
            values = tree_w2_pullouts.item(iid, "values")
            key = (str(values[1]), str(values[3]))  # Set ID(1), Equipment Type(3)
            if w2_pull_row_checks[iid]:
                w2_pull_persistent_checks.add(key)
            else:
                w2_pull_persistent_checks.discard(key)
            tree_w2_pullouts.set(iid, "CP0", "☑" if w2_pull_row_checks[iid] else "☐")

def select_item(event):
    selected = tree_warehouse.selection()
    if selected:
        iid = selected[0]
        values = tree_warehouse.item(iid, "values")
        # values: ☐(0), QR(1), Hostname(2), Checked By(3), Shelf(4), Status(5), Remarks(6), Date(7)

        # Toggle checkbox on single-click
        if iid in w1_row_checks:
            w1_row_checks[iid] = not w1_row_checks[iid]
            hostname = str(values[2])
            if w1_row_checks[iid]:
                w1_persistent_checks.add(hostname)
            else:
                w1_persistent_checks.discard(hostname)
            tree_warehouse.set(iid, "C0", "☑" if w1_row_checks[iid] else "☐")
            _w1_refresh_select_all_label()

        pull_item_entry.delete(0, tk.END)
        pull_item_entry.insert(0, values[2])
        w1_status_label.config(
            text=f"Selected → Hostname: {values[2]}  |  Shelf: {values[4]}",
            fg="#1a5276")

# ========== W1 RESET ==========

def reset_ui():
    _clear_input_fields()
    for s in tree_warehouse.selection(): tree_warehouse.selection_remove(s)
    w1_status_label.config(text="")
    w1_search_label.config(text="")
    w1_persistent_checks.clear()
    show_warehouse()

def reset_shelf_control():
    shelf_control_var.set("")
    w1_status_label.config(text="")

def reset_shelf_addition():
    remove_shelf_var.set("")
    w1_status_label.config(text="")

def reset_pull_out():
    pull_item_entry.delete(0, tk.END)
    pull_reason_filter_var.set("")
    w1_pull_date_from_var.set("")
    w1_pull_date_to_var.set("")
    for s in tree_warehouse.selection(): tree_warehouse.selection_remove(s)
    w1_status_label.config(text="")
    w1_search_label.config(text="")
    show_warehouse()

def clear_pull_filters():
    pull_shelf_var.set("")
    pull_remarks_var.set("")
    search_entry.delete(0, tk.END)
    pull_reason_filter_var.set("")
    w1_pull_date_from_var.set("")
    w1_pull_date_to_var.set("")
    w1_date_from_var.set("")
    w1_date_to_var.set("")
    w1_search_label.config(text="")
    w1_persistent_checks.clear()
    w1_pull_persistent_checks.clear()
    show_warehouse()

# ========== W2 STAGING ==========

def update_w2_staged_display():
    w2_staged_listbox.delete(0, tk.END)
    if not staged_sets:
        w2_staged_listbox.insert(tk.END, "No staged sets")
        return
    for s in staged_sets:
        items = s["items"]
        # Collect distinct shelves
        shelves = list(dict.fromkeys(i.get("Shelf", "") for i in items if i.get("Shelf")))
        shelf_str = shelves[0] if len(shelves) == 1 else (f"{len(shelves)} shelves" if shelves else "—")
        # Check completeness (all 4 types present)
        missing = [t for t in EQUIPMENT_TYPES if t not in [i["Equipment Type"] for i in items]]
        flag = " ⚠ missing: " + ", ".join(missing) if missing else ""
        w2_staged_listbox.insert(tk.END, f"{s['set_id']} | {len(items)} items | {shelf_str}{flag}")

def w2_build_set():
    """Open a dialog to build a new equipment set."""
    selected_types = []
    for eq_type, var in w2_equip_vars.items():
        if var.get():
            selected_types.append(eq_type)

    if not selected_types:
        messagebox.showerror("Error", "Please select at least one equipment type"); return

    set_id = next_set_id()

    build_win = tk.Toplevel(root)
    build_win.title(f"Build {set_id}")
    build_win.resizable(False, False)
    build_win.transient(root)

    tk.Label(build_win, text=f"Fill in details for {set_id}", font=("Helvetica", 10, "bold")).pack(pady=(10, 5))

    shelf_list = sorted(load_shelves_w2()["Shelf"].tolist())

    COL_WIDTHS = [12, 18, 18, 16, 16, 13, 20]
    HEADERS    = ["Equipment", "Hostname", "Serial Number", "Checked By", "Shelf", "Status", "Remarks"]

    outer = tk.Frame(build_win, padx=10, pady=5)
    outer.pack(fill="both", expand=True)

    # Fixed header
    hdr_frame = tk.Frame(outer, bg="#dce3f0")
    hdr_frame.pack(fill="x")
    for col, (h, cw) in enumerate(zip(HEADERS, COL_WIDTHS)):
        tk.Label(hdr_frame, text=h, font=("Helvetica", 9, "bold"), width=cw, anchor="w",
                 bg="#dce3f0", padx=5).grid(row=0, column=col, padx=5, pady=4, sticky="w")

    ROW_COLORS = ("#ffffff", "#f0f4ff")

    rows = {}
    for r, eq_type in enumerate(selected_types):
        bg = ROW_COLORS[r % 2]
        row_bg = tk.Frame(outer, bg=bg, bd=1, relief="flat")
        row_bg.pack(fill="x", pady=1)
        tk.Label(row_bg, text=eq_type, width=COL_WIDTHS[0], anchor="w", bg=bg,
                 font=("Helvetica", 9, "bold")).grid(row=0, column=0, padx=5, pady=8, sticky="w")
        hostname_e = tk.Entry(row_bg, width=COL_WIDTHS[1], font=("Helvetica", 9)); hostname_e.grid(row=0, column=1, padx=5, pady=8)
        serial_e   = tk.Entry(row_bg, width=COL_WIDTHS[2], font=("Helvetica", 9)); serial_e.grid(row=0, column=2, padx=5, pady=8)
        checked_e  = tk.Entry(row_bg, width=COL_WIDTHS[3], font=("Helvetica", 9)); checked_e.grid(row=0, column=3, padx=5, pady=8)
        shelf_v = tk.StringVar()
        ttk.Combobox(row_bg, textvariable=shelf_v, values=shelf_list, width=COL_WIDTHS[4], state="readonly",
                     font=("Helvetica", 9)).grid(row=0, column=4, padx=5, pady=8)
        status_v = tk.StringVar()
        ttk.Combobox(row_bg, textvariable=status_v, values=STATUS_CHOICES,
                     width=COL_WIDTHS[5], state="readonly", font=("Helvetica", 9)).grid(row=0, column=5, padx=5, pady=8)
        remarks_e = tk.Entry(row_bg, width=COL_WIDTHS[6], font=("Helvetica", 9)); remarks_e.grid(row=0, column=6, padx=5, pady=8)
        rows[eq_type] = (hostname_e, serial_e, checked_e, shelf_v, status_v, remarks_e)

    error_lbl = tk.Label(outer, text="", fg="red", font=("Helvetica", 8))
    error_lbl.pack(pady=(6, 0))

    def confirm_set():
        df_items_w2 = load_items_w2()
        existing_serials_w2 = df_items_w2["Serial Number"].astype(str).tolist() if "Serial Number" in df_items_w2.columns else []
        staged_serials = []
        for ss in staged_sets:
            for it in ss["items"]:
                if it.get("Serial Number"):
                    staged_serials.append(it["Serial Number"])

        items = []
        for eq_type, (hostname_e, serial_e, checked_e, shelf_v, status_v, remarks_e) in rows.items():
            hostname   = hostname_e.get().strip()
            serial     = serial_e.get().strip()
            checked_by = checked_e.get().strip()
            shelf      = shelf_v.get().strip()
            status     = status_v.get().strip()
            remarks    = remarks_e.get().strip()

            if not hostname:
                error_lbl.config(text=f"Please enter a Hostname for {eq_type}"); return
            if not serial:
                error_lbl.config(text=f"Please enter a Serial Number for {eq_type}"); return
            if not checked_by:
                error_lbl.config(text=f"Please enter Checked By for {eq_type}"); return
            if not shelf:
                error_lbl.config(text=f"Please select a Shelf for {eq_type}"); return
            if not status:
                error_lbl.config(text=f"Please select a Status for {eq_type}"); return

            _df_sh_chk = load_shelves_w2()
            _sh_status = _df_sh_chk[_df_sh_chk["Shelf"] == shelf]["Status"].values
            if len(_sh_status) > 0 and _sh_status[0] == "FULL":
                error_lbl.config(text=f"Shelf '{shelf}' is marked FULL ({eq_type})"); return

            if serial in existing_serials_w2:
                error_lbl.config(text=f"Serial '{serial}' already exists in Warehouse 2"); return
            if serial in staged_serials:
                error_lbl.config(text=f"Serial '{serial}' already staged in another set"); return
            if serial in [i.get("Serial Number") for i in items]:
                error_lbl.config(text=f"Duplicate serial '{serial}' within this set"); return

            items.append({
                "Equipment Type": eq_type,
                "Hostname":       hostname,
                "Serial Number":  serial,
                "Checked By":     checked_by,
                "Shelf":          shelf,
                "Status":         status,
                "Remarks":        remarks,
            })

        staged_sets.append({"set_id": set_id, "items": items})
        for var in w2_equip_vars.values():
            var.set(False)
        build_win.destroy()
        update_w2_staged_display()
        messagebox.showinfo("Staged", f"{set_id} added to staging with {len(items)} item(s)")

    btn_f = tk.Frame(outer)
    btn_f.pack(pady=8)
    tk.Button(btn_f, text="STAGE", command=confirm_set, width=16).pack(side="left", padx=5)
    tk.Button(btn_f, text="CANCEL", command=build_win.destroy, width=10).pack(side="left", padx=5)

    # Auto-size AFTER all widgets (including buttons) are packed
    build_win.update_idletasks()
    build_win.geometry(f"{build_win.winfo_reqwidth()}x{build_win.winfo_reqheight()}")
    build_win.grab_set()
    build_win.focus_force()

def w2_remove_staged_set():
    global selected_set_index
    sel = w2_staged_listbox.curselection()
    if sel:
        index = sel[0]
        if index >= len(staged_sets):
            return
        removed = staged_sets.pop(index)
        selected_set_index = None
        update_w2_staged_display()
        messagebox.showinfo("Removed", f"{removed['set_id']} removed from staging")
    else:
        if not staged_sets:
            messagebox.showinfo("Info", "No staged sets to clear"); return
        if not messagebox.askyesno("Confirm", f"Clear all {len(staged_sets)} staged set(s)?"):
            return
        staged_sets.clear()
        selected_set_index = None
        update_w2_staged_display()
        messagebox.showinfo("Cleared", "All staged sets cleared")

def w2_import_excel_to_staging():
    """Import rows from an Excel file into W2 staging list as individual sets."""
    filepath = filedialog.askopenfilename(
        title="Import Excel — Warehouse 2",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    if not filepath: return
    try:
        df_imp = pd.read_excel(filepath, dtype=str).fillna("")
    except Exception as e:
        messagebox.showerror("Import Error", f"Could not read file:\n{e}"); return

    required = ["Equipment Type", "Hostname", "Serial Number", "Shelf", "Checked By", "Status"]
    missing_cols = [c for c in required if c not in df_imp.columns]
    if missing_cols:
        messagebox.showerror("Import Error",
            f"Missing required column(s): {', '.join(missing_cols)}\n\n"
            f"Required: Equipment Type, Hostname, Serial Number, Shelf, Checked By, Status\n"
            f"Optional: Remarks"); return

    df_items_w2   = load_items_w2()
    df_shelves_w2 = load_shelves_w2()
    valid_shelves     = df_shelves_w2["Shelf"].tolist()
    existing_hosts    = df_items_w2["Hostname"].values.tolist() if "Hostname" in df_items_w2.columns else []
    existing_serials  = df_items_w2["Serial Number"].astype(str).values.tolist() if "Serial Number" in df_items_w2.columns else []
    staged_hosts_all  = [it["Hostname"]      for s in staged_sets for it in s["items"]]
    staged_serial_all = [it["Serial Number"] for s in staged_sets for it in s["items"]]

    added = skipped = 0
    skip_log = []
    import re as _re
    for i, row in df_imp.iterrows():
        eq_type    = str(row.get("Equipment Type", "")).strip()
        hostname   = str(row.get("Hostname",        "")).strip()
        serial     = str(row.get("Serial Number",   "")).strip()
        shelf      = str(row.get("Shelf",           "")).strip()
        checked_by = str(row.get("Checked By",      "")).strip()
        status     = str(row.get("Status",          "")).strip()
        remarks    = str(row.get("Remarks",         "")).strip()
        row_num    = i + 2

        if not eq_type:    skip_log.append(f"Row {row_num}: Missing Equipment Type"); skipped += 1; continue
        if not hostname:   skip_log.append(f"Row {row_num}: Missing Hostname ({eq_type})"); skipped += 1; continue
        if not serial:     skip_log.append(f"Row {row_num}: Missing Serial Number ({hostname})"); skipped += 1; continue
        if not shelf:      skip_log.append(f"Row {row_num}: Missing Shelf ({hostname})"); skipped += 1; continue
        if not checked_by: skip_log.append(f"Row {row_num}: Missing Checked By ({hostname})"); skipped += 1; continue
        if not status:     skip_log.append(f"Row {row_num}: Missing Status ({hostname})"); skipped += 1; continue
        if status not in STATUS_CHOICES:
            skip_log.append(f"Row {row_num}: Invalid Status '{status}' ({hostname})"); skipped += 1; continue
        if shelf not in valid_shelves:
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' not found ({hostname})"); skipped += 1; continue
        sh_status = df_shelves_w2[df_shelves_w2["Shelf"] == shelf]["Status"].values
        if len(sh_status) > 0 and sh_status[0] == "FULL":
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' is FULL ({hostname})"); skipped += 1; continue
        if hostname in existing_hosts or hostname in staged_hosts_all:
            skip_log.append(f"Row {row_num}: Hostname already exists ({hostname})"); skipped += 1; continue
        if serial in existing_serials or serial in staged_serial_all:
            skip_log.append(f"Row {row_num}: Serial already exists ({serial})"); skipped += 1; continue

        set_id = f"IMP-{hostname}"
        item = {"Equipment Type": eq_type, "Hostname": hostname, "Serial Number": serial,
                "Checked By": checked_by, "Shelf": shelf, "Status": status, "Remarks": remarks}
        staged_sets.append({"set_id": set_id, "items": [item]})
        staged_hosts_all.append(hostname)
        staged_serial_all.append(serial)
        added += 1

    update_w2_staged_display()
    msg = f"Import complete.\n\n✅ Added: {added}\n⛔ Skipped: {skipped}"
    if skip_log: msg += "\n\nSkipped details:\n" + "\n".join(skip_log[:20])
    if len(skip_log) > 20: msg += f"\n...and {len(skip_log)-20} more."
    (messagebox.showinfo if skipped == 0 else messagebox.showwarning)("Import Result", msg)

def w2_put_warehouse():
    if not staged_sets:
        messagebox.showerror("Error", "No staged sets to put"); return

    total_items = sum(len(s["items"]) for s in staged_sets)
    if not messagebox.askyesno("Confirm",
        f"Put {len(staged_sets)} set(s) ({total_items} item(s)) to Warehouse 2?"):
        return

    try:
        df_w2 = load_items_w2()
        _df_sh_w2 = load_shelves_w2()

        # Re-check shelf FULL status at commit time (shelf may have changed since staging)
        for _s in staged_sets:
            for _item in _s["items"]:
                _shelf_val = _item["Shelf"]
                _s_stat = _df_sh_w2[_df_sh_w2["Shelf"] == _shelf_val]["Status"].values
                if len(_s_stat) > 0 and _s_stat[0] == "FULL":
                    messagebox.showerror("Error",
                        f"Shelf '{_shelf_val}' is marked FULL.\n"
                        f"Set {_s['set_id']} — {_item['Equipment Type']} cannot be placed there.\n"
                        "Edit or remove it from staging first.")
                    return

        now_str = _now().strftime("%Y-%m-%d %H:%M:%S")
        for s in staged_sets:
            set_id = s["set_id"]
            for item in s["items"]:
                eq_type = item["Equipment Type"]
                qr_label = f"{set_id}-{eq_type}"
                qr_code = str(uuid.uuid4())
                generate_qr(qr_label, qr_label, warehouse=2)
                df_w2 = pd.concat([df_w2, pd.DataFrame([{
                    "QR":             qr_code,
                    "Set ID":         set_id,
                    "Hostname":       item.get("Hostname", ""),
                    "Equipment Type": eq_type,
                    "Serial Number":  item.get("Serial Number", ""),
                    "Checked By":     item.get("Checked By", ""),
                    "Shelf":          item["Shelf"],
                    "Status":         item.get("Status", ""),
                    "Remarks":        item.get("Remarks", ""),
                    "Date":           now_str
                }])], ignore_index=True)
            save_log("PUT WAREHOUSE", f"[W2] Set: {set_id} | Items: {len(s['items'])}")

        save_warehouse_2(df_w2, load_shelves_w2())

        count = len(staged_sets)
        staged_sets.clear()
        update_w2_staged_display()
        w2_pull_item_entry.delete(0, tk.END)
        w2_search_label.config(text="", fg="blue")
        w2_status_label.config(text="")
        w2_refresh_all()
        messagebox.showinfo("Success", f"{count} set(s) added to Warehouse 2.\nQR codes generated.\nUse 'GENERATE FILES' to create PDF labels and export to Excel.")

    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save:\n{str(e)}")

# ========== W2 DISPLAY ==========

def _show_w2_tree(tree):
    for t in (tree_w2_warehouse, tree_w2_available, tree_w2_pullouts, tree_w2_qr):
        if t is not tree:
            t.pack_forget()
    tree.pack(fill="both", expand=True)

def w2_show_warehouse():
    w2_update_full_shelves_display()
    try: w2_back_to_wh_btn.pack_forget()
    except Exception: pass
    try: w2_back_to_stage_btn.pack(side="left", padx=(0, 6))
    except Exception: pass
    df = load_items_w2()
    if "Date" not in df.columns:
        df["Date"] = ""
    # Apply only date filters (mirrors W1 show_warehouse behaviour).
    # Keyword/shelf/type filters are intentionally NOT applied here so that
    # clicking SHOW WAREHOUSE always resets to the full warehouse view,
    # exactly like Warehouse 1.
    try:
        date_from = w2_date_from_var.get().strip()
        date_to   = w2_date_to_var.get().strip()
        df = _filter_by_date(df, date_from, date_to)
    except (NameError, Exception):
        pass
    _populate_w2_warehouse_tree(df)
    w2_search_label.config(text="", fg="blue")

def w2_show_available():
    _show_w2_tree(tree_w2_available)
    try: w2_back_to_wh_btn.pack_forget()
    except Exception: pass
    try: w2_back_to_stage_btn.pack_forget()
    except Exception: pass
    tree_w2_available.delete(*tree_w2_available.get_children())
    try:
        keyword = w2_search_entry.get().strip().lower()
    except Exception:
        keyword = ""
    df = load_shelves_w2().sort_values("Shelf")
    df_items_w2 = load_items_w2()
    if keyword:
        mask = False
        for col in ["Shelf", "Status", "Date_Full"]:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
        df = df[mask]
    for _, row in df.iterrows():
        date_full  = row.get("Date_Full", "")
        shelf_name = row["Shelf"]
        item_count = int((df_items_w2["Shelf"] == shelf_name).sum()) if "Shelf" in df_items_w2.columns else 0
        tree_w2_available.insert("", "end", values=(shelf_name, row["Status"], item_count, date_full if pd.notna(date_full) else ""))

def w2_show_pullouts():
    _show_w2_tree(tree_w2_pullouts)
    try: w2_back_to_stage_btn.pack_forget()
    except Exception: pass
    w2_back_to_wh_btn.pack(side="left", padx=(0, 6))
    tree_w2_pullouts.delete(*tree_w2_pullouts.get_children())
    df_po2 = load_pullouts_w2()
    # Only apply dropdown/date filters — keyword is NOT applied here so that
    # clicking PULL HISTORY always shows the full pull history, not a
    # search-box-filtered subset.
    try:
        shelf_f   = w2_pull_shelf_var.get()
        type_f    = w2_type_filter_var.get()
        date_from = w2_date_from_var.get().strip()
        date_to   = w2_date_to_var.get().strip()
    except (NameError, Exception):
        shelf_f = type_f = date_from = date_to = ""
    if shelf_f: df_po2 = df_po2[df_po2["Shelf"] == shelf_f]
    if type_f:  df_po2 = df_po2[df_po2["Equipment Type"] == type_f]
    df_po2 = _filter_by_date(df_po2, date_from, date_to)
    w2_pull_row_checks.clear()
    for _, row in df_po2.iterrows():
        key     = (str(row.get("Set ID", "")), str(row.get("Equipment Type", "")))
        checked = key in w2_pull_persistent_checks
        iid = tree_w2_pullouts.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c, "") for c in ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
        ))
        w2_pull_row_checks[iid] = checked
    try:
        all_reasons = sorted(load_pullouts_w2()["Pull Reason"].dropna().unique().tolist())
        w2_pull_reason_filter_entry["values"] = [""] + all_reasons
    except Exception:
        pass
    w2_search_label.config(text="", fg="blue")

def _populate_w2_warehouse_tree(df):
    _show_w2_tree(tree_w2_warehouse)
    tree_w2_warehouse.delete(*tree_w2_warehouse.get_children())
    w2_row_checks.clear()
    for _, row in df.iterrows():
        key     = (str(row.get("Set ID", "")), str(row.get("Equipment Type", "")))
        checked = key in w2_persistent_checks
        iid = tree_w2_warehouse.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c, "") for c in ["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"])
        ))
        w2_row_checks[iid] = checked
    _w2_refresh_select_all_label()

def w2_search_item():
    keyword   = w2_search_entry.get().strip().lower()
    shelf_f   = w2_pull_shelf_var.get()
    type_f    = w2_type_filter_var.get()
    date_from = w2_date_from_var.get().strip()
    date_to   = w2_date_to_var.get().strip()

    # ── Pull history view is active: filter it instead of warehouse ──
    if tree_w2_pullouts.winfo_ismapped():
        df = load_pullouts_w2()
        if keyword:
            search_cols = ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_f: df = df[df["Shelf"] == shelf_f]
        if type_f:  df = df[df["Equipment Type"] == type_f]
        df = _filter_by_date(df, date_from, date_to)
        tree_w2_pullouts.delete(*tree_w2_pullouts.get_children())
        w2_pull_row_checks.clear()
        for _, row in df.iterrows():
            key     = (str(row.get("Set ID", "")), str(row.get("Equipment Type", "")))
            checked = key in w2_pull_persistent_checks
            iid = tree_w2_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c, "") for c in ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
            ))
            w2_pull_row_checks[iid] = checked
        parts = []
        if keyword:   parts.append(f"Search: \"{keyword}\"")
        if shelf_f:   parts.append(f"Shelf: {shelf_f}")
        if type_f:    parts.append(f"Type: {type_f}")
        if date_from: parts.append(f"From: {date_from}")
        if date_to:   parts.append(f"To: {date_to}")
        label = (f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")) if parts else ""
        w2_search_label.config(text=label, fg="darkorange" if parts else "blue")
        return

    # ── Default: filter warehouse view ───────────────────────────────
    df = load_items_w2()
    if keyword:
        search_cols = ["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]
        mask = False
        for col in search_cols:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
        df = df[mask]
    if shelf_f: df = df[df["Shelf"] == shelf_f]
    if type_f:  df = df[df["Equipment Type"] == type_f]
    df = _filter_by_date(df, date_from, date_to)
    _populate_w2_warehouse_tree(df)
    parts = []
    if keyword:   parts.append(f"Search: \"{keyword}\"")
    if shelf_f:   parts.append(f"Shelf: {shelf_f}")
    if type_f:    parts.append(f"Type: {type_f}")
    if date_from: parts.append(f"From: {date_from}")
    if date_to:   parts.append(f"To: {date_to}")
    label = (f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")) if parts else ""
    w2_search_label.config(text=label, fg="darkorange" if parts else "blue")

def w2_clear_filters():
    w2_pull_shelf_var.set("")
    w2_type_filter_var.set("")
    w2_search_entry.delete(0, tk.END)
    w2_date_from_var.set("")
    w2_date_to_var.set("")
    w2_pull_reason_filter_var.set("")
    w2_pull_date_from_var.set("")
    w2_pull_date_to_var.set("")
    w2_search_label.config(text="", fg="blue")
    w2_persistent_checks.clear()
    w2_pull_persistent_checks.clear()
    w2_show_warehouse()

def w2_update_full_shelves_display():
    df_shelves = load_shelves_w2()
    full_shelves = df_shelves[df_shelves["Status"] == "FULL"]["Shelf"].tolist()
    w2_full_label.config(text="FULL Shelves:\n" + "\n".join(full_shelves) if full_shelves else "FULL Shelves: None")

def w2_refresh_all():
    w2_show_warehouse()
    update_all_shelf_dropdowns()

# ========== W2 PULL OUT ==========

def w2_pull_search_live(event=None):
    """Filter whichever W2 view is currently active based on the search box."""
    keyword = w2_pull_item_entry.get().strip().lower()

    if tree_w2_available.winfo_ismapped():
        # Shelf status view is active
        df = load_shelves_w2().sort_values("Shelf")
        df_items_all = load_items_w2()
        if keyword:
            mask = False
            for col in ["Shelf", "Status", "Date_Full"]:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        tree_w2_available.delete(*tree_w2_available.get_children())
        for _, row in df.iterrows():
            date_full  = row.get("Date_Full", "")
            shelf_name = row["Shelf"]
            item_count = int((df_items_all["Shelf"] == shelf_name).sum()) if "Shelf" in df_items_all.columns else 0
            tree_w2_available.insert("", "end", values=(shelf_name, row["Status"], item_count, date_full if pd.notna(date_full) else ""))
        w2_search_label.config(text=f"{len(df)} match(es)" if keyword else "", fg="blue")

    elif tree_w2_pullouts.winfo_ismapped():
        # Pull history view is active — respect all active filters
        df = load_pullouts_w2()
        shelf_f   = w2_pull_shelf_var.get()
        type_f    = w2_type_filter_var.get()
        date_from = w2_date_from_var.get().strip()
        date_to   = w2_date_to_var.get().strip()
        if keyword:
            search_cols = ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_f: df = df[df["Shelf"] == shelf_f]
        if type_f:  df = df[df["Equipment Type"] == type_f]
        df = _filter_by_date(df, date_from, date_to)
        tree_w2_pullouts.delete(*tree_w2_pullouts.get_children())
        w2_pull_row_checks.clear()
        for _, row in df.iterrows():
            key     = (str(row.get("Set ID", "")), str(row.get("Equipment Type", "")))
            checked = key in w2_pull_persistent_checks
            iid = tree_w2_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c, "") for c in ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
            ))
            w2_pull_row_checks[iid] = checked
        active = bool(keyword or shelf_f or type_f or date_from or date_to)
        w2_search_label.config(text=f"{len(df)} match(es)" if active else "", fg="darkorange" if active else "blue")

    else:
        # Warehouse view (default)
        w2_show_warehouse()
        if keyword:
            df = load_items_w2()
            search_cols = ["QR", "Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Date"]
            mask = False
            for col in search_cols:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
            _populate_w2_warehouse_tree(df)
            w2_search_label.config(text=f"{len(df)} match(es)", fg="blue")
        else:
            w2_search_label.config(text="", fg="blue")

def w2_select_item(event):
    selected = tree_w2_warehouse.selection()
    if selected:
        iid = selected[0]
        values = tree_w2_warehouse.item(iid, "values")
        # values: ☐(0), QR(1), Set ID(2), Hostname(3), Equipment Type(4), Serial(5), Checked By(6), Shelf(7), Status(8), Remarks(9), Date(10)

        # Toggle checkbox on single-click
        if iid in w2_row_checks:
            w2_row_checks[iid] = not w2_row_checks[iid]
            key = (str(values[2]), str(values[4]))
            if w2_row_checks[iid]:
                w2_persistent_checks.add(key)
            else:
                w2_persistent_checks.discard(key)
            tree_w2_warehouse.set(iid, "C0", "☑" if w2_row_checks[iid] else "☐")
            _w2_refresh_select_all_label()

        w2_status_label.config(
            text=f"Selected → {values[2]} ({values[4]})  |  Hostname: {values[3]}  |  Shelf: {values[7]}  |  Serial: {values[5]}",
            fg="#1a5276")

def w2_unstage_from_warehouse(event=None):
    # Collect checked rows; fall back to treeview selection if nothing checked
    checked = [iid for iid, state in w2_row_checks.items() if state]
    if not checked:
        sel = tree_w2_warehouse.selection()
        if sel:
            checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Stage", "Check at least one row in the Warehouse table.")
        return

    # Single confirmation for all selected items
    preview = []
    for item_id in checked:
        v = tree_w2_warehouse.item(item_id, "values")
        if v:
            preview.append(f"  • {v[4]} ({v[2]}) | Hostname: {v[3]}  (Shelf: {v[7]})")
    if not messagebox.askyesno("Move to Staging",
            f"Move {len(preview)} item(s) back to staging?\n\n" + "\n".join(preview)):
        return

    moved = 0
    for item_id in checked:
        values = tree_w2_warehouse.item(item_id, "values")
        if not values:
            continue
        # values: ☐(0), QR(1), Set ID(2), Hostname(3), Equipment Type(4), Serial(5), Checked By(6), Shelf(7), Status(8), Remarks(9), Date(10)
        set_id, hostname, eq_type, serial, checked_by, shelf, status, remarks = values[2], values[3], values[4], values[5], values[6], values[7], values[8], values[9]

        df_w2 = load_items_w2()
        match = df_w2[(df_w2["Set ID"] == set_id) & (df_w2["Equipment Type"] == eq_type)]
        if match.empty:
            messagebox.showerror("Error", "Item not found in warehouse")
            continue

        qr_label = f"{set_id}-{eq_type}"
        remove_qr(qr_label, warehouse=2)
        df_w2 = df_w2.drop(match.index).reset_index(drop=True)
        save_warehouse_2(df_w2, load_shelves_w2())

        staged_sets.append({"set_id": set_id, "items": [{
            "Equipment Type": eq_type,
            "Hostname":       hostname,
            "Serial Number":  serial,
            "Checked By":     checked_by,
            "Shelf":          shelf,
            "Status":         status,
            "Remarks":        remarks,
        }]})
        save_log("UNSTAGE", f"[W2] Set: {set_id} | Item: {eq_type} | Shelf: {shelf}")
        moved += 1

    if moved:
        messagebox.showinfo("Moved", f"{moved} item(s) moved back to staging.")
        update_w2_staged_display()
        w2_pull_item_entry.delete(0, tk.END)
        w2_search_label.config(text="", fg="blue")
        w2_status_label.config(text="")
        w2_show_warehouse()
        update_all_shelf_dropdowns()

def w2_pull_item():
    reason = w2_pull_reason_filter_var.get().strip()
    if not reason:
        messagebox.showerror("Error", "Please enter a pull reason"); return

    # Collect checked rows; fall back to treeview selection; then fall back to entry text
    checked = [iid for iid, state in w2_row_checks.items() if state]
    if checked:
        target_iids = checked
    elif tree_w2_warehouse.selection():
        target_iids = list(tree_w2_warehouse.selection())
    else:
        # Legacy: single item from entry text "SET-001 - Monitor"
        selection_text = w2_pull_item_entry.get().strip()
        if not selection_text:
            messagebox.showerror("Error", "Select or check item(s) to pull, or select a row"); return
        try:
            set_id, eq_type = [x.strip() for x in selection_text.split(" - ", 1)]
        except ValueError:
            messagebox.showerror("Error", "Invalid selection format"); return
        df_w2  = load_items_w2()
        df_po2 = load_pullouts_w2()
        match  = df_w2[(df_w2["Set ID"] == set_id) & (df_w2["Equipment Type"] == eq_type)]
        if match.empty:
            messagebox.showerror("Error", f"'{selection_text}' not found in Warehouse 2"); return
        item_row = match.iloc[0]
        hostname = str(item_row.get("Hostname", ""))
        if not messagebox.askyesno("Confirm Pull Out",
                f"Pull out {eq_type}?\nSet ID: {set_id} | Hostname: {hostname}\nReason: {reason}"):
            return
        qr_label = f"{set_id}-{eq_type}"
        pull_qr(qr_label, warehouse=2)
        df_w2  = df_w2.drop(match.index).reset_index(drop=True)
        df_po2 = pd.concat([df_po2, pd.DataFrame([{
            "QR":             str(item_row.get("QR", "")),
            "Set ID":         set_id,
            "Hostname":       str(item_row.get("Hostname", "")),
            "Equipment Type": eq_type,
            "Serial Number":  str(item_row.get("Serial Number", "")),
            "Checked By":     str(item_row.get("Checked By", "")),
            "Shelf":          str(item_row.get("Shelf", "")),
            "Status":         str(item_row.get("Status", "")),
            "Remarks":        str(item_row.get("Remarks", "")),
            "Pull Reason":    reason,
            "Date":           _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_warehouse_2(df_w2, load_shelves_w2(), df_po2)
        save_log("WAREHOUSE PULL", f"[W2] Set: {set_id} | Item: {eq_type} | Reason: {reason}")
        try:
            all_reasons = sorted(load_pullouts_w2()["Pull Reason"].dropna().unique().tolist())
            w2_pull_reason_filter_entry["values"] = [""] + all_reasons
        except Exception:
            pass
        messagebox.showinfo("Success", f"{eq_type} from {set_id} pulled out successfully")
        w2_pull_item_entry.delete(0, tk.END)
        w2_pull_reason_filter_var.set("")
        w2_refresh_all()
        return

    # ── Bulk pull from checked/selected rows ──────────────────
    # W2 tree values: ☐(0), QR(1), Set ID(2), Hostname(3), Equip Type(4), Serial(5), Checked By(6), Shelf(7), Status(8), Remarks(9), Date(10)
    targets = [(tree_w2_warehouse.item(iid, "values")[2],
                tree_w2_warehouse.item(iid, "values")[3],
                tree_w2_warehouse.item(iid, "values")[4]) for iid in target_iids]

    if not messagebox.askyesno("Confirm Pull Out",
            f"Pull out {len(targets)} item(s)?\n" +
            "\n".join(f"  • {eq}  |  Set ID: {sid}  |  Hostname: {host}" for sid, host, eq in targets) +
            f"\n\nReason: {reason}"):
        return

    df_w2  = load_items_w2()
    df_po2 = load_pullouts_w2()
    pulled = 0
    for set_id, _, eq_type in targets:
        match = df_w2[(df_w2["Set ID"] == set_id) & (df_w2["Equipment Type"] == eq_type)]
        if match.empty:
            continue
        item_row = match.iloc[0]
        qr_label = f"{set_id}-{eq_type}"
        pull_qr(qr_label, warehouse=2)
        df_w2  = df_w2.drop(match.index).reset_index(drop=True)
        df_po2 = pd.concat([df_po2, pd.DataFrame([{
            "QR":             str(item_row.get("QR", "")),
            "Set ID":         set_id,
            "Hostname":       str(item_row.get("Hostname", "")),
            "Equipment Type": eq_type,
            "Serial Number":  str(item_row.get("Serial Number", "")),
            "Checked By":     str(item_row.get("Checked By", "")),
            "Shelf":          str(item_row.get("Shelf", "")),
            "Status":         str(item_row.get("Status", "")),
            "Remarks":        str(item_row.get("Remarks", "")),
            "Pull Reason":    reason,
            "Date":           _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_log("WAREHOUSE PULL", f"[W2] Set: {set_id} | Item: {eq_type} | Reason: {reason}")
        pulled += 1

    save_warehouse_2(df_w2, load_shelves_w2(), df_po2)
    try:
        all_reasons = sorted(load_pullouts_w2()["Pull Reason"].dropna().unique().tolist())
        w2_pull_reason_filter_entry["values"] = [""] + all_reasons
    except Exception:
        pass
    messagebox.showinfo("Success", f"{pulled} item(s) pulled out successfully")
    w2_pull_item_entry.delete(0, tk.END)
    w2_pull_reason_filter_var.set("")
    w2_refresh_all()

def w2_undo_pull(event=None):
    # Collect checked rows; fall back to treeview selection if nothing checked
    checked = [iid for iid, state in w2_pull_row_checks.items() if state]
    if not checked:
        sel = tree_w2_pullouts.selection()
        if sel:
            checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Warehouse", "Check at least one row in the Pull History table.")
        return

    # Single confirmation for all selected items
    preview = []
    for item_id in checked:
        v = tree_w2_pullouts.item(item_id, "values")
        if v:
            preview.append(f"  • {v[3]} — {v[2]}  |  Set ID: {v[1]}  (Shelf: {v[5]})")
    if not messagebox.askyesno("Undo Pull",
            f"Restore {len(preview)} item(s) back to Warehouse 2?\n\n" + "\n".join(preview)):
        return

    # ── Pre-validate: block entire restore if any target shelf is FULL ──
    _df_sh_w2_chk = load_shelves_w2()
    for item_id in checked:
        _v = tree_w2_pullouts.item(item_id, "values")
        if not _v:
            continue
        _shelf = _v[5]
        _s_stat = _df_sh_w2_chk[_df_sh_w2_chk["Shelf"] == _shelf]["Status"].values
        if len(_s_stat) > 0 and _s_stat[0] == "FULL":
            messagebox.showerror("Shelf Full",
                f"Cannot restore — shelf '{_shelf}' is marked FULL.\n"
                "Set it to AVAILABLE first, then retry.")
            return

    restored = 0
    for item_id in checked:
        values = tree_w2_pullouts.item(item_id, "values")
        if not values:
            continue
        # values: ☐(0), Set ID(1), Hostname(2), Equip Type(3), Serial(4), Shelf(5), Status(6), Remarks(7), PullReason(8), Date(9)
        set_id, hostname, eq_type, shelf, status, remarks = values[1], values[2], values[3], values[5], values[6], values[7]

        df_w2 = load_items_w2()
        df_po2 = load_pullouts_w2()

        match = df_po2[(df_po2["Set ID"] == set_id) & (df_po2["Equipment Type"] == eq_type)]
        if match.empty:
            messagebox.showerror("Error", "Record not found in pull history")
            continue

        pull_row = match.iloc[0]
        qr_label = f"{set_id}-{eq_type}"

        # ── Restore QR: move back from pull_qrs/ instead of regenerating ──
        import shutil
        pull_qr_file = pull_qr_path_for(qr_label, warehouse=2)
        wh_qr_file   = qr_path_for(qr_label, warehouse=2)
        qr_code = str(pull_row.get("QR", ""))
        if os.path.exists(pull_qr_file):
            try:
                os.makedirs(QR_FOLDER_W2, exist_ok=True)
                shutil.move(pull_qr_file, wh_qr_file)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR file could not be moved back: {e}")
        elif not os.path.exists(wh_qr_file):
            # Fallback: regenerate only if truly missing from both locations
            try:
                qr_code = str(uuid.uuid4())
                generate_qr(qr_label, qr_code, warehouse=2)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR not regenerated: {e}")

        df_w2 = pd.concat([df_w2, pd.DataFrame([{
            "QR":             qr_code,
            "Set ID":         set_id,
            "Hostname":       str(pull_row.get("Hostname", "")),
            "Equipment Type": eq_type,
            "Serial Number":  str(pull_row.get("Serial Number", "")),
            "Checked By":     str(pull_row.get("Checked By", "")),
            "Shelf":          shelf,
            "Status":         status,
            "Remarks":        remarks,
            "Date":           _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)

        df_po2 = df_po2.drop(match.index).reset_index(drop=True)
        save_warehouse_2(df_w2, load_shelves_w2(), df_po2)
        save_log("UNDO PULL", f"[W2] Set: {set_id} | Item: {eq_type} | Shelf: {shelf}")
        restored += 1

    if restored:
        messagebox.showinfo("Restored", f"{restored} item(s) restored to Warehouse 2.")
    w2_show_pullouts()

# ========== W2 ITEM MANAGEMENT ==========

def w2_update_item():
    """Update a staged set item OR a warehouse item, depending on what is selected."""

    # ── Branch A: edit a STAGED set ──────────────────────────────────────────
    staged_sel = w2_staged_listbox.curselection()
    if staged_sel:
        s_idx = staged_sel[0]
        if s_idx >= len(staged_sets):
            return
        staged_set = staged_sets[s_idx]
        set_id     = staged_set["set_id"]
        items      = staged_set["items"]
        shelf_list = sorted(load_shelves_w2()["Shelf"].tolist())

        edit_win = tk.Toplevel(root)
        edit_win.title(f"Edit Staged Set — {set_id}")
        edit_win.resizable(False, False)
        edit_win.transient(root)

        tk.Label(edit_win, text=f"Edit Staged Set: {set_id}",
                 font=("Helvetica", 10, "bold"), bg="#27ae60", fg="white",
                 padx=10, pady=6).grid(row=0, column=0, columnspan=2, sticky="we")

        COL_WIDTHS = [12, 18, 18, 16, 16, 13, 20]
        HEADERS    = ["Equipment", "Hostname", "Serial Number",
                      "Checked By", "Shelf", "Status", "Remarks"]

        outer = tk.Frame(edit_win, padx=10, pady=5)
        outer.grid(row=1, column=0, columnspan=2)

        hdr_frame = tk.Frame(outer, bg="#dce3f0")
        hdr_frame.pack(fill="x")
        for col, (h, cw) in enumerate(zip(HEADERS, COL_WIDTHS)):
            tk.Label(hdr_frame, text=h, font=("Helvetica", 9, "bold"), width=cw, anchor="w",
                     bg="#dce3f0", padx=5).grid(row=0, column=col, padx=5, pady=4, sticky="w")

        ROW_COLORS = ("#ffffff", "#f0f4ff")
        row_widgets = {}
        for r, item in enumerate(items):
            bg = ROW_COLORS[r % 2]
            eq_type = item["Equipment Type"]
            row_bg  = tk.Frame(outer, bg=bg, bd=1, relief="flat")
            row_bg.pack(fill="x", pady=1)
            tk.Label(row_bg, text=eq_type, width=COL_WIDTHS[0], anchor="w",
                     bg=bg, font=("Helvetica", 9, "bold")).grid(row=0, column=0, padx=5, pady=8, sticky="w")
            hostname_e = tk.Entry(row_bg, width=COL_WIDTHS[1], font=("Helvetica", 9))
            hostname_e.insert(0, item.get("Hostname", ""))
            hostname_e.grid(row=0, column=1, padx=5, pady=8)
            serial_e = tk.Entry(row_bg, width=COL_WIDTHS[2], font=("Helvetica", 9))
            serial_e.insert(0, item.get("Serial Number", ""))
            serial_e.grid(row=0, column=2, padx=5, pady=8)
            checked_e = tk.Entry(row_bg, width=COL_WIDTHS[3], font=("Helvetica", 9))
            checked_e.insert(0, item.get("Checked By", ""))
            checked_e.grid(row=0, column=3, padx=5, pady=8)
            shelf_v = tk.StringVar(value=item.get("Shelf", ""))
            ttk.Combobox(row_bg, textvariable=shelf_v, values=shelf_list,
                         width=COL_WIDTHS[4], state="readonly",
                         font=("Helvetica", 9)).grid(row=0, column=4, padx=5, pady=8)
            status_v = tk.StringVar(value=item.get("Status", ""))
            ttk.Combobox(row_bg, textvariable=status_v,
                         values=STATUS_CHOICES,
                         width=COL_WIDTHS[5], state="readonly",
                         font=("Helvetica", 9)).grid(row=0, column=5, padx=5, pady=8)
            remarks_e = tk.Entry(row_bg, width=COL_WIDTHS[6], font=("Helvetica", 9))
            remarks_e.insert(0, item.get("Remarks", ""))
            remarks_e.grid(row=0, column=6, padx=5, pady=8)
            row_widgets[eq_type] = (hostname_e, serial_e, checked_e,
                                    shelf_v, status_v, remarks_e)

        error_lbl = tk.Label(edit_win, text="", fg="red", font=("Helvetica", 8))
        error_lbl.grid(row=2, column=0, columnspan=2, pady=(4, 0))

        def save_staged_update():
            df_items_w2      = load_items_w2()
            existing_serials = df_items_w2["Serial Number"].astype(str).tolist() \
                               if "Serial Number" in df_items_w2.columns else []
            # serials from OTHER staged sets
            other_staged_serials = [
                it["Serial Number"]
                for si, ss in enumerate(staged_sets) if si != s_idx
                for it in ss["items"] if it.get("Serial Number")
            ]
            seen_in_dialog = []
            updated_items  = []
            for eq_type, (hostname_e, serial_e, checked_e,
                          shelf_v, status_v, remarks_e) in row_widgets.items():
                hn  = hostname_e.get().strip()
                sn  = serial_e.get().strip()
                cb  = checked_e.get().strip()
                sh  = shelf_v.get().strip()
                st  = status_v.get().strip()
                rm  = remarks_e.get().strip()
                if not hn:  error_lbl.config(text=f"Hostname required for {eq_type}");       return
                if not sn:  error_lbl.config(text=f"Serial Number required for {eq_type}");  return
                if not cb:  error_lbl.config(text=f"Checked By required for {eq_type}");     return
                if not sh:  error_lbl.config(text=f"Please select a Shelf for {eq_type}");   return
                if not st:  error_lbl.config(text=f"Please select a Status for {eq_type}");  return
                # Find original serial for this item to allow keeping same value
                orig_sn = next((it["Serial Number"] for it in items
                                if it["Equipment Type"] == eq_type), None)
                if sn != orig_sn:
                    if sn in existing_serials:
                        error_lbl.config(text=f"Serial '{sn}' already in Warehouse 2"); return
                    if sn in other_staged_serials:
                        error_lbl.config(text=f"Serial '{sn}' in another staged set"); return
                if sn in seen_in_dialog:
                    error_lbl.config(text=f"Duplicate serial '{sn}' within this set"); return
                seen_in_dialog.append(sn)
                updated_items.append({
                    "Equipment Type": eq_type, "Hostname": hn,
                    "Serial Number": sn, "Checked By": cb,
                    "Shelf": sh, "Status": st, "Remarks": rm,
                })
            staged_sets[s_idx]["items"] = updated_items
            edit_win.destroy()
            update_w2_staged_display()
            messagebox.showinfo("Updated", f"{set_id} staging updated successfully")

        btn_f = tk.Frame(edit_win)
        btn_f.grid(row=3, column=0, columnspan=2, pady=10)
        tk.Button(btn_f, text="SAVE UPDATE", command=save_staged_update, width=14,
                  bg="#27ae60", fg="white").pack(side="left", padx=6)
        tk.Button(btn_f, text="Cancel", command=edit_win.destroy, width=10).pack(side="left", padx=6)

        edit_win.update_idletasks()
        edit_win.geometry(f"{edit_win.winfo_reqwidth()}x{edit_win.winfo_reqheight()}")
        edit_win.grab_set()
        edit_win.focus_force()
        return

    # ── Branch B: warehouse-table update is intentionally disabled ────────────
    messagebox.showerror(
        "Update Not Allowed",
        "Items cannot be updated directly from the warehouse.\n\n"
        "Double-click the row to move it back to staging,\n"
        "then select it in the staged sets list and click UPDATE ITEM.")

# ========== W2 SHELF MANAGEMENT ==========

def w2_set_shelf_status(new_status):
    shelf = w2_shelf_control_var.get()
    if not shelf:
        messagebox.showerror("Error", "Select a shelf"); return
    df_items_w2 = load_items_w2()
    df_shelves_w2 = load_shelves_w2()
    idx = df_shelves_w2[df_shelves_w2["Shelf"] == shelf].index
    if len(idx) == 0:
        return
    df_shelves_w2.at[idx[0], "Status"] = new_status
    df_shelves_w2.at[idx[0], "Date_Full"] = _now().strftime("%Y-%m-%d %H:%M:%S") if new_status == "FULL" else None
    save_warehouse_2(df_items_w2, df_shelves_w2)
    save_log("SHELF STATUS", f"[W2] Shelf: {shelf} → {new_status}")
    w2_status_label.config(text=f"{shelf} → {new_status}")
    w2_refresh_all()

def w2_add_shelf():
    new_shelf = w2_remove_shelf_var.get().strip()
    if not new_shelf:
        messagebox.showerror("Error", "Enter shelf name"); return
    df_shelves_w2 = load_shelves_w2()
    if new_shelf in df_shelves_w2["Shelf"].values:
        messagebox.showerror("Error", "Shelf already exists in W2"); return
    df_shelves_w2 = pd.concat([df_shelves_w2, pd.DataFrame([{"Shelf": new_shelf, "Status": "AVAILABLE"}])], ignore_index=True)
    df_shelves_w2 = df_shelves_w2.sort_values("Shelf", ignore_index=True)
    save_warehouse_2(load_items_w2(), df_shelves_w2)
    messagebox.showinfo("Success", f"Shelf '{new_shelf}' added to Warehouse 2")
    save_log("ADD SHELF", f"[W2] Shelf: {new_shelf}")
    w2_remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def w2_remove_shelf():
    shelf_name = w2_remove_shelf_var.get().strip()
    if not shelf_name:
        messagebox.showerror("Error", "Select a shelf to remove"); return
    df_items_w2 = load_items_w2()
    df_shelves_w2 = load_shelves_w2()
    if not df_items_w2[df_items_w2["Shelf"] == shelf_name].empty:
        messagebox.showerror("Error", f"Cannot remove shelf '{shelf_name}' — it still has items"); return
    if shelf_name not in df_shelves_w2["Shelf"].values:
        messagebox.showerror("Error", f"Shelf '{shelf_name}' does not exist in W2"); return
    df_shelves_w2 = df_shelves_w2[df_shelves_w2["Shelf"] != shelf_name].sort_values("Shelf", ignore_index=True)
    save_warehouse_2(df_items_w2, df_shelves_w2)
    messagebox.showinfo("Success", f"Shelf '{shelf_name}' removed from Warehouse 2")
    save_log("REMOVE SHELF", f"[W2] Shelf: {shelf_name}")
    w2_remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def w2_reset_shelf_control():
    w2_shelf_control_var.set("")
    w2_status_label.config(text="")

def w2_reset_shelf_addition():
    w2_remove_shelf_var.set("")
    w2_status_label.config(text="")

def w2_reset_pull_out():
    w2_pull_item_entry.delete(0, tk.END)
    w2_pull_reason_filter_var.set("")
    w2_pull_date_from_var.set("")
    w2_pull_date_to_var.set("")
    w2_status_label.config(text="")
    w2_search_label.config(text="")
    w2_show_warehouse()

def w2_filter_pull_history():
    """Filter and display the W2 pull history table only. Completely separate from warehouse search."""
    reason = w2_pull_reason_filter_var.get().strip().lower()
    date_from = w2_pull_date_from_var.get().strip()
    date_to = w2_pull_date_to_var.get().strip()
    w2_show_pullouts()  # always switch to pull history view first
    if not reason and not date_from and not date_to:
        return
    tree_w2_pullouts.delete(*tree_w2_pullouts.get_children())
    df = load_pullouts_w2()
    if reason:
        search_cols = ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Checked By", "Shelf", "Status", "Remarks", "Pull Reason", "Date"]
        mask = False
        for col in search_cols:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(reason, na=False)
        df = df[mask]
    df = _filter_by_date(df, date_from, date_to)
    # Populate the pull reason dropdown with known values from data
    all_reasons = sorted(load_pullouts_w2()["Pull Reason"].dropna().unique().tolist())
    w2_pull_reason_filter_entry["values"] = [""] + all_reasons
    w2_pull_row_checks.clear()
    for _, row in df.iterrows():
        iid = tree_w2_pullouts.insert("", "end", values=(
            "☐",
            *tuple(row.get(c, "") for c in ["Set ID", "Hostname", "Equipment Type", "Serial Number", "Shelf", "Status", "Remarks", "Pull Reason", "Date"])
        ))
        w2_pull_row_checks[iid] = False
    w2_search_label.config(text=f"Pull History Filtered: {len(df)} result(s)", fg="darkorange")

# ========== YUBIKEY (YK) ==========

def update_yk_staged_display():
    yk_staged_listbox.delete(0, tk.END)
    if not staged_yk_items:
        yk_staged_listbox.insert(tk.END, "No staged items")
        return
    for item in staged_yk_items:
        yk_staged_listbox.insert(tk.END,
            f"{item['Hostname']} | S/N: {item['Serial Number']} → {item.get('Status','')}")

def select_yk_staged_item(event):
    global selected_staged_yk_index
    sel = yk_staged_listbox.curselection()
    if not sel:
        return
    idx = sel[0]
    selected_staged_yk_index = idx
    item = staged_yk_items[idx]
    yk_hostname_entry.delete(0, tk.END);    yk_hostname_entry.insert(0, item["Hostname"])
    yk_serial_entry.delete(0, tk.END);     yk_serial_entry.insert(0, item.get("Serial Number",""))
    yk_checked_by_entry.delete(0, tk.END); yk_checked_by_entry.insert(0, item.get("Checked By",""))
    yk_shelf_var.set(item.get("Shelf",""))
    yk_status_var.set(item.get("Status",""))
    yk_remarks_var.set(item.get("Remarks",""))

def _yk_clear_fields():
    yk_hostname_entry.delete(0, tk.END)
    yk_serial_entry.delete(0, tk.END)
    yk_checked_by_entry.delete(0, tk.END)
    yk_shelf_var.set("")
    yk_status_var.set("")
    yk_remarks_var.set("")

def yk_import_excel_to_staging():
    """Import rows from an Excel file into YK staging list."""
    filepath = filedialog.askopenfilename(
        title="Import Excel — Yubikey",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    if not filepath: return
    try:
        df_imp = pd.read_excel(filepath, dtype=str).fillna("")
    except Exception as e:
        messagebox.showerror("Import Error", f"Could not read file:\n{e}"); return

    required = ["Hostname", "Serial Number", "Shelf", "Checked By", "Status"]
    missing_cols = [c for c in required if c not in df_imp.columns]
    if missing_cols:
        messagebox.showerror("Import Error",
            f"Missing required column(s): {', '.join(missing_cols)}\n\n"
            f"Required: Hostname, Serial Number, Shelf, Checked By, Status\n"
            f"Optional: Remarks"); return

    df_yk         = load_items_yk()
    df_shelves_yk = load_shelves_yk()
    valid_shelves    = df_shelves_yk["Shelf"].tolist()
    existing_hosts   = df_yk["Hostname"].values.tolist() if "Hostname" in df_yk.columns else []
    existing_serials = df_yk["Serial Number"].astype(str).values.tolist() if "Serial Number" in df_yk.columns else []
    staged_hosts     = [i["Hostname"]      for i in staged_yk_items]
    staged_serials   = [i["Serial Number"] for i in staged_yk_items]

    added = skipped = 0
    skip_log = []
    for i, row in df_imp.iterrows():
        hostname   = str(row.get("Hostname",       "")).strip()
        serial     = str(row.get("Serial Number",  "")).strip()
        checked_by = str(row.get("Checked By",     "")).strip()
        shelf      = str(row.get("Shelf",          "")).strip()
        status     = str(row.get("Status",         "")).strip()
        remarks    = str(row.get("Remarks",        "")).strip()
        row_num    = i + 2

        if not hostname:   skip_log.append(f"Row {row_num}: Missing Hostname"); skipped += 1; continue
        if not serial:     skip_log.append(f"Row {row_num}: Missing Serial Number ({hostname})"); skipped += 1; continue
        if not checked_by: skip_log.append(f"Row {row_num}: Missing Checked By ({hostname})"); skipped += 1; continue
        if not shelf:      skip_log.append(f"Row {row_num}: Missing Shelf ({hostname})"); skipped += 1; continue
        if not status:     skip_log.append(f"Row {row_num}: Missing Status ({hostname})"); skipped += 1; continue
        if status not in STATUS_CHOICES:
            skip_log.append(f"Row {row_num}: Invalid Status '{status}' ({hostname})"); skipped += 1; continue
        if shelf not in valid_shelves:
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' not found ({hostname})"); skipped += 1; continue
        sh_status = df_shelves_yk[df_shelves_yk["Shelf"] == shelf]["Status"].values
        if len(sh_status) > 0 and sh_status[0] == "FULL":
            skip_log.append(f"Row {row_num}: Shelf '{shelf}' is FULL ({hostname})"); skipped += 1; continue
        if hostname in existing_hosts or hostname in staged_hosts:
            skip_log.append(f"Row {row_num}: Hostname already exists ({hostname})"); skipped += 1; continue
        if serial in existing_serials or serial in staged_serials:
            skip_log.append(f"Row {row_num}: Serial Number already exists ({serial})"); skipped += 1; continue

        staged_yk_items.append({"Hostname": hostname, "Serial Number": serial,
                                 "Checked By": checked_by, "Shelf": shelf,
                                 "Status": status, "Remarks": remarks})
        staged_hosts.append(hostname)
        staged_serials.append(serial)
        added += 1

    update_yk_staged_display()
    msg = f"Import complete.\n\n✅ Added: {added}\n⛔ Skipped: {skipped}"
    if skip_log: msg += "\n\nSkipped details:\n" + "\n".join(skip_log[:20])
    if len(skip_log) > 20: msg += f"\n...and {len(skip_log)-20} more."
    (messagebox.showinfo if skipped == 0 else messagebox.showwarning)("Import Result", msg)

def yk_put_item():
    hostname   = yk_hostname_entry.get().strip()
    serial     = yk_serial_entry.get().strip()
    checked_by = yk_checked_by_entry.get().strip()
    shelf      = yk_shelf_var.get()
    status     = yk_status_var.get()
    remarks    = yk_remarks_var.get().strip()
    if not hostname:   messagebox.showerror("Error", "Please enter a Hostname"); return
    if not serial:     messagebox.showerror("Error", "Please enter a Serial Number"); return
    if not checked_by: messagebox.showerror("Error", "Please enter Checked By"); return
    if not shelf:      messagebox.showerror("Error", "Please select a Shelf"); return
    if not status:     messagebox.showerror("Error", "Please select a Status"); return
    df_yk = load_items_yk()
    df_shelves_yk = load_shelves_yk()
    if hostname in df_yk["Hostname"].values:
        messagebox.showerror("Error", "Hostname already exists in Yubikey warehouse"); return
    if any(i["Hostname"] == hostname for i in staged_yk_items):
        messagebox.showerror("Error", "Hostname already staged"); return
    if serial in df_yk["Serial Number"].astype(str).values:
        messagebox.showerror("Error", "Serial Number already exists in Yubikey warehouse"); return
    if any(i.get("Serial Number") == serial for i in staged_yk_items):
        messagebox.showerror("Error", "Serial Number already staged"); return
    shelf_status = df_shelves_yk[df_shelves_yk["Shelf"] == shelf]["Status"].values
    if len(shelf_status) > 0 and shelf_status[0] == "FULL":
        messagebox.showerror("Error", "Shelf is marked FULL"); return
    staged_yk_items.append({"Hostname": hostname, "Serial Number": serial,
                             "Checked By": checked_by, "Shelf": shelf, "Status": status, "Remarks": remarks})
    _yk_clear_fields()
    messagebox.showinfo("Staged", f"'{hostname}' added to Yubikey staging queue")
    update_yk_staged_display()

def yk_update_item():
    global selected_staged_yk_index
    new_hostname   = yk_hostname_entry.get().strip()
    new_serial     = yk_serial_entry.get().strip()
    new_checked_by = yk_checked_by_entry.get().strip()
    new_shelf      = yk_shelf_var.get()
    new_status     = yk_status_var.get()
    new_remarks    = yk_remarks_var.get().strip()
    if not new_hostname:   messagebox.showerror("Error", "Hostname cannot be empty"); return
    if not new_serial:     messagebox.showerror("Error", "Serial Number cannot be empty"); return
    if not new_checked_by: messagebox.showerror("Error", "Checked By cannot be empty"); return
    if not new_shelf:      messagebox.showerror("Error", "Please select a Shelf"); return
    if not new_status:     messagebox.showerror("Error", "Please select a Status"); return
    if selected_staged_yk_index is not None:
        idx = selected_staged_yk_index
        if idx >= len(staged_yk_items):
            messagebox.showerror("Error", "Invalid staged selection"); selected_staged_yk_index = None; return
        if any(i != idx and item["Hostname"] == new_hostname for i, item in enumerate(staged_yk_items)):
            messagebox.showerror("Error", "Hostname already exists in staging"); return
        staged_yk_items[idx].update({"Hostname": new_hostname, "Serial Number": new_serial,
                                     "Checked By": new_checked_by, "Shelf": new_shelf,
                                     "Status": new_status, "Remarks": new_remarks})
        save_log("UPDATE ITEM", f"[YK] Hostname: {new_hostname}")
        messagebox.showinfo("Updated", "Staged item updated")
        update_yk_staged_display()
        selected_staged_yk_index = None
        return
    messagebox.showerror("Update Not Allowed",
        "Items cannot be updated directly from the warehouse.\n\n"
        "Click a row to move it back to staging,\n"
        "then select it in the staged list and click UPDATE.")

def yk_remove_from_staging():
    global selected_staged_yk_index
    sel = yk_staged_listbox.curselection()
    if sel:
        idx = sel[0]
        if idx >= len(staged_yk_items):
            selected_staged_yk_index = None; return
        removed = staged_yk_items.pop(idx)
        selected_staged_yk_index = None
        _yk_clear_fields()
        update_yk_staged_display()
        messagebox.showinfo("Removed", f"'{removed['Hostname']}' removed from staging")
    else:
        if not staged_yk_items:
            messagebox.showinfo("Info", "No staged items to clear"); return
        if not messagebox.askyesno("Confirm", f"Clear all {len(staged_yk_items)} staged item(s)?"):
            return
        staged_yk_items.clear()
        selected_staged_yk_index = None
        _yk_clear_fields()
        update_yk_staged_display()
        messagebox.showinfo("Cleared", "All staged items cleared")

def yk_put_warehouse():
    if not staged_yk_items:
        messagebox.showerror("Error", "No staged items to put"); return
    if not messagebox.askyesno("Confirm", f"Put {len(staged_yk_items)} item(s) to Yubikey warehouse?"):
        return
    try:
        df_yk = load_items_yk()
        df_shelves_yk = load_shelves_yk()
        # Re-check shelf FULL at commit time
        for _item in staged_yk_items:
            _shelf_val = _item.get("Shelf", "")
            if _shelf_val:
                _s_stat = df_shelves_yk[df_shelves_yk["Shelf"] == _shelf_val]["Status"].values
                if len(_s_stat) > 0 and _s_stat[0] == "FULL":
                    messagebox.showerror("Error",
                        f"Shelf '{_shelf_val}' is marked FULL.\n"
                        f"Item '{_item['Hostname']}' cannot be placed there.\n"
                        "Edit or remove it from staging first.")
                    return
        now_str = _now().strftime("%Y-%m-%d %H:%M:%S")
        for item in staged_yk_items:
            qr_code = str(uuid.uuid4())
            generate_qr(item["Hostname"], item["Hostname"], warehouse=3)
            df_yk = pd.concat([df_yk, pd.DataFrame([{
                "QR":            qr_code,
                "Hostname":      item["Hostname"],
                "Serial Number": item.get("Serial Number",""),
                "Checked By":    item.get("Checked By",""),
                "Shelf":         item.get("Shelf",""),
                "Status":        item.get("Status",""),
                "Remarks":       item.get("Remarks",""),
                "Date":          now_str
            }])], ignore_index=True)
        save_warehouse_yk(df_yk)
        count = len(staged_yk_items)
        for item in staged_yk_items:
            save_log("PUT WAREHOUSE", f"[YK] Hostname: {item['Hostname']} | Shelf: {item.get('Shelf','')}")
        staged_yk_items.clear()
        messagebox.showinfo("Success",
            f"{count} item(s) added to Yubikey warehouse.\nQR codes generated.\n"
            "Use 'GENERATE FILES' to create PDF labels and export to Excel.")
        update_yk_staged_display()
        yk_refresh_all()
    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save:\n{str(e)}")

def yk_delete_item():
    selected = tree_yk_warehouse.selection()
    if not selected:
        messagebox.showerror("Error", "Select item"); return
    df_yk = load_items_yk()
    hostname = tree_yk_warehouse.item(selected[0], "values")[2]
    if not messagebox.askyesno("Confirm Delete",
            f"Move '{hostname}' to the dump folder?\n"
            "It will no longer appear in the warehouse but can be recovered manually."):
        return
    delete_qr(hostname, warehouse=3)
    row = df_yk[df_yk["Hostname"] == hostname]
    if not row.empty:
        import shutil
        dump_file  = os.path.join(DUMP_FOLDER, "dumped_records.xlsx")
        dump_sheet = "dumped_yk"
        try:
            if os.path.exists(dump_file):
                existing = pd.read_excel(dump_file, sheet_name=None)
                df_dump  = existing.get(dump_sheet, pd.DataFrame())
            else:
                df_dump = pd.DataFrame()
            row_copy = row.copy()
            row_copy["Dumped At"] = _now().strftime("%Y-%m-%d %H:%M:%S")
            df_dump = pd.concat([df_dump, row_copy], ignore_index=True)
            with pd.ExcelWriter(dump_file, engine="openpyxl",
                                mode="a" if os.path.exists(dump_file) else "w",
                                if_sheet_exists="replace") as writer:
                df_dump.to_excel(writer, sheet_name=dump_sheet, index=False)
        except Exception as e:
            messagebox.showwarning("Dump Warning", f"Record moved but dump log failed:\n{e}")
    df_yk = df_yk[df_yk["Hostname"] != hostname].reset_index(drop=True)
    save_warehouse_yk(df_yk)
    save_log("DELETE ITEM", f"[YK] Hostname: {hostname} | Dumped")
    messagebox.showinfo("Moved to Dump", f"'{hostname}' has been moved to the dump folder.")
    yk_refresh_all()

def yk_pull_search_live(event=None):
    keyword = yk_search_entry.get().strip().lower()
    if tree_yk_pullouts.winfo_ismapped():
        df = load_pullouts_yk()
        shelf_f   = yk_shelf_filter_var.get()
        status_f  = yk_status_filter_var.get()
        date_from = yk_date_from_var.get().strip()
        date_to   = yk_date_to_var.get().strip()
        if keyword:
            mask = False
            for col in ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"]:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_f:  df = df[df["Shelf"] == shelf_f]
        if status_f: df = df[df["Status"] == status_f]
        df = _filter_by_date(df, date_from, date_to)
        tree_yk_pullouts.delete(*tree_yk_pullouts.get_children())
        yk_pull_row_checks.clear()
        for _, row in df.iterrows():
            hostname = str(row.get("Hostname",""))
            checked  = hostname in yk_pull_persistent_checks
            iid = tree_yk_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c,"") for c in ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"])
            ))
            yk_pull_row_checks[iid] = checked
        active = bool(keyword or shelf_f or status_f or date_from or date_to)
        yk_search_label.config(text=f"{len(df)} match(es)" if active else "", fg="darkorange" if active else "blue")
    else:
        yk_show_warehouse()
        if keyword:
            df = load_items_yk()
            mask = False
            for col in ["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"]:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
            _yk_populate_warehouse_tree(df)
            yk_search_label.config(text=f"{len(df)} match(es)")
        else:
            yk_search_label.config(text="")

def yk_pull_item():
    reason = yk_pull_reason_var.get().strip()
    if not reason:
        messagebox.showerror("Error", "Please enter a pull reason"); return
    checked = [iid for iid, state in yk_row_checks.items() if state]
    if checked:
        target_iids = checked
    elif tree_yk_warehouse.selection():
        target_iids = list(tree_yk_warehouse.selection())
    else:
        hostname_input = yk_search_entry.get().strip()
        if not hostname_input:
            messagebox.showerror("Error", "Select or check item(s) to pull, or type a hostname"); return
        df_yk = load_items_yk()
        match = df_yk[df_yk["Hostname"] == hostname_input]
        if match.empty:
            match = df_yk[df_yk["Hostname"].astype(str).str.lower().str.contains(hostname_input.lower(), na=False)]
        if match.empty:
            messagebox.showerror("Error", f"'{hostname_input}' not found in Yubikey warehouse"); return
        if len(match) > 1:
            messagebox.showerror("Error", "Multiple matches. Be more specific."); return
        target_iids = None
        hostname = match.iloc[0]["Hostname"]
        item_row = match.iloc[0]
        if not messagebox.askyesno("Confirm Pull Out", f"Pull out '{hostname}'?\nReason: {reason}"):
            return
        df_yk = load_items_yk(); df_po_yk = load_pullouts_yk()
        pull_qr(hostname, warehouse=3)
        df_yk = df_yk[df_yk["Hostname"] != hostname].reset_index(drop=True)
        df_po_yk = pd.concat([df_po_yk, pd.DataFrame([{
            "QR":            str(item_row.get("QR","")),
            "Hostname":      hostname,
            "Serial Number": str(item_row.get("Serial Number","")),
            "Checked By":    str(item_row.get("Checked By","")),
            "Shelf":         str(item_row.get("Shelf","")),
            "Status":        str(item_row.get("Status","")),
            "Remarks":       str(item_row.get("Remarks","")),
            "Pull Reason":   reason,
            "Date":          _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_warehouse_yk(df_yk, None, df_po_yk)
        save_log("WAREHOUSE PULL", f"[YK] Hostname: {hostname} | Reason: {reason}")
        try:
            all_reasons = sorted(load_pullouts_yk()["Pull Reason"].dropna().unique().tolist())
            yk_pull_reason_entry["values"] = [""] + all_reasons
        except Exception: pass
        messagebox.showinfo("Success", f"'{hostname}' pulled out successfully")
        yk_search_entry.delete(0, tk.END)
        yk_pull_reason_var.set("")
        yk_refresh_all()
        return

    # Bulk pull
    hostnames = [tree_yk_warehouse.item(iid, "values")[2] for iid in target_iids]
    if not messagebox.askyesno("Confirm Pull Out",
            f"Pull out {len(hostnames)} item(s)?\n" + "\n".join(hostnames) + f"\n\nReason: {reason}"):
        return
    df_yk = load_items_yk(); df_po_yk = load_pullouts_yk()
    pulled = 0
    for hostname in hostnames:
        match = df_yk[df_yk["Hostname"] == hostname]
        if match.empty: continue
        item_row = match.iloc[0]
        pull_qr(hostname, warehouse=3)
        df_yk = df_yk[df_yk["Hostname"] != hostname].reset_index(drop=True)
        df_po_yk = pd.concat([df_po_yk, pd.DataFrame([{
            "QR":            str(item_row.get("QR","")),
            "Hostname":      hostname,
            "Serial Number": str(item_row.get("Serial Number","")),
            "Checked By":    str(item_row.get("Checked By","")),
            "Shelf":         str(item_row.get("Shelf","")),
            "Status":        str(item_row.get("Status","")),
            "Remarks":       str(item_row.get("Remarks","")),
            "Pull Reason":   reason,
            "Date":          _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        save_log("WAREHOUSE PULL", f"[YK] Hostname: {hostname} | Reason: {reason}")
        pulled += 1
    save_warehouse_yk(df_yk, None, df_po_yk)
    try:
        all_reasons = sorted(load_pullouts_yk()["Pull Reason"].dropna().unique().tolist())
        yk_pull_reason_entry["values"] = [""] + all_reasons
    except Exception: pass
    messagebox.showinfo("Success", f"{pulled} item(s) pulled out successfully")
    yk_search_entry.delete(0, tk.END)
    yk_pull_reason_var.set("")
    yk_refresh_all()

def yk_undo_pull(event=None):
    checked = [iid for iid, state in yk_pull_row_checks.items() if state]
    if not checked:
        sel = tree_yk_pullouts.selection()
        if sel: checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Warehouse", "Check at least one row in the Pull History table.")
        return
    preview = []
    for iid in checked:
        v = tree_yk_pullouts.item(iid, "values")
        if v: preview.append(f"  • {v[1]}  (S/N: {v[2]})")
    if not messagebox.askyesno("Undo Pull",
            f"Restore {len(preview)} item(s) back to Yubikey warehouse?\n\n" + "\n".join(preview)):
        return
    restored = 0
    for iid in checked:
        values = tree_yk_pullouts.item(iid, "values")
        if not values: continue
        # values: ☐(0), Hostname(1), Serial(2), Checked By(3), Shelf(4), Status(5), Remarks(6), PullReason(7), Date(8)
        hostname, serial, checked_by, shelf, status, remarks = values[1], values[2], values[3], values[4], values[5], values[6]
        df_yk    = load_items_yk()
        df_po_yk = load_pullouts_yk()
        if hostname in df_yk["Hostname"].values:
            messagebox.showerror("Error", f"'{hostname}' already exists in warehouse"); continue
        match = df_po_yk[df_po_yk["Hostname"] == hostname]
        if match.empty:
            messagebox.showerror("Error", f"'{hostname}' not found in pull history"); continue
        pull_row = match.iloc[0]
        import shutil
        pull_qr_file = pull_qr_path_for(hostname, warehouse=3)
        wh_qr_file   = qr_path_for(hostname, warehouse=3)
        qr_code = str(pull_row.get("QR",""))
        if os.path.exists(pull_qr_file):
            try:
                os.makedirs(QR_FOLDER_YK, exist_ok=True)
                shutil.move(pull_qr_file, wh_qr_file)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR file could not be moved back: {e}")
        elif not os.path.exists(wh_qr_file):
            try:
                qr_code = str(uuid.uuid4())
                generate_qr(hostname, qr_code, warehouse=3)
            except Exception as e:
                messagebox.showwarning("Warning", f"QR code not regenerated: {e}")
        df_yk = pd.concat([df_yk, pd.DataFrame([{
            "QR":            qr_code,
            "Hostname":      hostname,
            "Serial Number": serial,
            "Checked By":    checked_by,
            "Shelf":         shelf,
            "Status":        status,
            "Remarks":       remarks,
            "Date":          _now().strftime("%Y-%m-%d %H:%M:%S")
        }])], ignore_index=True)
        df_po_yk = df_po_yk[df_po_yk["Hostname"] != hostname].reset_index(drop=True)
        save_warehouse_yk(df_yk, None, df_po_yk)
        save_log("UNDO PULL", f"[YK] Hostname: {hostname}")
        restored += 1
    if restored:
        messagebox.showinfo("Restored", f"{restored} item(s) restored to Yubikey warehouse.")
    yk_show_pullouts()

def yk_unstage_from_warehouse(event=None):
    checked = [iid for iid, state in yk_row_checks.items() if state]
    if not checked:
        sel = tree_yk_warehouse.selection()
        if sel: checked = [sel[0]]
    if not checked:
        messagebox.showinfo("Back to Stage", "Check at least one row in the Warehouse table.")
        return
    preview = []
    for iid in checked:
        v = tree_yk_warehouse.item(iid, "values")
        if v: preview.append(f"  • {v[2]}  (S/N: {v[3]})")
    if not messagebox.askyesno("Move to Staging",
            f"Move {len(preview)} item(s) back to staging?\n\n" + "\n".join(preview)):
        return
    moved = 0
    for iid in checked:
        values = tree_yk_warehouse.item(iid, "values")
        if not values: continue
        # values: ☐(0), QR(1), Hostname(2), Serial(3), Checked By(4), Shelf(5), Status(6), Remarks(7), Date(8)
        hostname, serial, checked_by, shelf, status, remarks = values[2], values[3], values[4], values[5], values[6], values[7]
        if any(item["Hostname"] == hostname for item in staged_yk_items):
            messagebox.showerror("Error", f"'{hostname}' is already in staging"); continue
        df_yk = load_items_yk()
        remove_qr(hostname, warehouse=3)
        df_yk = df_yk[df_yk["Hostname"] != hostname].reset_index(drop=True)
        save_warehouse_yk(df_yk)
        staged_yk_items.append({"Hostname": hostname, "Serial Number": serial,
                                 "Checked By": checked_by, "Shelf": shelf, "Status": status, "Remarks": remarks})
        save_log("UNSTAGE", f"[YK] Hostname: {hostname}")
        moved += 1
    if moved:
        messagebox.showinfo("Moved", f"{moved} item(s) moved back to staging.")
        update_yk_staged_display()
        yk_refresh_all()

def yk_show_warehouse():
    try: yk_back_to_wh_btn.pack_forget()
    except Exception: pass
    try: yk_back_to_stage_btn.pack(side="left", padx=(0,6))
    except Exception: pass
    yk_update_full_shelves_display()
    df_yk = load_items_yk()
    if "Date" not in df_yk.columns: df_yk["Date"] = ""
    try:
        date_from = yk_date_from_var.get().strip()
        date_to   = yk_date_to_var.get().strip()
        df_yk = _filter_by_date(df_yk, date_from, date_to)
    except Exception: pass
    _yk_populate_warehouse_tree(df_yk)
    yk_search_label.config(text="", fg="blue")

def yk_show_pullouts():
    try: yk_back_to_stage_btn.pack_forget()
    except Exception: pass
    try: yk_back_to_wh_btn.pack(side="left", padx=(0,6))
    except Exception: pass
    _yk_show_tree(tree_yk_pullouts)
    tree_yk_pullouts.delete(*tree_yk_pullouts.get_children())
    df_po = load_pullouts_yk()
    try:
        shelf_f   = yk_shelf_filter_var.get()
        status_f  = yk_status_filter_var.get()
        date_from = yk_date_from_var.get().strip()
        date_to   = yk_date_to_var.get().strip()
    except Exception:
        shelf_f = status_f = date_from = date_to = ""
    if shelf_f:  df_po = df_po[df_po["Shelf"] == shelf_f]
    if status_f: df_po = df_po[df_po["Status"] == status_f]
    df_po = _filter_by_date(df_po, date_from, date_to)
    yk_pull_row_checks.clear()
    for _, row in df_po.iterrows():
        hostname = str(row.get("Hostname",""))
        checked  = hostname in yk_pull_persistent_checks
        iid = tree_yk_pullouts.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c,"") for c in ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"])
        ))
        yk_pull_row_checks[iid] = checked
    try:
        all_reasons = sorted(load_pullouts_yk()["Pull Reason"].dropna().unique().tolist())
        yk_pull_reason_entry["values"] = [""] + all_reasons
    except Exception: pass
    yk_search_label.config(text="", fg="blue")

def _yk_show_tree(tree):
    for t in (tree_yk_warehouse, tree_yk_pullouts, tree_yk_available):
        if t is not tree: t.pack_forget()
    tree.pack(fill="both", expand=True)

def yk_show_available():
    try: yk_back_to_stage_btn.pack_forget()
    except Exception: pass
    try: yk_back_to_wh_btn.pack_forget()
    except Exception: pass
    _yk_show_tree(tree_yk_available)
    tree_yk_available.delete(*tree_yk_available.get_children())
    df_shelves = load_shelves_yk().sort_values("Shelf")
    df_items   = load_items_yk()
    for _, row in df_shelves.iterrows():
        shelf_name = row["Shelf"]
        item_count = int((df_items["Shelf"] == shelf_name).sum()) if "Shelf" in df_items.columns else 0
        date_full  = row.get("Date_Full", "")
        tree_yk_available.insert("", "end", values=(
            shelf_name, row["Status"], item_count,
            date_full if pd.notna(date_full) else ""
        ))
    yk_search_label.config(text="", fg="blue")

def _yk_populate_warehouse_tree(df):
    _yk_show_tree(tree_yk_warehouse)
    tree_yk_warehouse.delete(*tree_yk_warehouse.get_children())
    yk_row_checks.clear()
    for _, row in df.iterrows():
        hostname = str(row.get("Hostname",""))
        checked  = hostname in yk_persistent_checks
        iid = tree_yk_warehouse.insert("", "end", values=(
            "☑" if checked else "☐",
            *tuple(row.get(c,"") for c in ["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"])
        ))
        yk_row_checks[iid] = checked
    _yk_refresh_select_all_label()

def yk_search_item():
    keyword   = yk_search_entry.get().strip().lower()
    shelf_f   = yk_shelf_filter_var.get()
    status_f  = yk_status_filter_var.get()
    date_from = yk_date_from_var.get().strip()
    date_to   = yk_date_to_var.get().strip()
    if tree_yk_pullouts.winfo_ismapped():
        df = load_pullouts_yk()
        if keyword:
            mask = False
            for col in ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"]:
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
            df = df[mask]
        if shelf_f:  df = df[df["Shelf"] == shelf_f]
        if status_f: df = df[df["Status"] == status_f]
        df = _filter_by_date(df, date_from, date_to)
        tree_yk_pullouts.delete(*tree_yk_pullouts.get_children())
        yk_pull_row_checks.clear()
        for _, row in df.iterrows():
            hostname = str(row.get("Hostname",""))
            checked  = hostname in yk_pull_persistent_checks
            iid = tree_yk_pullouts.insert("", "end", values=(
                "☑" if checked else "☐",
                *tuple(row.get(c,"") for c in ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"])
            ))
            yk_pull_row_checks[iid] = checked
        parts = []
        if keyword:   parts.append(f"Search: \"{keyword}\"")
        if shelf_f:   parts.append(f"Shelf: {shelf_f}")
        if status_f:  parts.append(f"Status: {status_f}")
        if date_from: parts.append(f"From: {date_from}")
        if date_to:   parts.append(f"To: {date_to}")
        label = f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")
        yk_search_label.config(text=label if parts else "", fg="darkorange" if parts else "blue")
        return
    df = load_items_yk()
    if keyword:
        mask = False
        for col in ["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"]:
            if col in df.columns:
                mask = mask | df[col].astype(str).str.lower().str.contains(keyword, na=False)
        df = df[mask]
    if shelf_f:  df = df[df["Shelf"] == shelf_f]
    if status_f: df = df[df["Status"] == status_f]
    df = _filter_by_date(df, date_from, date_to)
    _yk_populate_warehouse_tree(df)
    parts = []
    if keyword:   parts.append(f"Search: \"{keyword}\"")
    if shelf_f:   parts.append(f"Shelf: {shelf_f}")
    if status_f:  parts.append(f"Status: {status_f}")
    if date_from: parts.append(f"From: {date_from}")
    if date_to:   parts.append(f"To: {date_to}")
    label = f"{len(df)} result(s)" + (" — " + " | ".join(parts) if parts else "")
    yk_search_label.config(text=label if parts else "")
    if parts:
        save_log("SEARCH", f"[YK] {' | '.join(parts)} → {len(df)} result(s)")

def yk_select_item(event):
    selected = tree_yk_warehouse.selection()
    if selected:
        iid = selected[0]
        values = tree_yk_warehouse.item(iid, "values")
        # values: ☐(0), QR(1), Hostname(2), Serial(3), Checked By(4), Shelf(5), Status(6), Remarks(7), Date(8)
        if iid in yk_row_checks:
            yk_row_checks[iid] = not yk_row_checks[iid]
            hostname = str(values[2])
            if yk_row_checks[iid]: yk_persistent_checks.add(hostname)
            else:                  yk_persistent_checks.discard(hostname)
            tree_yk_warehouse.set(iid, "C0", "☑" if yk_row_checks[iid] else "☐")
            _yk_refresh_select_all_label()
        yk_search_entry.delete(0, tk.END)
        yk_search_entry.insert(0, values[2])
        yk_status_label.config(text=f"Selected → Hostname: {values[2]}  |  S/N: {values[3]}  |  Shelf: {values[5]}", fg="#1a5276")

def yk_select_pull_item(event):
    selected = tree_yk_pullouts.selection()
    if selected:
        iid = selected[0]
        if iid in yk_pull_row_checks:
            yk_pull_row_checks[iid] = not yk_pull_row_checks[iid]
            values   = tree_yk_pullouts.item(iid, "values")
            hostname = str(values[1])
            if yk_pull_row_checks[iid]: yk_pull_persistent_checks.add(hostname)
            else:                       yk_pull_persistent_checks.discard(hostname)
            tree_yk_pullouts.set(iid, "CP0", "☑" if yk_pull_row_checks[iid] else "☐")

def yk_reset_ui():
    _yk_clear_fields()
    for s in tree_yk_warehouse.selection(): tree_yk_warehouse.selection_remove(s)
    yk_status_label.config(text="")
    yk_search_label.config(text="")
    yk_persistent_checks.clear()
    yk_show_warehouse()

def yk_reset_pull_out():
    yk_search_entry.delete(0, tk.END)
    yk_pull_reason_var.set("")
    yk_pull_date_from_var.set("")
    yk_pull_date_to_var.set("")
    for s in tree_yk_warehouse.selection(): tree_yk_warehouse.selection_remove(s)
    yk_status_label.config(text="")
    yk_search_label.config(text="")
    yk_show_warehouse()

def yk_clear_filters():
    yk_shelf_filter_var.set("")
    yk_status_filter_var.set("")
    yk_search_entry.delete(0, tk.END)
    yk_pull_reason_var.set("")
    yk_pull_date_from_var.set("")
    yk_pull_date_to_var.set("")
    yk_date_from_var.set("")
    yk_date_to_var.set("")
    yk_search_label.config(text="")
    yk_persistent_checks.clear()
    yk_pull_persistent_checks.clear()
    yk_show_warehouse()

def yk_refresh_all():
    yk_show_warehouse()
    update_all_shelf_dropdowns()

# ========== YK SHELF MANAGEMENT ==========

def yk_set_shelf_status(new_status):
    shelf = yk_shelf_control_var.get()
    if not shelf:
        messagebox.showerror("Error", "Select a shelf"); return
    df_shelves_yk = load_shelves_yk()
    idx = df_shelves_yk[df_shelves_yk["Shelf"] == shelf].index
    if len(idx) == 0:
        return
    df_shelves_yk.at[idx[0], "Status"] = new_status
    df_shelves_yk.at[idx[0], "Date_Full"] = _now().strftime("%Y-%m-%d %H:%M:%S") if new_status == "FULL" else None
    save_warehouse_yk(load_items_yk(), df_shelves_yk)
    save_log("SHELF STATUS", f"[YK] Shelf: {shelf} → {new_status}")
    yk_status_label.config(text=f"{shelf} → {new_status}")
    yk_refresh_all()

def yk_add_shelf():
    new_shelf = yk_remove_shelf_var.get().strip()
    if not new_shelf:
        messagebox.showerror("Error", "Enter shelf name"); return
    df_shelves_yk = load_shelves_yk()
    if new_shelf in df_shelves_yk["Shelf"].values:
        messagebox.showerror("Error", "Shelf already exists in Yubikey"); return
    df_shelves_yk = pd.concat([df_shelves_yk, pd.DataFrame([{"Shelf": new_shelf, "Status": "AVAILABLE"}])], ignore_index=True)
    df_shelves_yk = df_shelves_yk.sort_values("Shelf", ignore_index=True)
    save_warehouse_yk(load_items_yk(), df_shelves_yk)
    messagebox.showinfo("Success", f"Shelf '{new_shelf}' added to Yubikey")
    save_log("ADD SHELF", f"[YK] Shelf: {new_shelf}")
    yk_remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def yk_remove_shelf():
    shelf_name = yk_remove_shelf_var.get().strip()
    if not shelf_name:
        messagebox.showerror("Error", "Select a shelf to remove"); return
    df_items_yk   = load_items_yk()
    df_shelves_yk = load_shelves_yk()
    if not df_items_yk[df_items_yk["Shelf"] == shelf_name].empty:
        messagebox.showerror("Error", f"Cannot remove shelf '{shelf_name}' — it still has items"); return
    if shelf_name not in df_shelves_yk["Shelf"].values:
        messagebox.showerror("Error", f"Shelf '{shelf_name}' does not exist in Yubikey"); return
    df_shelves_yk = df_shelves_yk[df_shelves_yk["Shelf"] != shelf_name].sort_values("Shelf", ignore_index=True)
    save_warehouse_yk(df_items_yk, df_shelves_yk)
    messagebox.showinfo("Success", f"Shelf '{shelf_name}' removed from Yubikey")
    save_log("REMOVE SHELF", f"[YK] Shelf: {shelf_name}")
    yk_remove_shelf_var.set("")
    update_all_shelf_dropdowns()

def yk_reset_shelf_control():
    yk_shelf_control_var.set("")
    yk_status_label.config(text="")

def yk_reset_shelf_addition():
    yk_remove_shelf_var.set("")
    yk_status_label.config(text="")

def yk_update_full_shelves_display():
    df_shelves = load_shelves_yk()
    full_shelves = df_shelves[df_shelves["Status"] == "FULL"]["Shelf"].tolist()
    yk_full_label.config(text="FULL Shelves:\n" + "\n".join(full_shelves) if full_shelves else "FULL Shelves: None")

def _yk_refresh_select_all_label():
    try:
        all_iids = list(tree_yk_warehouse.get_children())
        checked  = [iid for iid in all_iids if yk_row_checks.get(iid)]
        yk_select_all_btn.config(
            text="DESELECT ALL" if all_iids and len(checked) == len(all_iids) else "SELECT ALL")
    except (NameError, tk.TclError, AttributeError):
        pass

def _get_yk_selected_rows():
    checked = [iid for iid, state in yk_row_checks.items() if state]
    if checked:
        return [tree_yk_warehouse.item(iid, "values") for iid in checked]
    return [tree_yk_warehouse.item(iid, "values") for iid in tree_yk_warehouse.get_children()]

def yk_generate_stored_qr():
    """Generate Files (QR + PDF + Excel) or export pull history for Yubikey."""
    if tree_yk_pullouts.winfo_ismapped():
        _yk_export_pull_history(); return
    rows = _get_yk_selected_rows()
    if not rows:
        messagebox.showinfo("Generate Files", "No items to generate QR codes for."); return
    # Preview
    preview_lines = [f"  • {v[2]}" for v in rows]
    if not messagebox.askyesno("Confirm — Generate Files",
            f"Generate files for {len(rows)} item(s):\n\n" + "\n".join(preview_lines) + "\n\nProceed?"):
        return
    # Generate QR PNGs
    count_ok = count_skip = 0
    qr_keys = []
    for values in rows:
        try:
            hostname = str(values[2])
            generate_qr(hostname, hostname, warehouse=3)
            qr_keys.append(hostname)
            count_ok += 1
        except Exception: count_skip += 1
    if count_skip:
        messagebox.showwarning("Generate Files", f"{count_ok} QR code(s) generated.\n{count_skip} skipped.")
    # Write selection sheet
    try:
        cols  = ["QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"]
        # values: ☐(0),QR(1),Hostname(2),Serial(3),CheckedBy(4),Shelf(5),Status(6),Remarks(7),Date(8)
        records = [{c: v for c, v in zip(cols, [values[1],values[2],values[3],values[4],values[5],values[6],values[7],values[8]])} for values in rows]
        df_sel  = pd.DataFrame(records, columns=cols)
        from openpyxl import load_workbook
        wb = load_workbook(FILE)
        sname = "qr_selection_yk"
        if sname in wb.sheetnames: del wb[sname]
        ws = wb.create_sheet(sname)
        ws.append(cols)
        for rec in df_sel.itertuples(index=False): ws.append(list(rec))
        ws.protection.sheet = True; ws.protection.enable()
        wb.save(FILE)
    except Exception as e:
        messagebox.showwarning("Generate Files", f"QR codes generated but selection sheet failed:\n{e}")
    # PDF + Excel dialog
    _yk_do_generate_files(rows, qr_keys)
    _yk_open_qr_gallery(filter_keys=qr_keys)

def _yk_do_generate_files(rows, qr_keys):
    existing_pdfs = sorted(
        [os.path.splitext(f)[0] for f in os.listdir(QR_LABELS_FOLDER_YK) if f.lower().endswith(".pdf")],
        reverse=True) if os.path.exists(QR_LABELS_FOLDER_YK) else []
    os.makedirs(EXCEL_FOLDER_YK, exist_ok=True)
    existing_excels = sorted(
        [os.path.splitext(f)[0] for f in os.listdir(EXCEL_FOLDER_YK) if f.lower().endswith(".xlsx")],
        reverse=True)
    name_win = tk.Toplevel(root)
    name_win.title("Generate Files — Name Your Export")
    name_win.resizable(False, False); name_win.transient(root); name_win.grab_set()
    tk.Label(name_win, text="Generate PDF Labels & Excel Export",
             font=("Helvetica",10,"bold"), bg="#6c3483", fg="white", padx=10, pady=6).pack(fill="x")
    form = tk.Frame(name_win, padx=16, pady=12); form.pack()
    def _section(text, row):
        tk.Label(form, text=text, font=("Helvetica",8,"bold"), fg="#6c3483", anchor="w").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(10,2))
    tk.Label(form, text="Generate:", anchor="w", width=14).grid(row=0, column=0, sticky="w", pady=(0,4))
    gen_pdf_var = tk.BooleanVar(value=True); gen_excel_var = tk.BooleanVar(value=True)
    chk_f = tk.Frame(form); chk_f.grid(row=0, column=1, columnspan=2, sticky="w", pady=(0,4))
    pdf_chk   = tk.Checkbutton(chk_f, text="PDF Labels",   variable=gen_pdf_var);   pdf_chk.pack(side="left", padx=(0,12))
    excel_chk = tk.Checkbutton(chk_f, text="Excel Export", variable=gen_excel_var); excel_chk.pack(side="left")
    _section("▸ PDF Label File", 1)
    pdf_name_lbl = tk.Label(form, text="File Name:", anchor="w", width=14); pdf_name_lbl.grid(row=2, column=0, sticky="w", pady=3)
    pdf_name_var = tk.StringVar(value=""); pdf_name_cb = ttk.Combobox(form, textvariable=pdf_name_var, width=28, values=existing_pdfs)
    pdf_name_cb.grid(row=2, column=1, pady=3, padx=(4,0))
    tk.Label(form, text=".pdf", fg="gray", font=("Helvetica",8)).grid(row=2, column=2, sticky="w", padx=(3,0))
    pdf_hint = tk.Label(form, text="  ↳ Select existing PDF to append pages to, or type a new name", fg="gray", font=("Helvetica",7))
    pdf_hint.grid(row=3, column=1, columnspan=2, sticky="w")
    _section("▸ Excel File", 4)
    excel_name_lbl = tk.Label(form, text="File Name:", anchor="w", width=14); excel_name_lbl.grid(row=5, column=0, sticky="w", pady=3)
    file_name_var = tk.StringVar(value=""); file_name_cb = ttk.Combobox(form, textvariable=file_name_var, width=28, values=existing_excels)
    file_name_cb.grid(row=5, column=1, pady=3, padx=(4,0))
    tk.Label(form, text=".xlsx", fg="gray", font=("Helvetica",8)).grid(row=5, column=2, sticky="w", padx=(3,0))
    excel_hint = tk.Label(form, text="  ↳ Select existing Excel to append a sheet to, or type a new name", fg="gray", font=("Helvetica",7))
    excel_hint.grid(row=6, column=1, columnspan=2, sticky="w")
    _section("▸ Excel Sheet", 7)
    sheet_name_lbl = tk.Label(form, text="Sheet Name:", anchor="w", width=14); sheet_name_lbl.grid(row=8, column=0, sticky="w", pady=3)
    sheet_name_var = tk.StringVar(value=""); sheet_name_cb = ttk.Combobox(form, textvariable=sheet_name_var, width=28)
    sheet_name_cb.grid(row=8, column=1, pady=3, padx=(4,0))
    tk.Label(form, text="  ↳ New sheet name to add (existing sheet of same name will be replaced)", fg="gray", font=("Helvetica",7)).grid(row=9, column=1, columnspan=2, sticky="w")
    def _refresh_sheet_list(*_):
        chosen = file_name_var.get().strip()
        if not chosen: sheet_name_cb["values"] = []; return
        safe = chosen if chosen.lower().endswith(".xlsx") else chosen + ".xlsx"
        path = os.path.join(EXCEL_FOLDER_YK, safe.replace("/","-").replace("\\","-"))
        if os.path.exists(path):
            try:
                from openpyxl import load_workbook
                wb_p = load_workbook(path, read_only=True); sheet_name_cb["values"] = wb_p.sheetnames; wb_p.close(); return
            except Exception: pass
        sheet_name_cb["values"] = []
    file_name_cb.bind("<<ComboboxSelected>>", _refresh_sheet_list); file_name_var.trace_add("write", _refresh_sheet_list)
    def _tog_pdf(*_):
        s = "normal" if gen_pdf_var.get() else "disabled"; pdf_name_cb.config(state=s); pdf_name_lbl.config(fg="black" if gen_pdf_var.get() else "gray")
    def _tog_xl(*_):
        s = "normal" if gen_excel_var.get() else "disabled"
        file_name_cb.config(state=s); sheet_name_cb.config(state=s)
        excel_name_lbl.config(fg="black" if gen_excel_var.get() else "gray"); sheet_name_lbl.config(fg="black" if gen_excel_var.get() else "gray")
    gen_pdf_var.trace_add("write", _tog_pdf); gen_excel_var.trace_add("write", _tog_xl)
    tk.Label(form, text="(Files saved to excel_exports/yubikey/)", fg="gray", font=("Helvetica",8)).grid(row=10, column=0, columnspan=3, sticky="w", pady=(8,0))
    error_lbl = tk.Label(form, text="", fg="red", font=("Helvetica",8)); error_lbl.grid(row=11, column=0, columnspan=3, sticky="w", pady=(4,0))
    confirmed = [False]
    def on_confirm():
        if not gen_pdf_var.get() and not gen_excel_var.get(): error_lbl.config(text="Please select at least one file type."); return
        if gen_pdf_var.get() and not pdf_name_var.get().strip(): error_lbl.config(text="Please enter a PDF file name."); return
        if gen_excel_var.get() and not file_name_var.get().strip(): error_lbl.config(text="Please enter an Excel file name."); return
        if gen_excel_var.get() and not sheet_name_var.get().strip(): error_lbl.config(text="Please enter a sheet name."); return
        confirmed[0] = True; name_win.destroy()
    btn_r = tk.Frame(name_win, pady=10); btn_r.pack()
    tk.Button(btn_r, text="GENERATE", command=on_confirm, bg="#6c3483", fg="white", width=12).pack(side="left", padx=6)
    tk.Button(btn_r, text="Cancel", command=name_win.destroy, width=10).pack(side="left", padx=6)
    name_win.update_idletasks()
    px,py,pw,ph = root.winfo_rootx(),root.winfo_rooty(),root.winfo_width(),root.winfo_height()
    nw,nh = name_win.winfo_reqwidth(),name_win.winfo_reqheight()
    name_win.geometry(f"+{px+(pw-nw)//2}+{py+(ph-nh)//2}"); name_win.focus_force()
    root.wait_window(name_win)
    if not confirmed[0]: return
    pdf_name_str  = pdf_name_var.get().strip()
    file_name_str = file_name_var.get().strip()
    sheet_str     = sheet_name_var.get().strip()

    # ── Check for already-generated items (mirrors W1/W2 behaviour) ──
    already_in_pdf   = []
    already_in_excel = []
    import json
    safe_pdf = pdf_name_str.replace(" ", "_").replace("/", "-").replace("\\", "-")
    if not safe_pdf.lower().endswith(".pdf"): safe_pdf += ".pdf"
    sidecar_path = os.path.join(QR_LABELS_FOLDER_YK, safe_pdf + ".keys.json")
    if os.path.exists(sidecar_path):
        try:
            with open(sidecar_path, "r") as kf:
                existing_keys = set(json.load(kf).get("keys", []))
            for values in rows:
                key   = str(values[2])
                label = f"  • {values[2]}"
                if key in existing_keys:
                    already_in_pdf.append(label)
        except Exception:
            pass
    safe_xl = file_name_str.replace("/", "-").replace("\\", "-")
    if not safe_xl.lower().endswith(".xlsx"): safe_xl += ".xlsx"
    excel_check_path = os.path.join(EXCEL_FOLDER_YK, safe_xl)
    if os.path.exists(excel_check_path):
        try:
            from openpyxl import load_workbook as _lw
            wb_chk = _lw(excel_check_path, data_only=True)
            for ws_p in wb_chk.worksheets: ws_p.protection.sheet = False
            if sheet_str in wb_chk.sheetnames:
                ws_chk = wb_chk[sheet_str]
                existing_xl_keys = set()
                for xl_row in ws_chk.iter_rows(min_row=2, values_only=True):
                    if xl_row and xl_row[0] is not None:
                        existing_xl_keys.add((str(xl_row[0]), str(xl_row[1])))
                for values in rows:
                    key   = (str(values[2]), str(values[3]))
                    label = f"  • {values[2]}"
                    if key in existing_xl_keys:
                        already_in_excel.append(label)
            wb_chk.close()
        except Exception:
            pass
    if already_in_pdf or already_in_excel:
        parts = []
        if already_in_pdf:
            parts.append(
                f"Already in PDF '{pdf_name_str}'  ({len(already_in_pdf)} item(s)):\n"
                + "\n".join(already_in_pdf)
            )
        if already_in_excel:
            parts.append(
                f"Already in Excel '{file_name_str}' / sheet '{sheet_str}'  ({len(already_in_excel)} item(s)):\n"
                + "\n".join(already_in_excel)
            )
        msg = (
            "Some selected items were already generated before:\n\n"
            + "\n\n".join(parts)
            + "\n\nDo you want to continue? (duplicates will be skipped)"
        )
        if not messagebox.askyesno("Already Generated", msg):
            return

    save_log("GENERATE FILES", f"[YK] {len(rows)} item(s)")
    # PDF
    pdf_msg = ""
    if gen_pdf_var.get():
        try:
            pdf_items = [{"Hostname": str(v[2]), "Serial Number": str(v[3]),
                          "Checked By": str(v[4]), "_warehouse": 3} for v in rows]
            _yk_pdf_name = pdf_name_str.replace(" ", "_")
            if not _yk_pdf_name.lower().endswith(".pdf"): _yk_pdf_name += ".pdf"
            _yk_pdf_existed = os.path.exists(os.path.join(QR_LABELS_FOLDER_YK, _yk_pdf_name))
            pdf_path = _yk_generate_qr_pdf(pdf_items, custom_name=pdf_name_str)
            pdf_msg  = f"PDF saved to:\n{pdf_path}"
            save_log("FILE UPDATED" if _yk_pdf_existed else "FILE CREATED",
                     f"[YK] PDF: {os.path.basename(pdf_path)} | Path: {pdf_path}")
        except Exception as e: pdf_msg = f"PDF generation failed: {e}"
    else: pdf_msg = "PDF skipped."
    # Excel
    excel_msg = ""
    if gen_excel_var.get():
        try:
            import stat
            from openpyxl import load_workbook
            from openpyxl import Workbook as _WB
            safe_fname = file_name_str.replace("/","-").replace("\\","-")
            if not safe_fname.lower().endswith(".xlsx"): safe_fname += ".xlsx"
            excel_path = os.path.join(EXCEL_FOLDER_YK, safe_fname)
            cols_xl   = ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"]
            # values: ☐(0),QR(1),Hostname(2),Serial(3),CB(4),Shelf(5),Status(6),Remarks(7),Date(8)
            records_xl = [[v[2],v[3],v[4],v[5],v[6],v[7],v[8]] for v in rows]
            gen_at = _now().strftime("%Y-%m-%d %H:%M:%S")
            header = cols_xl
            if os.path.exists(excel_path):
                try:
                    import ctypes; ctypes.windll.kernel32.SetFileAttributesW(excel_path, 0x80)
                except Exception: pass
                os.chmod(excel_path, stat.S_IWRITE | stat.S_IREAD)
                wb_r = load_workbook(excel_path, read_only=True)
                ex_sheets = {sn: [list(r) for r in wb_r[sn].iter_rows(values_only=True)] for sn in wb_r.sheetnames}
                wb_r.close()
                wb_xl = _WB(); wb_xl.remove(wb_xl.active)
                for sn, srows in ex_sheets.items():
                    ws_p = wb_xl.create_sheet(sn)
                    for r in srows: ws_p.append([v if v is not None else "" for v in r])
                if sheet_str in ex_sheets:
                    ws_xl = wb_xl[sheet_str]
                    ex_keys = set()
                    for r in ex_sheets[sheet_str][1:]:
                        if r and len(r) > 2 and r[1] is not None: ex_keys.add((str(r[1]), str(r[2])))
                    for rec in records_xl:
                        if (str(rec[0]), str(rec[1])) not in ex_keys: ws_xl.append([str(v) for v in rec])
                else:
                    ws_xl = wb_xl.create_sheet(sheet_str); ws_xl.append(header)
                    for rec in records_xl: ws_xl.append([str(v) for v in rec])
                for ws_p in wb_xl.worksheets: ws_p.protection.sheet=True; ws_p.protection.enable()
                wb_xl.save(excel_path); wb_xl.close()
            else:
                wb_xl = _WB(); ws_xl = wb_xl.active; ws_xl.title = sheet_str
                ws_xl.append(header)
                for rec in records_xl: ws_xl.append([str(v) for v in rec])
                for ws_p in wb_xl.worksheets: ws_p.protection.sheet=True; ws_p.protection.enable()
                wb_xl.save(excel_path); wb_xl.close()
            try:
                import ctypes; ctypes.windll.kernel32.SetFileAttributesW(excel_path, 0x01)
            except Exception: pass
            os.chmod(excel_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)
            excel_msg = f"Excel saved to:\n{excel_path}"
            _last_excel_path[3] = excel_path
            _yk_xl_action = "FILE UPDATED" if os.path.exists(excel_path) else "FILE CREATED"
            save_log(_yk_xl_action, f"[YK] Excel: {os.path.basename(excel_path)} | Sheet: {sheet_str} | Path: {excel_path}")
        except Exception as e: excel_msg = f"Excel export failed: {e}"
    else: excel_msg = "Excel skipped."
    messagebox.showinfo("Generate Files", f"{len(rows)} QR code(s) processed.\n\n{pdf_msg}\n\n{excel_msg}")

def _yk_generate_qr_pdf(items_batch, custom_name=None):
    """Generate a PDF label sheet for Yubikey items (same layout as W1/W2)."""
    import json
    from fpdf import FPDF
    if not items_batch: return None
    os.makedirs(QR_LABELS_FOLDER_YK, exist_ok=True)
    if custom_name:
        safe_name = custom_name.replace(" ","_").replace("/","-").replace("\\","-")
        if safe_name.lower().endswith(".pdf"): safe_name = safe_name[:-4]
    else:
        existing = [f for f in os.listdir(QR_LABELS_FOLDER_YK) if f.startswith("BATCH_") and f.endswith(".pdf")]
        batch_nums = []
        for f in existing:
            try: batch_nums.append(int(f.split("_")[1]))
            except: pass
        safe_name = f"BATCH_{max(batch_nums, default=0)+1}"
    pdf_path   = os.path.join(QR_LABELS_FOLDER_YK, f"{safe_name}.pdf")
    index_path = pdf_path + ".keys.json"
    existing_items = []; existing_keys = set()
    if os.path.exists(index_path) and os.path.exists(pdf_path):
        try:
            with open(index_path,"r") as kf: sidecar = json.load(kf)
            existing_items = sidecar.get("items",[]); existing_keys = set(sidecar.get("keys",[]))
        except Exception: pass
    def _item_key(it): return str(it.get("Hostname",""))
    new_items = [it for it in items_batch if _item_key(it) not in existing_keys]
    if not new_items: return pdf_path
    all_items = existing_items + new_items
    if os.path.exists(pdf_path):
        try:
            import stat; os.chmod(pdf_path, stat.S_IWRITE | stat.S_IREAD)
        except Exception: pass
    LABEL_W=80; LABEL_H=24; QR_SIZE=20; QR_PAD=2; COLS=2; MARGIN_X=10; MARGIN_Y=10; GAP_X=10; GAP_Y=4
    TEXT_X_OFF=QR_SIZE+QR_PAD*2+2; TEXT_W=LABEL_W-TEXT_X_OFF-2; LINE_H=4.8
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=False); pdf.add_page()
    col = row = 0
    for item in all_items:
        fields = [("Hostname:",   str(item.get("Hostname",""))),
                  ("Serial No:",  str(item.get("Serial Number",""))),
                  ("Checked By:", str(item.get("Checked By",""))),
                  ("Date:",       str(item.get("_date", _now().strftime("%Y-%m-%d"))))]
        path = qr_path_for(item.get("Hostname",""), warehouse=3)
        x = MARGIN_X + col*(LABEL_W+GAP_X); y = MARGIN_Y + row*(LABEL_H+GAP_Y)
        if y+LABEL_H > 210-MARGIN_Y: pdf.add_page(); col=row=0; x=MARGIN_X; y=MARGIN_Y
        pdf.set_draw_color(150,150,150); pdf.rect(x,y,LABEL_W,LABEL_H)
        if os.path.exists(path): pdf.image(path, x=x+QR_PAD, y=y+(LABEL_H-QR_SIZE)/2, w=QR_SIZE, h=QR_SIZE)
        total_h = len(fields)*LINE_H; text_y = y+(LABEL_H-total_h)/2; label_x = x+TEXT_X_OFF
        for lbl, val in fields:
            pdf.set_font("Helvetica","B",5.5); pdf.set_xy(label_x, text_y)
            pdf.cell(TEXT_W, LINE_H, f"{lbl} {val[:28]}", ln=0); text_y += LINE_H
        col += 1
        if col >= COLS: col=0; row+=1
    pdf.output(pdf_path)
    today = _now().strftime("%Y-%m-%d")
    for it in new_items: it.setdefault("_date", today)
    all_keys = list(existing_keys | {_item_key(it) for it in new_items})
    try:
        with open(index_path,"w") as kf: json.dump({"keys": all_keys, "items": all_items}, kf)
    except Exception: pass
    try:
        import stat; os.chmod(pdf_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)
    except Exception: pass
    return pdf_path

def _yk_export_pull_history():
    """Export Yubikey pull history rows to pull_excel/yubikey/."""
    checked = [iid for iid, state in yk_pull_row_checks.items() if state]
    iids    = checked if checked else list(tree_yk_pullouts.get_children())
    if not iids: messagebox.showinfo("Export Pull History", "No pull history rows to export."); return
    rows = [tree_yk_pullouts.item(iid,"values") for iid in iids]
    # values: ☐(0),Hostname(1),Serial(2),CheckedBy(3),Shelf(4),Status(5),Remarks(6),PullReason(7),Date(8)
    preview_lines = [f"  • {v[1]}" for v in rows]
    if not messagebox.askyesno("Confirm — Export Pull History",
            f"Export {len(rows)} pull history row(s):\n\n" + "\n".join(preview_lines) + "\n\nProceed?"):
        return
    cols = ["Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"]
    col_indices = [1,2,3,4,5,6,7,8]
    records = [{c: v[i] for c,i in zip(cols, col_indices)} for v in rows]
    os.makedirs(PULL_EXCEL_FOLDER_YK, exist_ok=True)
    existing_excels = sorted([f for f in os.listdir(PULL_EXCEL_FOLDER_YK) if f.lower().endswith(".xlsx")], reverse=True)
    exp_win = tk.Toplevel(root); exp_win.title("Export Pull History — Yubikey")
    exp_win.resizable(False,False); exp_win.transient(root); exp_win.grab_set()
    tk.Label(exp_win, text=f"Export {len(rows)} Pull History Row(s) to Excel",
             font=("Helvetica",10,"bold"), bg="#922b21", fg="white", padx=10, pady=6).pack(fill="x")
    form = tk.Frame(exp_win, padx=16, pady=12); form.pack()
    def _sec(text, row):
        tk.Label(form, text=text, font=("Helvetica",8,"bold"), fg="#922b21", anchor="w").grid(row=row, column=0, columnspan=3, sticky="w", pady=(10,2))
    _sec("▸ Pull Excel File", 0)
    tk.Label(form, text="File Name:", anchor="w", width=14).grid(row=1, column=0, sticky="w", pady=3)
    fn_var = tk.StringVar(); fn_cb = ttk.Combobox(form, textvariable=fn_var, width=28, values=existing_excels)
    fn_cb.grid(row=1, column=1, pady=3, padx=(4,0))
    tk.Label(form, text=".xlsx", fg="gray", font=("Helvetica",8)).grid(row=1, column=2, sticky="w", padx=(3,0))
    tk.Label(form, text="  ↳ Select existing file to append a sheet to, or type a new name", fg="gray", font=("Helvetica",7)).grid(row=2, column=1, columnspan=2, sticky="w")
    _sec("▸ Excel Sheet", 3)
    tk.Label(form, text="Sheet Name:", anchor="w", width=14).grid(row=4, column=0, sticky="w", pady=3)
    sn_var = tk.StringVar(); sn_cb = ttk.Combobox(form, textvariable=sn_var, width=28)
    sn_cb.grid(row=4, column=1, pady=3, padx=(4,0))
    tk.Label(form, text="  ↳ New sheet name (existing sheet of same name will be replaced)", fg="gray", font=("Helvetica",7)).grid(row=5, column=1, columnspan=2, sticky="w")
    def _refresh_sn(*_):
        chosen = fn_var.get().strip()
        if not chosen: sn_cb["values"]=[]; return
        safe = chosen if chosen.lower().endswith(".xlsx") else chosen+".xlsx"
        path = os.path.join(PULL_EXCEL_FOLDER_YK, safe.replace("/","-").replace("\\","-"))
        if os.path.exists(path):
            try:
                from openpyxl import load_workbook
                wb_p = load_workbook(path, read_only=True); sn_cb["values"]=wb_p.sheetnames; wb_p.close(); return
            except Exception: pass
        sn_cb["values"]=[]
    fn_cb.bind("<<ComboboxSelected>>", _refresh_sn); fn_var.trace_add("write", _refresh_sn)
    tk.Label(form, text="(Files saved to pull_excel/yubikey/)", fg="gray", font=("Helvetica",8)).grid(row=6, column=0, columnspan=3, sticky="w", pady=(8,0))
    err_lbl = tk.Label(form, text="", fg="red", font=("Helvetica",8)); err_lbl.grid(row=7, column=0, columnspan=3, sticky="w", pady=(4,0))
    confirmed = [False]
    def on_confirm():
        if not fn_var.get().strip(): err_lbl.config(text="Please enter an Excel file name."); return
        if not sn_var.get().strip(): err_lbl.config(text="Please enter a sheet name."); return
        confirmed[0]=True; exp_win.destroy()
    btn_r = tk.Frame(exp_win, pady=10); btn_r.pack()
    tk.Button(btn_r, text="EXPORT", command=on_confirm, bg="#922b21", fg="white", width=12).pack(side="left", padx=6)
    tk.Button(btn_r, text="Cancel", command=exp_win.destroy, width=10).pack(side="left", padx=6)
    exp_win.update_idletasks()
    px,py,pw,ph = root.winfo_rootx(),root.winfo_rooty(),root.winfo_width(),root.winfo_height()
    nw,nh=exp_win.winfo_reqwidth(),exp_win.winfo_reqheight()
    exp_win.geometry(f"+{px+(pw-nw)//2}+{py+(ph-nh)//2}"); exp_win.focus_force()
    root.wait_window(exp_win)
    if not confirmed[0]: return
    file_name_str = fn_var.get().strip(); sheet_str = sn_var.get().strip()
    safe_xl = file_name_str.replace("/","-").replace("\\","-")
    if not safe_xl.lower().endswith(".xlsx"): safe_xl += ".xlsx"
    out_path = os.path.join(PULL_EXCEL_FOLDER_YK, safe_xl)
    try:
        import openpyxl
        from openpyxl import load_workbook
        df_new = pd.DataFrame(records, columns=cols)
        df_new.insert(0, "Exported At", _now().strftime("%Y-%m-%d %H:%M:%S"))
        if os.path.exists(out_path):
            wb = load_workbook(out_path)
            for ws_p in wb.worksheets: ws_p.protection.sheet=False
            if sheet_str in wb.sheetnames:
                ws = wb[sheet_str]
                for rd in df_new.itertuples(index=False, name=None): ws.append(list(rd))
            else:
                ws = wb.create_sheet(title=sheet_str); ws.append(list(df_new.columns))
                for rd in df_new.itertuples(index=False, name=None): ws.append(list(rd))
        else:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet_str
            ws.append(list(df_new.columns))
            for rd in df_new.itertuples(index=False, name=None): ws.append(list(rd))
        for ws_p in wb.worksheets: ws_p.protection.sheet=True; ws_p.protection.enable()
        wb.save(out_path)
        _last_excel_path[3] = out_path
        messagebox.showinfo("Export Complete", f"{len(records)} row(s) exported to:\n{out_path}\nSheet: '{sheet_str}'\n\nUse VIEW EXCEL to open it.")
    except PermissionError: _excel_locked_error()
    except Exception as e: messagebox.showerror("Export Failed", f"Export failed:\n{e}")

def _yk_open_qr_gallery(filter_keys=None):
    from PIL import Image, ImageTk
    qr_win = tk.Toplevel(root)
    qr_win.title("Stored QR Codes — Yubikey" + (f"  [{len(filter_keys)} selected]" if filter_keys else ""))
    qr_win.geometry("860x560")
    toolbar = tk.Frame(qr_win, bg="#117a65"); toolbar.pack(fill="x")
    tk.Label(toolbar, text="Stored QR Codes — Yubikey", bg="#117a65", fg="white",
             font=("Helvetica",10,"bold")).pack(side="left", padx=10, pady=6)
    search_var = tk.StringVar()
    tk.Label(toolbar, text="Search:", bg="#117a65", fg="white").pack(side="left", padx=(20,2))
    tk.Entry(toolbar, textvariable=search_var, width=18).pack(side="left", pady=4)
    count_lbl = tk.Label(toolbar, text="", bg="#117a65", fg="#a9dfbf"); count_lbl.pack(side="left", padx=10)
    tk.Button(toolbar, text="↻", command=lambda: _load_gallery(search_var.get()),
              bg="#0e6655", fg="white", relief="flat", padx=8).pack(side="right", padx=8, pady=4)
    container = tk.Frame(qr_win); container.pack(fill="both", expand=True)
    canvas_qr = tk.Canvas(container, bg="#f4f6f7", highlightthickness=0)
    sb_qr = ttk.Scrollbar(container, orient="vertical", command=canvas_qr.yview)
    canvas_qr.configure(yscrollcommand=sb_qr.set); sb_qr.pack(side="right", fill="y"); canvas_qr.pack(side="left", fill="both", expand=True)
    canvas_qr.bind("<MouseWheel>", lambda e: canvas_qr.yview_scroll(int(-1*(e.delta/120)),"units"))
    inner = tk.Frame(canvas_qr, bg="#f4f6f7")
    cw_id = canvas_qr.create_window((0,0), window=inner, anchor="nw")
    _img_refs = []
    def _load_gallery(keyword=""):
        for w in inner.winfo_children(): w.destroy()
        _img_refs.clear()
        COLS,THUMB,PAD = 4,120,14; row_f=col_f=shown=0
        df = load_items_yk()
        for _, row in df.iterrows():
            key   = str(row.get("Hostname",""))
            shelf = ""
            path  = qr_path_for(key, warehouse=3)
            if filter_keys is not None and key not in filter_keys: continue
            if keyword and keyword.lower() not in key.lower(): continue
            cell = tk.Frame(inner, bg="white", bd=1, relief="solid", padx=PAD, pady=PAD)
            cell.grid(row=row_f, column=col_f, padx=8, pady=8, sticky="n")
            if os.path.exists(path):
                try:
                    img = Image.open(path).resize((THUMB,THUMB), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img); _img_refs.append(photo)
                    tk.Label(cell, image=photo, bg="white").pack()
                except Exception: tk.Label(cell, text="[Error]", bg="white", fg="red", width=14).pack()
            else: tk.Label(cell, text="[No QR File]", bg="#fdf2f2", fg="#c0392b", width=14, height=6, font=("Helvetica",8)).pack()
            tk.Label(cell, text=key, bg="white", font=("Helvetica",8,"bold"), fg="#117a65", wraplength=130).pack(pady=(4,0))
            sn = str(row.get("Serial Number",""))
            if sn: tk.Label(cell, text=f"S/N: {sn}", bg="white", font=("Helvetica",7), fg="#555").pack()
            col_f += 1; shown += 1
            if col_f >= COLS: col_f=0; row_f+=1
        if shown == 0:
            tk.Label(inner, text="No QR codes found.", bg="#f4f6f7", font=("Helvetica",10), fg="gray").grid(row=0, column=0, padx=20, pady=40)
        count_lbl.config(text=f"{shown} QR code(s)")
        inner.update_idletasks(); canvas_qr.configure(scrollregion=canvas_qr.bbox("all"))
        canvas_qr.itemconfigure(cw_id, width=canvas_qr.winfo_width())
    canvas_qr.bind("<Configure>", lambda e: canvas_qr.itemconfigure(cw_id, width=e.width))
    search_var.trace_add("write", lambda *_: _load_gallery(search_var.get()))
    _load_gallery()

def yk_view_stored_qr():
    pull_active = tree_yk_pullouts.winfo_ismapped()
    if pull_active:
        folder = PULL_QR_FOLDER_YK
        if not os.path.exists(folder): messagebox.showinfo("View QR","No pull QR codes folder found."); return
        files = [f for f in os.listdir(folder) if f.lower().endswith(".png")]
        if not files: messagebox.showinfo("View QR","No QR codes for pulled items yet."); return
        visible_iids = list(tree_yk_pullouts.get_children())
        filter_keys  = [str(tree_yk_pullouts.item(iid,"values")[1]) for iid in visible_iids] if visible_iids else None
        from PIL import Image, ImageTk
        qr_win = tk.Toplevel(root); qr_win.title("Pull QR Codes — Yubikey"); qr_win.geometry("860x560")
        toolbar = tk.Frame(qr_win, bg="#117a65"); toolbar.pack(fill="x")
        tk.Label(toolbar, text="Pull QR Codes — Yubikey", bg="#117a65", fg="white", font=("Helvetica",10,"bold")).pack(side="left", padx=10, pady=6)
        sv = tk.StringVar(); tk.Label(toolbar, text="Search:", bg="#117a65", fg="white").pack(side="left", padx=(20,2))
        tk.Entry(toolbar, textvariable=sv, width=18).pack(side="left", pady=4)
        cnt_lbl = tk.Label(toolbar, text="", bg="#117a65", fg="#a9dfbf"); cnt_lbl.pack(side="left", padx=10)
        container = tk.Frame(qr_win); container.pack(fill="both", expand=True)
        canvas_qr = tk.Canvas(container, bg="#f4f6f7", highlightthickness=0)
        sb = ttk.Scrollbar(container, orient="vertical", command=canvas_qr.yview)
        canvas_qr.configure(yscrollcommand=sb.set); sb.pack(side="right", fill="y"); canvas_qr.pack(side="left", fill="both", expand=True)
        canvas_qr.bind("<MouseWheel>", lambda e: canvas_qr.yview_scroll(int(-1*(e.delta/120)),"units"))
        inner = tk.Frame(canvas_qr, bg="#f4f6f7"); cw_id = canvas_qr.create_window((0,0), window=inner, anchor="nw")
        _img_refs = []
        def _load_pg(keyword=""):
            for w in inner.winfo_children(): w.destroy(); _img_refs.clear()
            COLS,THUMB,PAD=4,120,14; row_f=col_f=shown=0
            for fname in sorted(files):
                key = os.path.splitext(fname)[0].replace("_"," ")
                if filter_keys is not None and key not in filter_keys: continue
                if keyword and keyword.lower() not in key.lower(): continue
                path = os.path.join(folder, fname)
                cell = tk.Frame(inner, bg="white", bd=1, relief="solid", padx=PAD, pady=PAD)
                cell.grid(row=row_f, column=col_f, padx=8, pady=8, sticky="n")
                try:
                    img=Image.open(path).resize((THUMB,THUMB),Image.LANCZOS); photo=ImageTk.PhotoImage(img); _img_refs.append(photo)
                    tk.Label(cell, image=photo, bg="white").pack()
                except Exception: tk.Label(cell, text="[Error]", bg="white", fg="red", width=14).pack()
                tk.Label(cell, text=key, bg="white", font=("Helvetica",8,"bold"), fg="#117a65", wraplength=130).pack(pady=(4,0))
                col_f+=1; shown+=1
                if col_f>=COLS: col_f=0; row_f+=1
            if shown==0: tk.Label(inner, text="No QR codes found.", bg="#f4f6f7", font=("Helvetica",10), fg="gray").grid(row=0, column=0, padx=20, pady=40)
            cnt_lbl.config(text=f"{shown} QR code(s)"); inner.update_idletasks()
            canvas_qr.configure(scrollregion=canvas_qr.bbox("all")); canvas_qr.itemconfigure(cw_id, width=canvas_qr.winfo_width())
        tk.Button(toolbar, text="↻", command=lambda: _load_pg(sv.get()), bg="#0e6655", fg="white", relief="flat", padx=8).pack(side="right", padx=8, pady=4)
        canvas_qr.bind("<Configure>", lambda e: canvas_qr.itemconfigure(cw_id, width=e.width))
        sv.trace_add("write", lambda *_: _load_pg(sv.get())); _load_pg()
    else:
        folder = QR_FOLDER_YK
        if not os.path.exists(folder): messagebox.showinfo("View QR","No QR codes folder found."); return
        files = [f for f in os.listdir(folder) if f.lower().endswith(".png")]
        if not files: messagebox.showinfo("View QR","No QR codes have been generated yet."); return
        visible_iids = list(tree_yk_warehouse.get_children())
        filter_keys  = [str(tree_yk_warehouse.item(iid,"values")[2]) for iid in visible_iids] if visible_iids else [os.path.splitext(f)[0].replace("_"," ") for f in files]
        _yk_open_qr_gallery(filter_keys=filter_keys)

def yk_open_label_manager():
    """QR Label file manager for Yubikey."""
    manager = tk.Toplevel(root); manager.title("QR Label Manager — Yubikey"); manager.geometry("950x500"); manager.resizable(True, False)
    hdr = tk.Frame(manager, bg="#117a65"); hdr.pack(fill="x")
    tk.Label(hdr, text="QR Label Files — Yubikey", font=("Helvetica",10,"bold"), bg="#117a65", fg="white").pack(side="left", padx=10, pady=6)
    sel_count_lbl = tk.Label(hdr, text="", font=("Helvetica",9), bg="#117a65", fg="#f0b429"); sel_count_lbl.pack(side="right", padx=10)
    list_frame = tk.Frame(manager, bd=1, relief="sunken"); list_frame.pack(fill="both", expand=True, padx=10, pady=(8,4))
    canvas_cl = tk.Canvas(list_frame, bg="white", highlightthickness=0)
    sb_cl = ttk.Scrollbar(list_frame, orient="vertical", command=canvas_cl.yview)
    canvas_cl.configure(yscrollcommand=sb_cl.set); sb_cl.pack(side="right", fill="y"); canvas_cl.pack(side="left", fill="both", expand=True)
    canvas_cl.bind("<MouseWheel>", lambda e: canvas_cl.yview_scroll(int(-1*(e.delta/120)),"units"))
    inner_cl = tk.Frame(canvas_cl, bg="white"); cw_id = canvas_cl.create_window((0,0), window=inner_cl, anchor="nw")
    canvas_cl.bind("<Configure>", lambda e: canvas_cl.itemconfigure(cw_id, width=e.width))
    row_data=[]; check_vars={}
    def _refresh_sel():
        n = sum(1 for v in check_vars.values() if v.get()); sel_count_lbl.config(text=f"{n} selected" if n else ""); clear_btn.config(state="normal" if n else "disabled")
    def _toggle_all():
        want = not all(v.get() for v in check_vars.values())
        for v in check_vars.values(): v.set(want)
        _refresh_sel(); _repaint()
    def _repaint():
        for iid,_,_,_,_,_,var in row_data:
            try:
                fr = inner_cl.nametowidget(f"row_{iid}")
                fr.config(bg="#e8f5e9" if var.get() else ("white" if row_data.index(next(r for r in row_data if r[0]==iid))%2==0 else "#f7f9fc"))
            except Exception: pass
    def load_files():
        for w in inner_cl.winfo_children(): w.destroy(); row_data.clear(); check_vars.clear()
        now = _now()
        hdr_row = tk.Frame(inner_cl, bg="#dce3f0"); hdr_row.pack(fill="x")
        tk.Label(hdr_row, text="✔", width=3, bg="#dce3f0", font=("Helvetica",9,"bold")).pack(side="left", padx=(6,0))
        for txt,w in [("Type",12),("Filename",28),("Created",22),("Size",10)]:
            tk.Label(hdr_row, text=txt, width=w, bg="#dce3f0", font=("Helvetica",9,"bold"), anchor="w").pack(side="left", padx=4, pady=5)
        idx = 0
        for xl_folder, folder_type in [(QR_LABELS_FOLDER_YK,"PDF Labels")]:
            if not os.path.exists(xl_folder): continue
            ext = ".pdf" if folder_type == "PDF Labels" else ".xlsx"
            files = sorted([f for f in os.listdir(xl_folder) if f.lower().endswith(ext)], reverse=True)
            for f in files:
                full_path = os.path.join(xl_folder, f)
                size_kb = round(os.path.getsize(full_path)/1024, 1)
                try:
                    mtime = os.path.getmtime(full_path); file_dt = datetime.fromtimestamp(mtime)
                    delta = now-file_dt; age = "Today" if delta.days==0 else ("1 day ago" if delta.days==1 else f"{delta.days} days ago")
                    date_str = f"{file_dt.strftime('%Y-%m-%d')}  ({age})"
                except Exception: date_str="Unknown"
                iid=f"row{idx}"; var=tk.BooleanVar(value=False); check_vars[iid]=var
                bg="white" if idx%2==0 else "#f7f9fc"
                row_fr=tk.Frame(inner_cl,bg=bg,name=f"row_{iid}",cursor="hand2"); row_fr.pack(fill="x")
                def _make_tog(v=var):
                    def _t(e=None): v.set(not v.get()); _refresh_sel(); _repaint()
                    return _t
                tk.Checkbutton(row_fr, variable=var, bg=bg, command=lambda: [_refresh_sel(),_repaint()]).pack(side="left", padx=(6,0), pady=4)
                size_str = f"{size_kb} KB" if size_kb<1024 else f"{round(size_kb/1024,2)} MB"
                for txt,w in [(folder_type,14),(f,28),(date_str,22),(size_str,10)]:
                    lbl=tk.Label(row_fr,text=txt,bg=bg,anchor="w",width=w,font=("Helvetica",9)); lbl.pack(side="left",padx=4,pady=4); lbl.bind("<Button-1>",_make_tog(var))
                row_fr.bind("<Button-1>",_make_tog(var))
                row_data.append((iid,full_path,"Yubikey",f,date_str,f"{size_kb}kb",var)); idx+=1
        if idx==0: tk.Label(inner_cl, text="No label/excel files found.", fg="gray", font=("Helvetica",10), bg="white").pack(pady=30)
        inner_cl.update_idletasks(); canvas_cl.configure(scrollregion=canvas_cl.bbox("all")); _refresh_sel()
    def open_sel():
        chosen=[rd for rd in row_data if rd[6].get()]
        if not chosen: messagebox.showwarning("Warning","Check a file to open.", parent=manager); return
        if len(chosen)>1: messagebox.showerror("Error","Only 1 file can be opened at a time.", parent=manager); return
        fp=chosen[0][1]
        if os.path.exists(fp): os.startfile(fp)
        else: messagebox.showerror("Error","File not found.", parent=manager)
    def clear_sel():
        chosen=[rd for rd in row_data if rd[6].get()]
        if not chosen: messagebox.showwarning("Warning","Check at least one file.", parent=manager); return
        if not messagebox.askyesno("Confirm Move to Dump", f"Move {len(chosen)} file(s) to dump?", parent=manager): return
        import shutil, stat, ctypes
        failed=[]
        for rd in chosen:
            _,fp,_,fname,_,_,_ = rd
            dump_folder = os.path.join(DUMP_FOLDER,"qr_labels","yubikey")
            os.makedirs(dump_folder, exist_ok=True)
            try:
                if os.path.exists(fp):
                    try: ctypes.windll.kernel32.SetFileAttributesW(fp, 0x80)
                    except: pass
                    os.chmod(fp, stat.S_IWRITE | stat.S_IREAD)
                    dest=os.path.join(dump_folder, fname)
                    if os.path.exists(dest):
                        base,ext=os.path.splitext(fname); dest=os.path.join(dump_folder,f"{base}_{_now().strftime('%Y%m%d%H%M%S')}{ext}")
                    shutil.move(fp, dest)
            except Exception as e: failed.append(f"{fname}: {e}")
        load_files()
        if failed: messagebox.showerror("Error","Some files could not be moved:\n"+"\n".join(failed), parent=manager)
        else: messagebox.showinfo("Moved to Dump",f"{len(chosen)} file(s) moved to dump folder.", parent=manager)
    btn_frame_m = tk.Frame(manager); btn_frame_m.pack(pady=8)
    tk.Button(btn_frame_m, text="☑", command=_toggle_all, width=4).pack(side="left", padx=4)
    tk.Button(btn_frame_m, text="OPEN", command=open_sel, width=10).pack(side="left", padx=4)
    clear_btn = tk.Button(btn_frame_m, text="✕", command=clear_sel, width=10, bg="#922b21", fg="white", state="disabled")
    clear_btn.pack(side="left", padx=4)
    load_files()

def yk_view_excel():
    """Open the Excel file manager for Yubikey."""
    manager = tk.Toplevel(root); manager.title("Excel File Manager — Yubikey"); manager.geometry("700x500"); manager.resizable(False,False)
    hdr = tk.Frame(manager, bg="#1e8449"); hdr.pack(fill="x")
    tk.Label(hdr, text="Generated Excel Files — Yubikey", font=("Helvetica",10,"bold"), bg="#1e8449", fg="white").pack(side="left", padx=10, pady=6)
    sel_count_lbl = tk.Label(hdr, text="", font=("Helvetica",9), bg="#1e8449", fg="#f0b429"); sel_count_lbl.pack(side="right", padx=10)
    list_frame = tk.Frame(manager, bd=1, relief="sunken"); list_frame.pack(fill="both", expand=True, padx=10, pady=(8,4))
    canvas_cl = tk.Canvas(list_frame, bg="white", highlightthickness=0)
    sb_cl = ttk.Scrollbar(list_frame, orient="vertical", command=canvas_cl.yview)
    canvas_cl.configure(yscrollcommand=sb_cl.set); sb_cl.pack(side="right", fill="y"); canvas_cl.pack(side="left", fill="both", expand=True)
    canvas_cl.bind("<MouseWheel>", lambda e: canvas_cl.yview_scroll(int(-1*(e.delta/120)),"units"))
    inner_cl = tk.Frame(canvas_cl, bg="white"); cw_id = canvas_cl.create_window((0,0), window=inner_cl, anchor="nw")
    canvas_cl.bind("<Configure>", lambda e: canvas_cl.itemconfigure(cw_id, width=e.width))
    row_data=[]; check_vars={}
    def _refresh_sel():
        n=sum(1 for v in check_vars.values() if v.get()); sel_count_lbl.config(text=f"{n} selected" if n else ""); clear_btn.config(state="normal" if n else "disabled")
    def _toggle_all():
        want=not all(v.get() for v in check_vars.values())
        for v in check_vars.values(): v.set(want)
        _refresh_sel()
    def load_excel_files():
        for w in inner_cl.winfo_children(): w.destroy(); row_data.clear(); check_vars.clear()
        now=_now()
        hdr_row=tk.Frame(inner_cl,bg="#dce3f0"); hdr_row.pack(fill="x")
        tk.Label(hdr_row,text="✔",width=3,bg="#dce3f0",font=("Helvetica",9,"bold")).pack(side="left",padx=(6,0))
        for txt,w in [("Type",12),("Filename",24),("Created",22),("Size",10)]:
            tk.Label(hdr_row,text=txt,width=w,bg="#dce3f0",font=("Helvetica",9,"bold"),anchor="w").pack(side="left",padx=4,pady=5)
        idx=0
        for xl_folder, folder_type in [(EXCEL_FOLDER_YK,"Generated"),(PULL_EXCEL_FOLDER_YK,"Pull History")]:
            if not os.path.exists(xl_folder): continue
            files=sorted([f for f in os.listdir(xl_folder) if f.lower().endswith(".xlsx")],reverse=True)
            for f in files:
                full_path=os.path.join(xl_folder,f); size_kb=round(os.path.getsize(full_path)/1024,1)
                try:
                    mtime=os.path.getmtime(full_path); file_dt=datetime.fromtimestamp(mtime); delta=now-file_dt
                    age="Today" if delta.days==0 else ("1 day ago" if delta.days==1 else f"{delta.days} days ago")
                    date_str=f"{file_dt.strftime('%Y-%m-%d')}  ({age})"
                except: date_str="Unknown"
                iid=f"row{idx}"; var=tk.BooleanVar(value=False); check_vars[iid]=var
                bg="white" if idx%2==0 else "#f7f9fc"
                row_fr=tk.Frame(inner_cl,bg=bg,cursor="hand2"); row_fr.pack(fill="x")
                def _make_tog(v=var,fp=full_path):
                    def _t(e=None): v.set(not v.get()); _refresh_sel()
                    return _t
                tk.Checkbutton(row_fr,variable=var,bg=bg,command=_refresh_sel).pack(side="left",padx=(6,0),pady=4)
                size_str=f"{size_kb} KB" if size_kb<1024 else f"{round(size_kb/1024,2)} MB"
                for txt,w in [(folder_type,12),(f,24),(date_str,22),(size_str,10)]:
                    lbl=tk.Label(row_fr,text=txt,bg=bg,anchor="w",width=w,font=("Helvetica",9)); lbl.pack(side="left",padx=4,pady=4); lbl.bind("<Button-1>",_make_tog(var,full_path))
                row_fr.bind("<Button-1>",_make_tog(var,full_path))
                row_data.append((iid,full_path,"Yubikey",f,date_str,f"{size_kb}kb",var,folder_type,3)); idx+=1
        if idx==0: tk.Label(inner_cl,text="No generated Excel files found.",fg="gray",font=("Helvetica",10),bg="white").pack(pady=30)
        inner_cl.update_idletasks(); canvas_cl.configure(scrollregion=canvas_cl.bbox("all")); _refresh_sel()
    def open_selected():
        chosen=[rd for rd in row_data if rd[6].get()]
        if not chosen: messagebox.showwarning("Warning","Check a file to open.", parent=manager); return
        if len(chosen)>1: messagebox.showerror("Error","Only 1 file can be opened at a time.", parent=manager); return
        fp=chosen[0][1]
        if os.path.exists(fp): os.startfile(fp)
        else: messagebox.showerror("Error","File not found.", parent=manager)
    def clear_selected():
        chosen=[rd for rd in row_data if rd[6].get()]
        if not chosen: messagebox.showwarning("Warning","Check at least one file.", parent=manager); return
        if not messagebox.askyesno("Confirm Move to Dump",f"Move {len(chosen)} file(s) to dump?", parent=manager): return
        import shutil,stat,ctypes
        failed=[]
        for rd in chosen:
            _,fp,_,fname,_,_,_,folder_type,_ = rd
            dump_folder=DUMP_EXCEL_YK
            os.makedirs(dump_folder, exist_ok=True)
            try:
                if os.path.exists(fp):
                    try: ctypes.windll.kernel32.SetFileAttributesW(fp,0x80)
                    except: pass
                    os.chmod(fp,stat.S_IWRITE|stat.S_IREAD)
                    dest=os.path.join(dump_folder,fname)
                    if os.path.exists(dest):
                        base,ext=os.path.splitext(fname); dest=os.path.join(dump_folder,f"{base}_{_now().strftime('%Y%m%d%H%M%S')}{ext}")
                    shutil.move(fp,dest)
                    save_log("FILE DELETED", f"File: {fname} | Warehouse: [YK] | Moved to dump: {dest}")
            except Exception as e: failed.append(f"{fname}: {e}")
        load_excel_files()
        if failed: messagebox.showerror("Error","Some files could not be moved:\n"+"\n".join(failed), parent=manager)
        else: messagebox.showinfo("Moved to Dump",f"{len(chosen)} file(s) moved to dump folder.", parent=manager)
    btn_frame_m=tk.Frame(manager); btn_frame_m.pack(pady=8)
    tk.Button(btn_frame_m,text="☑",command=_toggle_all,width=4).pack(side="left",padx=4)
    tk.Button(btn_frame_m,text="OPEN",command=open_selected,width=10).pack(side="left",padx=4)
    clear_btn=tk.Button(btn_frame_m,text="✕",command=clear_selected,width=10,bg="#922b21",fg="white",state="disabled"); clear_btn.pack(side="left",padx=4)
    load_excel_files()

# ========== QR LABEL PDF ==========

def _lock_pdf(pdf_path):
    """Restrict PDF to view-only: tries pikepdf encryption first, always falls back to OS read-only."""
    tmp_lock = pdf_path + ".~lock.pdf"
    try:
        import pikepdf
        permissions = pikepdf.Permissions(
            accessibility=True,
            extract=True,
            modify_annotation=False,
            modify_assembly=False,
            modify_form=False,
            modify_other=False,
            print_highres=True,
            print_lowres=True,
        )
        with pikepdf.open(pdf_path) as pdf_obj:
            pdf_obj.save(
                tmp_lock,
                encryption=pikepdf.Encryption(
                    owner="WMS_READONLY_2026",
                    user="",
                    R=4,
                    allow=permissions,
                )
            )
        os.replace(tmp_lock, pdf_path)
    except ImportError:
        pass
    except Exception:
        try:
            os.remove(tmp_lock)
        except Exception:
            pass
    # Always apply OS-level read-only regardless of pikepdf outcome
    try:
        import stat
        os.chmod(pdf_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass


def generate_qr_pdf(items_batch, custom_name=None):
    import json
    from fpdf import FPDF
    PAGE_W, LABEL_W, COLS = 210, 54, 3
    MARGIN_X, MARGIN_Y, ROW_GAP = 12, 10, 4
    GAP_X = (PAGE_W - (COLS * LABEL_W) - (2 * MARGIN_X)) / (COLS - 1)
    QR_SIZE, QR_PAD_TOP, QR_PAD_BTM = 20, 3, 3
    LINE_H, FIELD_PAD_BTM = 4.8, 2

    if not items_batch:
        return None

    # ── Resolve output path first (needed for sidecar lookup) ──
    is_w2_batch = items_batch[0].get('_warehouse', 1) == 2
    if is_w2_batch:
        output_folder = QR_LABELS_FOLDER_W2
    else:
        output_folder = QR_LABELS_FOLDER_W1

    os.makedirs(output_folder, exist_ok=True)

    if custom_name:
        safe_name = custom_name.replace(" ", "_").replace("/", "-").replace("\\", "-")
        if safe_name.lower().endswith(".pdf"):
            safe_name = safe_name[:-4]
    else:
        if is_w2_batch:
            set_ids = list(dict.fromkeys(item.get("Set ID", "") for item in items_batch))
            label_name = "_".join(set_ids[:3])
            if len(set_ids) > 3:
                label_name += f"_and_{len(set_ids)-3}_more"
        else:
            existing = [f for f in os.listdir(output_folder) if f.startswith("BATCH_") and f.endswith(".pdf")]
            batch_nums = []
            for f in existing:
                try:
                    batch_nums.append(int(f.split("_")[1]))
                except (IndexError, ValueError):
                    pass
            label_name = f"BATCH_{max(batch_nums, default=0) + 1}"
        safe_name = label_name.replace(" ", "_").replace("/", "-")

    pdf_path   = os.path.join(output_folder, f"{safe_name}.pdf")
    index_path = pdf_path + ".keys.json"

    # ── Load existing sidecar (item data + keys) ──
    existing_items = []
    existing_keys  = set()
    if os.path.exists(index_path) and os.path.exists(pdf_path):
        try:
            with open(index_path, "r") as kf:
                sidecar = json.load(kf)
            existing_items = sidecar.get("items", [])
            existing_keys  = set(sidecar.get("keys",  []))
            # Ensure _warehouse is always present on re-loaded sidecar items
            # so they render with the correct label layout (W1 vs W2)
            _wh_tag = 2 if is_w2_batch else 1
            for _it in existing_items:
                _it.setdefault("_warehouse", _wh_tag)
        except Exception:
            existing_items = []
            existing_keys  = set()

    def _item_key(item):
        if item.get('_warehouse', 1) == 2:
            return f"{item.get('Set ID', '')}-{item.get('Equipment Type', '')}"
        return str(item.get('Hostname', ''))

    # Only keep truly new items
    new_items = [it for it in items_batch if _item_key(it) not in existing_keys]
    if not new_items:
        return pdf_path  # nothing to add

    # ── Combine: existing items first, then new ones ──
    all_items = existing_items + new_items

    # ── Lift OS read-only before writing (re-applied after via _lock_pdf) ──
    if os.path.exists(pdf_path):
        try:
            import stat
            os.chmod(pdf_path, stat.S_IWRITE | stat.S_IREAD)
        except Exception:
            pass

    # ── Render all items into a fresh PDF so grid is always packed ──
    # ── Label dimensions — landscape orientation, QR left, fields right ──
    LABEL_W   = 80    # wider label to fit QR + text side by side
    LABEL_H   = 24    # fixed height (matches QR size + padding)
    QR_SIZE   = 20    # QR image size
    QR_PAD    = 2     # padding around QR
    COLS      = 2     # two labels per row
    MARGIN_X  = 10
    MARGIN_Y  = 10
    GAP_X     = 10    # horizontal gap between label columns
    GAP_Y     = 4     # vertical gap between label rows
    TEXT_X_OFF = QR_SIZE + QR_PAD * 2 + 2   # text starts after QR
    TEXT_W     = LABEL_W - TEXT_X_OFF - 2    # available text width
    LINE_H     = 4.8

    pdf = FPDF(orientation='L', unit='mm', format='A4')  # landscape page
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    col = row = 0

    for item in all_items:
        is_w2 = item.get('_warehouse', 1) == 2

        if is_w2:
            fields = [
                ("Hostname:",   str(item.get("Hostname", ""))),
                ("Serial No:",  str(item.get("Serial Number", ""))),
                ("Checked By:", str(item.get("Checked By", ""))),
                ("Date:",       str(item.get("_date", _now().strftime("%Y-%m-%d")))),
            ]
            qr_key = f"{item.get('Set ID', '')}-{item.get('Equipment Type', '')}"
            path   = qr_path_for(qr_key, warehouse=2)
        else:
            fields = [
                ("Hostname:",   str(item.get("Hostname", ""))),
                ("Checked By:", str(item.get("Checked By", ""))),
                ("Date:",       str(item.get("_date", _now().strftime("%Y-%m-%d")))),
            ]
            path = qr_path_for(item.get('Hostname', ''), warehouse=1)

        x = MARGIN_X + col * (LABEL_W + GAP_X)
        y = MARGIN_Y + row * (LABEL_H + GAP_Y)

        # New page if label would overflow
        if y + LABEL_H > 210 - MARGIN_Y:   # 210 = A4 landscape height
            pdf.add_page(); col = row = 0
            x = MARGIN_X
            y = MARGIN_Y

        # Label border
        pdf.set_draw_color(150, 150, 150)
        pdf.rect(x, y, LABEL_W, LABEL_H)

        # QR on the left, vertically centred
        if os.path.exists(path):
            qr_y = y + (LABEL_H - QR_SIZE) / 2
            pdf.image(path, x=x + QR_PAD, y=qr_y, w=QR_SIZE, h=QR_SIZE)

        # Fields to the right of QR, vertically centred as a block
        # TEXT_X_OFF = 26, TEXT_W = 52
        # LABEL_COL_W must match value_x offset exactly — both use same constant
        total_text_h = len(fields) * LINE_H
        text_y       = y + (LABEL_H - total_text_h) / 2
        label_x      = x + TEXT_X_OFF
        for lbl, val in fields:
            pdf.set_font("Helvetica", style="B", size=5.5)
            pdf.set_xy(label_x, text_y)
            pdf.cell(TEXT_W, LINE_H, f"{lbl} {val[:28]}", ln=0)
            text_y += LINE_H

        col += 1
        if col >= COLS:
            col = 0; row += 1

    pdf.output(pdf_path)

    # ── Stamp date on new items before saving to sidecar ──
    today = _now().strftime("%Y-%m-%d")
    for it in new_items:
        it.setdefault("_date", today)

    # ── Update sidecar with all item data + keys ──
    all_keys = list(existing_keys | {_item_key(it) for it in new_items})
    try:
        with open(index_path, "w") as kf:
            json.dump({"keys": all_keys, "items": all_items}, kf)
    except Exception:
        pass

    _lock_pdf(pdf_path)
    try:
        import stat
        os.chmod(pdf_path, stat.S_IREAD | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass
    return pdf_path

# ========== DIALOGS ==========

def open_label_manager(warehouse=None):
    manager = tk.Toplevel(root)
    manager.title(f"QR Label Manager — Warehouse {warehouse}" if warehouse else "QR Label Manager")
    manager.geometry("950x500")
    manager.resizable(True, False)

    # ── Header ──────────────────────────────────────────────
    hdr = tk.Frame(manager, bg="#2c3e50")
    hdr.pack(fill="x")
    tk.Label(hdr, text="QR Label Files", font=("Helvetica", 10, "bold"),
             bg="#2c3e50", fg="white").pack(side="left", padx=10, pady=6)
    sel_count_lbl = tk.Label(hdr, text="", font=("Helvetica", 9),
                              bg="#2c3e50", fg="#f0b429")
    sel_count_lbl.pack(side="right", padx=10)

    # ── Checklist canvas area ────────────────────────────────
    list_frame = tk.Frame(manager, bd=1, relief="sunken")
    list_frame.pack(fill="both", expand=True, padx=10, pady=(8, 4))

    canvas_cl = tk.Canvas(list_frame, bg="white", highlightthickness=0)
    sb_cl = ttk.Scrollbar(list_frame, orient="vertical", command=canvas_cl.yview)
    canvas_cl.configure(yscrollcommand=sb_cl.set)
    sb_cl.pack(side="right", fill="y")
    canvas_cl.pack(side="left", fill="both", expand=True)
    canvas_cl.bind("<MouseWheel>", lambda e: canvas_cl.yview_scroll(int(-1*(e.delta/120)), "units"))

    inner_cl = tk.Frame(canvas_cl, bg="white")
    cw_id = canvas_cl.create_window((0, 0), window=inner_cl, anchor="nw")
    canvas_cl.bind("<Configure>", lambda e: canvas_cl.itemconfigure(cw_id, width=e.width))

    # ── State ────────────────────────────────────────────────
    row_data    = []   # list of (iid, full_path, wh_label, filename, date_str, size_str, var)
    check_vars  = {}   # iid -> BooleanVar

    def _refresh_sel_count():
        n = sum(1 for v in check_vars.values() if v.get())
        sel_count_lbl.config(text=f"{n} selected" if n else "")
        clear_btn.config(state="normal" if n else "disabled")

    def _toggle_all():
        want = not all(v.get() for v in check_vars.values())
        for v in check_vars.values():
            v.set(want)
        _refresh_sel_count()
        _repaint_rows()

    def _repaint_rows():
        for iid, _, _, _, _, _, var in row_data:
            try:
                fr = inner_cl.nametowidget(f"row_{iid}")
                fr.config(bg="#e8f4fd" if var.get() else ("white" if row_data.index(
                    next(r for r in row_data if r[0] == iid)) % 2 == 0 else "#f7f9fc"))
            except Exception:
                pass

    EVEN_BG, ODD_BG, SEL_BG = "white", "#f7f9fc", "#e8f4fd"

    def load_label_files():
        # clear
        for w in inner_cl.winfo_children():
            w.destroy()
        row_data.clear()
        check_vars.clear()

        now = _now()
        all_warehouses = [("Warehouse 1", QR_LABELS_FOLDER_W1), ("Warehouse 2", QR_LABELS_FOLDER_W2)]
        filtered = [(wl, f) for wl, f in all_warehouses if warehouse is None or wl == f"Warehouse {warehouse}"]

        # Column header row
        hdr_row = tk.Frame(inner_cl, bg="#dce3f0")
        hdr_row.pack(fill="x")
        tk.Label(hdr_row, text="✔", width=3, bg="#dce3f0", font=("Helvetica", 9, "bold")).pack(side="left", padx=(6,0))
        for txt, w in [("Warehouse", 110), ("Filename", 240), ("Created", 155), ("Size", 60)]:
            tk.Label(hdr_row, text=txt, width=w//7, bg="#dce3f0",
                     font=("Helvetica", 9, "bold"), anchor="w").pack(side="left", padx=4, pady=5)

        idx = 0
        for warehouse_label, folder in filtered:
            if not os.path.exists(folder):
                continue
            for f in sorted([fn for fn in os.listdir(folder) if fn.endswith(".pdf")], reverse=True):
                full_path = os.path.join(folder, f)
                size_kb = round(os.path.getsize(full_path) / 1024, 1)
                try:
                    mtime = os.path.getmtime(full_path)
                    file_dt = datetime.fromtimestamp(mtime)
                    delta = now - file_dt
                    age = "Today" if delta.days == 0 else ("1 day ago" if delta.days == 1 else f"{delta.days} days ago")
                    date_str = f"{file_dt.strftime('%Y-%m-%d')}  ({age})"
                except Exception:
                    date_str = "Unknown"

                iid = f"row{idx}"
                var = tk.BooleanVar(value=False)
                check_vars[iid] = var
                bg = EVEN_BG if idx % 2 == 0 else ODD_BG

                row_fr = tk.Frame(inner_cl, bg=bg, name=f"row_{iid}", cursor="hand2")
                row_fr.pack(fill="x")

                def _make_toggle(v=var):
                    def _toggle(e=None):
                        v.set(not v.get())
                        _refresh_sel_count()
                        _repaint_rows()
                    return _toggle

                cb = tk.Checkbutton(row_fr, variable=var, bg=bg,
                                    command=lambda: [_refresh_sel_count(), _repaint_rows()])
                cb.pack(side="left", padx=(6, 0), pady=4)
                size_str = f"{size_kb} KB" if size_kb < 1024 else f"{round(size_kb/1024, 2)} MB"
                for txt, w in [(warehouse_label, 16), (f, 34), (date_str, 22), (size_str, 10)]:
                    lbl = tk.Label(row_fr, text=txt, bg=bg, anchor="w", width=w,
                                   font=("Helvetica", 9))
                    lbl.pack(side="left", padx=4, pady=4)
                    lbl.bind("<Button-1>", _make_toggle(var))
                row_fr.bind("<Button-1>", _make_toggle(var))

                row_data.append((iid, full_path, warehouse_label, f, date_str, f"{size_kb} kb", var))
                idx += 1

        if idx == 0:
            tk.Label(inner_cl, text="No QR label PDF files found.", fg="gray",
                     font=("Helvetica", 10), bg="white").pack(pady=30)

        inner_cl.update_idletasks()
        canvas_cl.configure(scrollregion=canvas_cl.bbox("all"))
        _refresh_sel_count()

    def open_selected():
        chosen = [rd for rd in row_data if rd[6].get()]
        if not chosen:
            messagebox.showwarning("Warning", "Check a file to open.", parent=manager); return
        if len(chosen) > 1:
            messagebox.showerror("Error", "Only 1 file can be opened at a time.\nPlease check only one file.", parent=manager); return
        full_path = chosen[0][1]
        if os.path.exists(full_path):
            os.startfile(full_path)
        else:
            messagebox.showerror("Error", "File not found.", parent=manager)

    def clear_selected():
        chosen = [rd for rd in row_data if rd[6].get()]
        if not chosen:
            messagebox.showwarning("Warning", "Check at least one file to move to dump.", parent=manager); return
        count = len(chosen)
        prompt = (f"Move '{chosen[0][3]}' to dump?" if count == 1
                  else f"Move {count} checked file(s) to dump?\nThey can be recovered from the dump folder.")
        if not messagebox.askyesno("Confirm Move to Dump", prompt, parent=manager):
            return
        import shutil, stat
        dump_folder = DUMP_LABELS_W1 if warehouse == 1 else DUMP_LABELS_W2
        os.makedirs(dump_folder, exist_ok=True)
        failed = []
        for _, full_path, _, fname, _, _, _ in chosen:
            try:
                if os.path.exists(full_path):
                    os.chmod(full_path, stat.S_IWRITE | stat.S_IREAD)
                    dest = os.path.join(dump_folder, fname)
                    if os.path.exists(dest):
                        base, ext = os.path.splitext(fname)
                        dest = os.path.join(dump_folder,
                                            f"{base}_{_now().strftime('%Y%m%d%H%M%S')}{ext}")
                    shutil.move(full_path, dest)
                    save_log("FILE DELETED", f"File: {fname} | Warehouse: {'[W1]' if warehouse == 1 else '[W2]'} | Moved to dump: {dest}")
                    sidecar = full_path + ".keys.json"
                    if os.path.exists(sidecar):
                        shutil.move(sidecar, dest + ".keys.json")
            except Exception as e:
                failed.append(f"{fname}: {e}")
        load_label_files()
        if failed:
            messagebox.showerror("Error", "Some files could not be moved:\n" + "\n".join(failed), parent=manager)
        else:
            messagebox.showinfo("Moved to Dump", f"{count} file(s) moved to dump folder.", parent=manager)

    # ── Bottom toolbar ───────────────────────────────────────
    btn_frame_m = tk.Frame(manager)
    btn_frame_m.pack(pady=8)
    tk.Button(btn_frame_m, text="☑", command=_toggle_all,   width=4).pack(side="left", padx=4)
    tk.Button(btn_frame_m, text="OPEN", command=open_selected, width=10).pack(side="left", padx=4)
    clear_btn = tk.Button(btn_frame_m, text="✕", command=clear_selected,
                           width=4, bg="#922b21", fg="white", state="disabled")
    clear_btn.pack(side="left", padx=4)

    load_label_files()

def open_activity_log():
    log_win = tk.Toplevel(root)
    log_win.title("Activity Log")
    log_win.geometry("900x560")

    # ── Shared filter bar (applies to All Logs tab) ────────────
    filter_frame = tk.Frame(log_win)
    filter_frame.pack(fill="x", padx=10, pady=5)

    tk.Label(filter_frame, text="Filter Action:").pack(side="left", padx=(0, 2))
    filter_action_var = tk.StringVar()
    filter_action_cb = ttk.Combobox(filter_frame, textvariable=filter_action_var, state="readonly", width=18,
        values=["", "LOGIN", "LOGOUT",
                "PUT WAREHOUSE", "WAREHOUSE PULL", "UPDATE ITEM", "DELETE ITEM",
                "UNDO PULL", "UNSTAGE",
                "SHELF STATUS", "ADD SHELF", "REMOVE SHELF",
                "GENERATE FILES", "FILE CREATED", "FILE UPDATED", "FILE DELETED",
                "SEARCH",
                "OPEN ADMIN PANEL", "CREATE ACCOUNT", "DELETE ACCOUNT", "CHANGE PASSWORD"]
    )
    filter_action_cb.pack(side="left", padx=(0, 10))

    log_date_from_var = tk.StringVar()
    log_date_to_var   = tk.StringVar()
    _date_picker_widget(filter_frame, log_date_from_var, "From:").pack(side="left", padx=(5, 4))
    _date_picker_widget(filter_frame, log_date_to_var,   "To:").pack(side="left", padx=(0, 6))
    tk.Button(filter_frame, text="↻", command=lambda: reset_filters(), width=3).pack(side="left", padx=(0, 10))

    count_label = tk.Label(filter_frame, text="", fg="blue")
    count_label.pack(side="right", padx=10)

    # ── Notebook with two tabs ─────────────────────────────────
    notebook = ttk.Notebook(log_win)
    notebook.pack(fill="both", expand=True, padx=10, pady=(0, 5))

    # ── TAB 1: All Logs ───────────────────────────────────────
    tab_all = tk.Frame(notebook)
    notebook.add(tab_all, text="📋  All Logs")

    content_frame = tk.Frame(tab_all)
    content_frame.pack(fill="both", expand=True)

    user_panel = tk.LabelFrame(content_frame, text="Users", padx=5, pady=5)
    user_panel.pack(side="left", fill="y", padx=(0, 8))

    user_scrollbar = ttk.Scrollbar(user_panel, orient="vertical")
    user_scrollbar.pack(side="right", fill="y")
    user_listbox = tk.Listbox(user_panel, width=18, yscrollcommand=user_scrollbar.set,
                               selectmode="single", exportselection=False, font=("Helvetica", 9))
    user_listbox.pack(side="left", fill="y")
    user_scrollbar.config(command=user_listbox.yview)

    btn_frame = tk.Frame(user_panel)
    btn_frame.pack(pady=(5, 0))

    table_frame_l = tk.Frame(content_frame)
    table_frame_l.pack(side="left", fill="both", expand=True)
    scrollbar_y = ttk.Scrollbar(table_frame_l, orient="vertical")
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x = ttk.Scrollbar(table_frame_l, orient="horizontal")
    scrollbar_x.pack(side="bottom", fill="x")
    tree_log = ttk.Treeview(table_frame_l,
                             columns=("Timestamp", "User", "Action", "Details"),
                             show="headings",
                             yscrollcommand=scrollbar_y.set,
                             xscrollcommand=scrollbar_x.set)
    for col, width in [("Timestamp", 140), ("User", 110), ("Action", 140), ("Details", 390)]:
        tree_log.heading(col, text=col); tree_log.column(col, width=width)
    tree_log.pack(fill="both", expand=True)
    scrollbar_y.config(command=tree_log.yview)
    scrollbar_x.config(command=tree_log.xview)

    # ── TAB 2: Generated Files ────────────────────────────────
    tab_files = tk.Frame(notebook)
    notebook.add(tab_files, text="📁  Generated Files")

    # Sub-filter bar inside Generated Files tab
    gf_filter_frame = tk.Frame(tab_files, pady=4)
    gf_filter_frame.pack(fill="x", padx=8)
    tk.Label(gf_filter_frame, text="Show:", font=("Helvetica", 9)).pack(side="left", padx=(0, 4))
    gf_action_var = tk.StringVar(value="")
    gf_action_cb = ttk.Combobox(gf_filter_frame, textvariable=gf_action_var, state="readonly", width=16,
        values=["", "GENERATE FILES", "FILE CREATED", "FILE UPDATED", "FILE DELETED"])
    gf_action_cb.pack(side="left", padx=(0, 8))
    tk.Label(gf_filter_frame, text="Warehouse:", font=("Helvetica", 9)).pack(side="left", padx=(0, 4))
    gf_wh_var = tk.StringVar(value="")
    gf_wh_cb = ttk.Combobox(gf_filter_frame, textvariable=gf_wh_var, state="readonly", width=12,
        values=["", "[W1]", "[W2]", "[YK]"])
    gf_wh_cb.pack(side="left", padx=(0, 8))
    gf_date_from_var = tk.StringVar()
    gf_date_to_var   = tk.StringVar()
    _date_picker_widget(gf_filter_frame, gf_date_from_var, "From:").pack(side="left", padx=(4, 4))
    _date_picker_widget(gf_filter_frame, gf_date_to_var,   "To:").pack(side="left", padx=(0, 6))
    tk.Button(gf_filter_frame, text="↻", width=3, command=lambda: [
        gf_action_var.set(""), gf_wh_var.set(""), gf_date_from_var.set(""), gf_date_to_var.set(""),
        load_gf_data()]).pack(side="left", padx=(0, 10))
    gf_count_label = tk.Label(gf_filter_frame, text="", fg="blue")
    gf_count_label.pack(side="right", padx=8)

    gf_table_frame = tk.Frame(tab_files)
    gf_table_frame.pack(fill="both", expand=True, padx=8, pady=(0, 5))
    gf_vsb = ttk.Scrollbar(gf_table_frame, orient="vertical")
    gf_vsb.pack(side="right", fill="y")
    gf_hsb = ttk.Scrollbar(gf_table_frame, orient="horizontal")
    gf_hsb.pack(side="bottom", fill="x")

    # Tag colours for created vs deleted rows
    tree_gf = ttk.Treeview(gf_table_frame,
                            columns=("Timestamp", "User", "Action", "File Name", "Details"),
                            show="headings",
                            yscrollcommand=gf_vsb.set,
                            xscrollcommand=gf_hsb.set)
    for col, width in [("Timestamp", 140), ("User", 110), ("Action", 130), ("File Name", 180), ("Details", 340)]:
        tree_gf.heading(col, text=col)
        tree_gf.column(col, width=width, anchor="w")
    tree_gf.tag_configure("created",  background="#eafaf1", foreground="#1e8449")  # green tint
    tree_gf.tag_configure("deleted",  background="#fdedec", foreground="#922b21")  # red tint
    tree_gf.tag_configure("updated",  background="#fef9e7", foreground="#9a7d0a")  # yellow tint
    tree_gf.tag_configure("generate", background="#eaf4fb", foreground="#1a5276")  # blue tint
    tree_gf.pack(fill="both", expand=True)
    gf_vsb.config(command=tree_gf.yview)
    gf_hsb.config(command=tree_gf.xview)

    # ── Helper: parse file name from Details field ─────────────
    def _extract_filename(action, details):
        """Return a short display name from the log Details string."""
        details = str(details)
        # FILE CREATED / FILE DELETED: "PDF: name.pdf | Path: ..." or "File: name.xlsx | ..."
        for prefix in ("PDF: ", "Excel: ", "File: "):
            if prefix in details:
                part = details.split(prefix, 1)[1]
                return part.split(" |")[0].strip()
        return ""

    # ── Data loader for Generated Files tab ───────────────────
    def load_gf_data(*_):
        tree_gf.delete(*tree_gf.get_children())
        try:
            df_log = load_logs()
        except Exception:
            gf_count_label.config(text="0 record(s)")
            return
        # Keep only file-related actions
        file_actions = {"GENERATE FILES", "FILE CREATED", "FILE UPDATED", "FILE DELETED"}
        df_log = df_log[df_log["Action"].isin(file_actions)]
        # Sub-filter by action
        act_f = gf_action_var.get().strip()
        if act_f:
            df_log = df_log[df_log["Action"] == act_f]
        # Sub-filter by warehouse tag
        wh_f = gf_wh_var.get().strip()
        if wh_f:
            df_log = df_log[df_log["Details"].str.contains(re.escape(wh_f), na=False)]
        # Date filter
        df_log = _filter_by_date(df_log, gf_date_from_var.get().strip(),
                                 gf_date_to_var.get().strip(), col="Timestamp")
        df_log = df_log.iloc[::-1].reset_index(drop=True)
        for _, row in df_log.iterrows():
            action  = str(row.get("Action", ""))
            details = str(row.get("Details", ""))
            fname   = _extract_filename(action, details)
            tag = "created" if action == "FILE CREATED" else (
                  "deleted" if action == "FILE DELETED" else (
                  "updated" if action == "FILE UPDATED" else "generate"))
            tree_gf.insert("", "end", values=(
                row.get("Timestamp", ""),
                row.get("User", ""),
                action,
                fname,
                details,
            ), tags=(tag,))
        gf_count_label.config(text=f"{len(df_log)} record(s)")

    gf_action_cb.bind("<<ComboboxSelected>>", load_gf_data)
    gf_wh_cb.bind("<<ComboboxSelected>>", load_gf_data)
    gf_date_from_var.trace_add("write", load_gf_data)
    gf_date_to_var.trace_add("write", load_gf_data)

    # ── All Logs helpers ──────────────────────────────────────
    def populate_user_listbox():
        try:
            df_log = load_logs()
            users = sorted(df_log["User"].dropna().unique().tolist())
        except Exception:
            users = []
        user_listbox.delete(0, tk.END)
        user_listbox.insert(tk.END, "(All Users)")
        for u in users:
            user_listbox.insert(tk.END, u)
        user_listbox.selection_set(0)

    def get_selected_user():
        sel = user_listbox.curselection()
        if not sel: return None
        val = user_listbox.get(sel[0])
        return None if val == "(All Users)" else val

    def load_log_data(*_):
        tree_log.delete(*tree_log.get_children())
        df_log = load_logs()
        action_f  = filter_action_var.get().strip()
        user_f    = get_selected_user()
        date_from = log_date_from_var.get().strip()
        date_to   = log_date_to_var.get().strip()
        if user_f:   df_log = df_log[df_log["User"] == user_f]
        if action_f: df_log = df_log[df_log["Action"] == action_f]
        df_log = _filter_by_date(df_log, date_from, date_to, col="Timestamp")
        df_log = df_log.iloc[::-1].reset_index(drop=True)
        for _, row in df_log.iterrows():
            tree_log.insert("", "end", values=tuple(
                row.get(c, "") for c in ["Timestamp", "User", "Action", "Details"]))
        count_label.config(text=f"{len(df_log)} record(s)")

    def reset_filters():
        filter_action_var.set("")
        log_date_from_var.set("")
        log_date_to_var.set("")
        user_listbox.selection_clear(0, tk.END)
        user_listbox.selection_set(0)
        load_log_data()

    filter_action_cb.bind("<<ComboboxSelected>>", lambda e: load_log_data())
    user_listbox.bind("<<ListboxSelect>>", lambda e: load_log_data())
    log_date_from_var.trace_add("write", lambda *_: load_log_data())
    log_date_to_var.trace_add("write", lambda *_: load_log_data())
    populate_user_listbox()
    load_log_data()
    load_gf_data()



# ========== SWITCH USER ==========

def switch_user():
    global current_user, current_is_admin, session_start

    sw_win = tk.Toplevel(root)
    sw_win.title("Switch User — Login")
    sw_win.geometry("320x290")
    sw_win.resizable(False, False)
    sw_win.transient(root)
    sw_win.grab_set()

    hdr = tk.Frame(sw_win, bg="#2c3e50")
    hdr.pack(fill="x")
    tk.Label(hdr, text="Switch User", font=("Helvetica", 10, "bold"),
             bg="#2c3e50", fg="white", pady=8).pack()
    tk.Label(hdr, text="Enter credentials of the account to switch to",
             font=("Helvetica", 8), bg="#2c3e50", fg="#95a5a6", pady=0).pack(pady=(0, 8))

    form = tk.Frame(sw_win, padx=26, pady=14)
    form.pack(fill="x")

    tk.Label(form, text="Username", anchor="w", font=("Helvetica", 9)).grid(row=0, column=0, sticky="w")
    uv = tk.StringVar()
    u_entry = tk.Entry(form, textvariable=uv, width=26, font=("Helvetica", 10))
    u_entry.grid(row=1, column=0, pady=(2, 10), sticky="we")

    tk.Label(form, text="Password", anchor="w", font=("Helvetica", 9)).grid(row=2, column=0, sticky="w")
    pv = tk.StringVar()
    pw_e = tk.Entry(form, textvariable=pv, width=26, show="●", font=("Helvetica", 10))
    pw_e.grid(row=3, column=0, pady=(2, 4), sticky="we")

    show_var = tk.BooleanVar(value=False)
    tk.Checkbutton(form, text="Show password", variable=show_var,
                   command=lambda: pw_e.config(show="" if show_var.get() else "●"),
                   font=("Helvetica", 8)).grid(row=4, column=0, sticky="w")

    err = tk.Label(form, text="", fg="#e74c3c", font=("Helvetica", 8), wraplength=260)
    err.grid(row=5, column=0, sticky="w", pady=(6, 0))

    def do_switch(event=None):
        global current_user, current_is_admin, session_start
        uname = uv.get().strip()
        pw    = pv.get()
        if not uname or not pw:
            err.config(text="Please enter both username and password."); return
        if uname.lower() == current_user.lower():
            err.config(text="You are already logged in as this user."); return
        ok, role = authenticate_user(uname, pw)
        if not ok:
            err.config(text="Invalid username or password."); return
        save_log("LOGOUT", f"Session ended for '{current_user}'")
        current_user     = uname
        current_is_admin = (role == "admin")
        session_start    = _now().strftime("%Y-%m-%d %H:%M:%S")
        _refresh_user_bar()
        save_log("LOGIN", f"Session started by '{current_user}' (role: {role})")
        sw_win.destroy()
        messagebox.showinfo("User Switched", f"Switched to: {current_user}\nRole: {role.capitalize()}")

    pw_e.bind("<Return>", do_switch)
    u_entry.bind("<Return>", lambda e: pw_e.focus_set())
    tk.Button(sw_win, text="SWITCH USER", command=do_switch,
              bg="#2c3e50", fg="white", font=("Helvetica", 9, "bold"),
              pady=4, padx=20).pack(pady=(0, 10))
    u_entry.focus_set()
    sw_win.wait_window()


# ========== LOGIN ==========

def show_login():
    global current_user, current_is_admin, session_start
    initialize_users()

    login_win = tk.Tk()
    login_win.title("Warehouse System — Login")
    login_win.geometry("360x310")
    login_win.resizable(False, False)
    login_win.eval('tk::PlaceWindow . center')

    # ── Header ────────────────────────────────────────────────
    hdr = tk.Frame(login_win, bg="#2c3e50")
    hdr.pack(fill="x")
    tk.Label(hdr, text="Warehouse Management System",
             font=("Helvetica", 12, "bold"), bg="#2c3e50", fg="white",
             pady=12).pack()
    tk.Label(hdr, text="Please log in to continue",
             font=("Helvetica", 9), bg="#2c3e50", fg="#95a5a6",
             pady=0).pack(pady=(0, 10))

    form = tk.Frame(login_win, padx=30, pady=18)
    form.pack(fill="x")

    tk.Label(form, text="Username", anchor="w", font=("Helvetica", 9)).grid(row=0, column=0, sticky="w", pady=(0, 2))
    user_var = tk.StringVar()
    user_entry = tk.Entry(form, textvariable=user_var, width=26, font=("Helvetica", 10))
    user_entry.grid(row=1, column=0, pady=(0, 10), sticky="we")

    tk.Label(form, text="Password", anchor="w", font=("Helvetica", 9)).grid(row=2, column=0, sticky="w", pady=(0, 2))
    pw_var = tk.StringVar()
    pw_entry = tk.Entry(form, textvariable=pw_var, width=26, show="●", font=("Helvetica", 10))
    pw_entry.grid(row=3, column=0, pady=(0, 6), sticky="we")

    show_pw_var = tk.BooleanVar(value=False)
    def _toggle_pw():
        pw_entry.config(show="" if show_pw_var.get() else "●")
    tk.Checkbutton(form, text="Show password", variable=show_pw_var,
                   command=_toggle_pw, font=("Helvetica", 8)).grid(row=4, column=0, sticky="w")

    error_label = tk.Label(form, text="", fg="#e74c3c", font=("Helvetica", 8), wraplength=260)
    error_label.grid(row=5, column=0, pady=(6, 0), sticky="w")

    def attempt_login(event=None):
        global current_user, current_is_admin, session_start
        uname = user_var.get().strip()
        pw    = pw_var.get()
        if not uname or not pw:
            error_label.config(text="Please enter both username and password."); return
        ok, role = authenticate_user(uname, pw)
        if not ok:
            error_label.config(text="Invalid username or password."); return
        current_user     = uname
        current_is_admin = (role == "admin")
        session_start    = _now().strftime("%Y-%m-%d %H:%M:%S")
        login_win.quit()

    btn_row = tk.Frame(login_win, padx=30)
    btn_row.pack(fill="x")
    tk.Button(btn_row, text="LOGIN", command=attempt_login,
              bg="#2c3e50", fg="white", font=("Helvetica", 10, "bold"),
              width=14, pady=4).pack(side="left", padx=(0, 8))

    def open_register():
        _show_register_window(login_win)

    tk.Button(btn_row, text="CREATE ACCOUNT", command=open_register,
              font=("Helvetica", 10), width=17, pady=4).pack(side="left")

    pw_entry.bind("<Return>", attempt_login)
    user_entry.bind("<Return>", lambda e: pw_entry.focus_set())

    def on_close():
        if messagebox.askyesno("Exit", "Exit the Warehouse System?", parent=login_win):
            login_win.destroy()
            import sys; sys.exit(0)

    login_win.protocol("WM_DELETE_WINDOW", on_close)
    user_entry.focus_set()
    login_win.mainloop()
    login_win.destroy()
    initialize_log()
    save_log("LOGIN", f"Session started by '{current_user}' (role: {('admin' if current_is_admin else 'user')})")


def _show_register_window(parent):
    """Registration dialog callable from login screen or admin panel."""
    reg_win = tk.Toplevel(parent)
    reg_win.title("Create Account")
    reg_win.geometry("320x300")
    reg_win.resizable(False, False)
    reg_win.transient(parent)
    reg_win.grab_set()

    hdr = tk.Frame(reg_win, bg="#1a5276")
    hdr.pack(fill="x")
    tk.Label(hdr, text="Create New Account", font=("Helvetica", 10, "bold"),
             bg="#1a5276", fg="white", pady=8).pack()

    form = tk.Frame(reg_win, padx=24, pady=14)
    form.pack(fill="x")

    tk.Label(form, text="Username", anchor="w", font=("Helvetica", 9)).grid(row=0, column=0, sticky="w")
    uv = tk.StringVar()
    tk.Entry(form, textvariable=uv, width=26, font=("Helvetica", 10)).grid(row=1, column=0, pady=(2, 10), sticky="we")

    tk.Label(form, text="Password", anchor="w", font=("Helvetica", 9)).grid(row=2, column=0, sticky="w")
    pv = tk.StringVar()
    pw_e = tk.Entry(form, textvariable=pv, width=26, show="●", font=("Helvetica", 10))
    pw_e.grid(row=3, column=0, pady=(2, 4), sticky="we")

    tk.Label(form, text="Confirm Password", anchor="w", font=("Helvetica", 9)).grid(row=4, column=0, sticky="w")
    cpv = tk.StringVar()
    tk.Entry(form, textvariable=cpv, width=26, show="●", font=("Helvetica", 10)).grid(row=5, column=0, pady=(2, 4), sticky="we")

    err = tk.Label(form, text="", fg="#e74c3c", font=("Helvetica", 8), wraplength=260)
    err.grid(row=6, column=0, sticky="w", pady=(6, 0))

    def do_create():
        pw   = pv.get()
        cpw  = cpv.get()
        if pw != cpw:
            err.config(text="Passwords do not match."); return
        msg = create_account(uv.get(), pw)
        if msg:
            err.config(text=msg); return
        role = "Admin" if _is_admin_password(pw) else "User"
        messagebox.showinfo("Account Created",
            f"Account '{uv.get()}' created successfully!\nRole: {role}", parent=reg_win)
        reg_win.destroy()

    btn_row = tk.Frame(reg_win, pady=10)
    btn_row.pack()
    tk.Button(btn_row, text="CREATE ACCOUNT", command=do_create,
              font=("Helvetica", 10), bg="#1a5276", fg="white", padx=16, pady=4).pack(side="left")
    reg_win.focus_force()
    reg_win.wait_window()



# ========== ADMIN PANEL ==========

def open_admin_panel():
    """Account management window — admin only."""
    if not current_is_admin:
        messagebox.showerror("Access Denied", "Only admin accounts can access the Account Manager.")
        return
    save_log("OPEN ADMIN PANEL", "Admin panel opened")
    panel = tk.Toplevel(root)
    panel.title("Account Manager")
    panel.geometry("700x480")
    panel.resizable(False, False)

    hdr = tk.Frame(panel, bg="#1a5276")
    hdr.pack(fill="x")
    tk.Label(hdr, text="Account Manager", font=("Helvetica", 11, "bold"),
             bg="#1a5276", fg="white", pady=8).pack(side="left", padx=12)
    tk.Label(hdr, text="(Admin only)", font=("Helvetica", 8, "italic"),
             bg="#1a5276", fg="#aed6f1").pack(side="left")

    # ── Table ─────────────────────────────────────────────────
    tbl_frame = tk.Frame(panel, bd=1, relief="sunken")
    tbl_frame.pack(fill="both", expand=True, padx=10, pady=(8, 4))

    sb_y = ttk.Scrollbar(tbl_frame, orient="vertical")
    sb_y.pack(side="right", fill="y")
    tree_acc = ttk.Treeview(tbl_frame,
        columns=("Username", "Role", "Created"),
        show="headings",
        yscrollcommand=sb_y.set)
    for col, w in [("Username", 200), ("Role", 120), ("Created", 180)]:
        tree_acc.heading(col, text=col)
        tree_acc.column(col, width=w, anchor="w")
    tree_acc.pack(fill="both", expand=True)
    sb_y.config(command=tree_acc.yview)

    def _load_table():
        tree_acc.delete(*tree_acc.get_children())
        df = load_users()
        for _, row in df.iterrows():
            role = str(row.get("Role", "user")).capitalize()
            tree_acc.insert("", "end", values=(
                row.get("Username", ""),
                role,
                row.get("Created", ""),
            ))

    # ── Bottom action area ────────────────────────────────────
    act_frame = tk.Frame(panel, padx=10, pady=8)
    act_frame.pack(fill="x")

    # Create account sub-frame
    create_lf = tk.LabelFrame(act_frame, text="Create Account", padx=8, pady=6)
    create_lf.pack(side="left", fill="y", padx=(0, 10))

    tk.Label(create_lf, text="Username", font=("Helvetica", 8)).grid(row=0, column=0, sticky="w")
    new_user_var = tk.StringVar()
    tk.Entry(create_lf, textvariable=new_user_var, width=16, font=("Helvetica", 9)).grid(row=1, column=0, pady=(2, 6))

    tk.Label(create_lf, text="Password", font=("Helvetica", 8)).grid(row=2, column=0, sticky="w")
    new_pw_var = tk.StringVar()
    tk.Entry(create_lf, textvariable=new_pw_var, width=16, show="●", font=("Helvetica", 9)).grid(row=3, column=0, pady=(2, 4))

    tk.Label(create_lf,
        text="Include ! @ # → Admin role",
        font=("Helvetica", 7, "italic"), fg="#7f8c8d").grid(row=4, column=0, sticky="w")

    create_err = tk.Label(create_lf, text="", fg="#e74c3c", font=("Helvetica", 7), wraplength=140)
    create_err.grid(row=5, column=0, sticky="w")

    def do_create():
        uname = new_user_var.get()
        err = create_account(uname, new_pw_var.get())
        if err:
            create_err.config(text=err); return
        role = "Admin" if _is_admin_password(new_pw_var.get()) else "User"
        save_log("CREATE ACCOUNT", f"Username: {uname} | Role: {role}")
        create_err.config(text="")
        new_user_var.set("")
        new_pw_var.set("")
        _load_table()
        messagebox.showinfo("Created", f"Account '{uname}' added as {role}.", parent=panel)

    tk.Button(create_lf, text="CREATE", command=do_create,
              bg="#1a5276", fg="white", font=("Helvetica", 8, "bold"), pady=3).grid(row=6, column=0, pady=(0, 0))

    # Delete / change-password sub-frame
    del_lf = tk.LabelFrame(act_frame, text="Manage Selected Account", padx=8, pady=6)
    del_lf.pack(side="left", fill="y", padx=(0, 10))

    del_err = tk.Label(del_lf, text="", fg="#e74c3c", font=("Helvetica", 7), wraplength=160)
    del_err.grid(row=0, column=0, columnspan=2, sticky="w")

    def do_delete():
        sel = tree_acc.selection()
        if not sel:
            del_err.config(text="Select an account first."); return
        uname = tree_acc.item(sel[0], "values")[0]
        if uname.lower() == current_user.lower():
            del_err.config(text="Cannot delete your own account."); return
        if not messagebox.askyesno("Confirm Delete",
                f"Delete account '{uname}'?\nThis cannot be undone.", parent=panel):
            return
        err = delete_account(uname)
        if err:
            del_err.config(text=err); return
        save_log("DELETE ACCOUNT", f"Username: {uname}")
        del_err.config(text="")
        _load_table()
        messagebox.showinfo("Deleted", f"Account '{uname}' deleted.", parent=panel)

    tk.Button(del_lf, text="DELETE ACCOUNT", command=do_delete,
              bg="#922b21", fg="white", font=("Helvetica", 8, "bold"), pady=3,
              width=16).grid(row=1, column=0, pady=(6, 10), sticky="w")

    tk.Label(del_lf, text="New Password", font=("Helvetica", 8)).grid(row=2, column=0, sticky="w")
    chpw_var = tk.StringVar()
    tk.Entry(del_lf, textvariable=chpw_var, width=18, show="●",
             font=("Helvetica", 9)).grid(row=3, column=0, pady=(2, 4))

    def do_change_pw():
        sel = tree_acc.selection()
        if not sel:
            del_err.config(text="Select an account first."); return
        uname = tree_acc.item(sel[0], "values")[0]
        err = change_password(uname, chpw_var.get())
        if err:
            del_err.config(text=err); return
        save_log("CHANGE PASSWORD", f"Username: {uname}")
        del_err.config(text="")
        chpw_var.set("")
        _load_table()
        messagebox.showinfo("Updated", f"Password for '{uname}' updated.", parent=panel)

    tk.Button(del_lf, text="CHANGE PASSWORD", command=do_change_pw,
              bg="#117a65", fg="white", font=("Helvetica", 8, "bold"), pady=3,
              width=16).grid(row=4, column=0, pady=(4, 0), sticky="w")

    _load_table()


def _refresh_user_bar():
    """Update the user bar label and show/hide admin-only buttons after a user switch."""
    role_badge = " 🔑" if current_is_admin else ""
    user_label.config(text=f"👤  {current_user}{role_badge}")
    session_label.config(text=f"Session started: {session_start}")
    # Show/hide admin buttons
    if current_is_admin:
        activity_log_btn.pack(side="right", padx=4, pady=2)
        admin_panel_btn.pack(side="right", padx=4, pady=2)
    else:
        activity_log_btn.pack_forget()
        admin_panel_btn.pack_forget()


def _guarded_activity_log():
    if not current_is_admin:
        messagebox.showerror("Access Denied", "Only admin accounts can view the Activity Log.")
        return
    open_activity_log()


# ========== UI SETUP ==========

show_login()

root = tk.Tk()
root.title("Warehouse System — Developed by Mark Benjamin (IT Intern 2026)")
root.geometry("1280x780")
root.eval('tk::PlaceWindow . center')

# ── User bar ──────────────────────────────────────────────
user_bar = tk.Frame(root, bg="#2c3e50", height=28)
user_bar.pack(fill="x")

clock_label = tk.Label(user_bar, text="", bg="#2c3e50", fg="#95a5a6", font=("Helvetica", 8))
clock_label.pack(side="right", padx=10, pady=4)

tip(tk.Button(user_bar, text="Change User", command=switch_user,
          bg="#34495e", fg="white", bd=0, padx=10),
    "Switch to a different user session. Credentials required.").pack(side="right", padx=10, pady=2)

admin_panel_btn = tip(tk.Button(user_bar, text="👥 Accounts", command=open_admin_panel,
          bg="#34495e", fg="white", bd=0, padx=10),
    "Manage user accounts (Admin only).")
activity_log_btn = tip(tk.Button(user_bar, text="📋 Activity Log", command=_guarded_activity_log,
          bg="#34495e", fg="white", bd=0, padx=10),
    "View the full history of all actions performed in this system (Admin only).")

# Pack admin buttons only if current user is admin
if current_is_admin:
    activity_log_btn.pack(side="right", padx=4, pady=2)
    admin_panel_btn.pack(side="right", padx=4, pady=2)

role_badge = " 🔑" if current_is_admin else ""
user_label = tk.Label(user_bar, text=f"👤  {current_user}{role_badge}", bg="#2c3e50", fg="white", font=("Helvetica", 9, "bold"))
user_label.pack(side="left", padx=10, pady=4)

session_label = tk.Label(user_bar, text=f"Session started: {session_start}", bg="#2c3e50", fg="#95a5a6", font=("Helvetica", 8))
session_label.pack(side="left", padx=5, pady=4)

def update_clock():
    clock_label.config(text=_now().strftime("%Y-%m-%d %H:%M:%S"))
    root.after(1000, update_clock)
update_clock()

# ── Notebook (tabs) ───────────────────────────────────────
notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=6, pady=6)

tab1 = tk.Frame(notebook)
tab2 = tk.Frame(notebook)
tab3 = tk.Frame(notebook)
notebook.add(tab1, text="  Warehouse 1 — Laptops")
notebook.add(tab2, text="  Warehouse 2 — Computer Peripherals / Equipment")
notebook.add(tab3, text="  Yubikey  ")
# ══════════════════════════════════════════════════════════
#  WAREHOUSE 1 TAB
# ══════════════════════════════════════════════════════════

w1_main = tk.Frame(tab1)
w1_main.pack(fill="both", expand=True, padx=8, pady=8)

w1_row1 = tk.Frame(w1_main)
w1_row1.pack(fill="both", expand=True)

# Item Management
input_frame = tk.LabelFrame(w1_row1, text="Item Management", padx=10, pady=5)
input_frame.pack(side="left", fill="both", padx=5)
tip(input_frame, "Add new items to the staging queue before committing to the warehouse.")

tk.Label(input_frame, text="Hostname").grid(row=0, column=0, sticky="w")
hostname_entry = tk.Entry(input_frame, width=22); hostname_entry.grid(row=0, column=1, pady=3)
tip(hostname_entry, "Enter the device hostname (e.g. PC-001). Must be unique.")

serial_entry = tk.Entry(input_frame, width=22)  # kept as hidden stub so downstream code doesn't break

tk.Label(input_frame, text="Checked By").grid(row=2, column=0, sticky="w")
checked_by_entry = tk.Entry(input_frame, width=22); checked_by_entry.grid(row=2, column=1, pady=3)
tip(checked_by_entry, "Name of the person who physically checked this item.")

tk.Label(input_frame, text="Shelf").grid(row=3, column=0, sticky="w")
shelf_var = tk.StringVar()
shelf_dropdown = ttk.Combobox(input_frame, textvariable=shelf_var, width=19)
shelf_dropdown.grid(row=3, column=1, pady=3)
tip(shelf_dropdown, "Select the shelf/area where this item will be stored.")

tk.Label(input_frame, text="Status").grid(row=4, column=0, sticky="w")
remarks_var = tk.StringVar()
ttk.Combobox(input_frame, textvariable=remarks_var, values=STATUS_CHOICES, width=19, state="readonly").grid(row=4, column=1, pady=3)

tk.Label(input_frame, text="Remarks").grid(row=5, column=0, sticky="w")
remarks_text_var = tk.StringVar()
remarks_text_entry = tk.Entry(input_frame, textvariable=remarks_text_var, width=22)
remarks_text_entry.grid(row=5, column=1, pady=3)
tip(remarks_text_entry, "Optional notes about the item's condition.")

crud_frame = tk.Frame(input_frame)
crud_frame.grid(row=6, column=0, columnspan=2, pady=5)
tip(tk.Button(crud_frame, text="STAGE ITEM",  command=put_item,    bg="#2980b9", fg="white", width=14), "Add this item to the staging queue.").grid(row=0, column=0, padx=3)
tip(tk.Button(crud_frame, text="UPDATE ITEM", command=update_item, bg="#1a5276", fg="white", width=13), "Update the selected staged item with new values.").grid(row=0, column=1, padx=3)
tip(tk.Button(crud_frame, text="↻",           command=reset_ui,    width=3), "Clear all input fields and reset the view.").grid(row=0, column=2, padx=3)
tip(tk.Button(crud_frame, text="📥 IMPORT EXCEL", command=import_excel_to_staging, bg="#1e8449", fg="white", width=16),
    "Import items from an Excel file directly into the staging list.").grid(row=1, column=0, columnspan=3, pady=(4,0))

tk.Label(input_frame, text="Staged Items (Click to Edit)", fg="green", font=("Helvetica", 9, "bold")).grid(row=7, column=0, columnspan=2, sticky="w")
staged_listbox = tk.Listbox(input_frame, width=32)
staged_listbox.grid(row=8, column=0, columnspan=2, sticky="nswe", pady=3)
staged_listbox.bind("<<ListboxSelect>>", select_staged_item)
tip(staged_listbox, "Items waiting to be committed. Click a row to load it back into the fields for editing.")
input_frame.rowconfigure(8, weight=1)
input_frame.columnconfigure(0, weight=1)
input_frame.columnconfigure(1, weight=1)

staging_btn_frame = tk.Frame(input_frame)
staging_btn_frame.grid(row=9, column=0, columnspan=2, pady=3)
tip(tk.Button(staging_btn_frame, text="CLEAR ITEMS",   command=remove_from_staging, width=13),
    "Remove the selected staged item (or all if none selected).").pack(side="left", padx=2)
tip(tk.Button(staging_btn_frame, text="↻", command=lambda: staged_listbox.selection_clear(0, tk.END), width=3),
    "Deselect the current staged item so CLEAR ITEMS will clear all.").pack(side="left", padx=2)
tip(tk.Button(staging_btn_frame, text="PUT WAREHOUSE", command=put_warehouse,       width=13),
    "Commit all staged items to Warehouse 1 and generate QR codes. Use GENERATE FILES to create PDF labels.").pack(side="left", padx=2)

# Shelf Controls W1
shelf_mid_frame = tk.Frame(w1_row1)
shelf_mid_frame.pack(side="left", fill="both", expand=True, padx=5)

# Shelf + View sub-row (side by side inside shelf_mid_frame)
w1_shelf_view_row = tk.Frame(shelf_mid_frame)
w1_shelf_view_row.pack(fill="x")

shelf_control_frame = tk.LabelFrame(w1_shelf_view_row, text="Shelf Control & Management", padx=10, pady=5)
shelf_control_frame.pack(side="left", fill="both", expand=True)
tip(shelf_control_frame, "Manage shelf availability and add/remove shelves for Warehouse 1.")

status_control_frame = tk.LabelFrame(shelf_control_frame, text="Status Control", padx=8, pady=5)
status_control_frame.pack(fill="x", pady=(0, 5))
tip(status_control_frame, "Mark a shelf as FULL (no more items) or AVAILABLE.")
shelf_control_var = tk.StringVar()
shelf_control_dropdown = ttk.Combobox(status_control_frame, textvariable=shelf_control_var, width=22, state="readonly")
shelf_control_dropdown.pack(side="left", padx=5)
tip(shelf_control_dropdown, "Select the shelf whose status you want to change.")
tip(tk.Button(status_control_frame, text="SET FULL",      command=lambda: set_shelf_status("FULL"),      width=10),
    "Mark selected shelf as FULL — prevents new items from being placed there.").pack(side="left", padx=3)
tip(tk.Button(status_control_frame, text="SET AVAILABLE", command=lambda: set_shelf_status("AVAILABLE"), width=12),
    "Mark selected shelf as AVAILABLE — allows new items to be placed.").pack(side="left", padx=3)
tip(tk.Button(status_control_frame, text="↻",             command=reset_shelf_control,                   width=3),
    "Clear the shelf selection.").pack(side="left", padx=3)

add_remove_frame = tk.LabelFrame(shelf_control_frame, text="Add / Remove", padx=8, pady=5)
add_remove_frame.pack(fill="x")
tip(add_remove_frame, "Add a new shelf by typing its name, or remove an existing empty shelf.")
remove_shelf_var = tk.StringVar()
remove_shelf_dropdown = ttk.Combobox(add_remove_frame, textvariable=remove_shelf_var, width=22)
remove_shelf_dropdown.pack(side="left", padx=5)
tip(remove_shelf_dropdown, "Type a new shelf name to add, or select an existing one to remove.")
tip(tk.Button(add_remove_frame, text="ADD",    command=add_shelf   ), "Add the typed shelf name as a new shelf.").pack(side="left", padx=3)
tip(tk.Button(add_remove_frame, text="REMOVE", command=remove_shelf ), "Remove the selected shelf (only if it has no items).").pack(side="left", padx=3)
tip(tk.Button(add_remove_frame, text="↻",      command=reset_shelf_addition, width=3), "Clear the shelf name field.").pack(side="left", padx=3)

# View W1 — sits to the right of Shelf Control inside shelf_mid_frame
view_frame = tk.LabelFrame(w1_shelf_view_row, text="View", padx=10, pady=5)
view_frame.pack(side="left", fill="y", padx=(5, 0))
tip(view_frame, "Switch between different table views for Warehouse 1.")
for text, cmd, tooltip in [
    ("SHOW WAREHOUSE", show_warehouse,  "View all items currently stored in Warehouse 1."),
    ("SHELF STATUS",   show_available,  "View each shelf's status and how many items it holds."),
    ("PULL HISTORY",   show_pullouts,   "View items that have been pulled out of Warehouse 1."),
    ("QR LABELS",      lambda: open_label_manager(warehouse=1), "Open and manage printable QR label PDF files."),
    ("VIEW EXCEL",     w1_view_excel,   "Open the last Excel file generated by GENERATE FILES for Warehouse 1."),
]:
    tip(tk.Button(view_frame, text=text, command=cmd, width=15), tooltip).pack(anchor="w", pady=3)

# Search & Filter W1
w1_pullout_frame = tk.LabelFrame(shelf_mid_frame, text="Warehouse 1", padx=10, pady=8)
w1_pullout_frame.pack(fill="x", pady=5)

w1_search_filter = tk.LabelFrame(w1_pullout_frame, text="Search & Filter", padx=8, pady=5)
w1_search_filter.pack(fill="x", pady=(0, 5))

# Row 1: search + shelf + remarks + single 🔍 button + clear
w1_sf_row1 = tk.Frame(w1_search_filter)
w1_sf_row1.pack(fill="x", pady=(0, 3))
tk.Label(w1_sf_row1, text="Search:").pack(side="left", padx=(5, 2))
search_entry = tk.Entry(w1_sf_row1, width=20); search_entry.pack(side="left", padx=(0, 2))
search_entry.bind("<Return>", lambda e: search_item())
search_entry.bind("<KeyRelease>", pull_search_live)

tk.Label(w1_sf_row1, text="Shelf:").pack(side="left", padx=(5, 2))
pull_shelf_var = tk.StringVar()
pull_shelf_dropdown = ttk.Combobox(w1_sf_row1, textvariable=pull_shelf_var, width=16, state="readonly")
pull_shelf_dropdown.pack(side="left", padx=(0, 8))

tk.Label(w1_sf_row1, text="Remarks:").pack(side="left", padx=(5, 2))
pull_remarks_var = tk.StringVar()
ttk.Combobox(w1_sf_row1, textvariable=pull_remarks_var, values=[""] + STATUS_CHOICES, width=14, state="readonly").pack(side="left", padx=(0, 8))

tip(tk.Button(w1_sf_row1, text="🔍", command=search_item,       width=3), "Search and filter the warehouse table.").pack(side="left", padx=3)
tip(tk.Button(w1_sf_row1, text="↻",  command=clear_pull_filters, width=3), "Clear all search and filter fields.").pack(side="left", padx=3)

# Row 2: date range (calendar pickers)
w1_sf_row2 = tk.Frame(w1_search_filter)
w1_sf_row2.pack(fill="x", pady=(0, 2))
w1_date_from_var = tk.StringVar()
w1_date_to_var   = tk.StringVar()
_date_picker_widget(w1_sf_row2, w1_date_from_var, "From:").pack(side="left", padx=(5, 8))
_date_picker_widget(w1_sf_row2, w1_date_to_var,   "To:").pack(side="left", padx=(0, 5))

pull_item_entry = search_entry

# Pull Out W1
w1_pull_action = tk.LabelFrame(w1_pullout_frame, text="Pull Out", padx=8, pady=5)
w1_pull_action.pack(fill="x", pady=(5, 0))

# Pull Out Row 1: pull reason dropdown + redo + warehouse pull buttons only
w1_po_row1 = tk.Frame(w1_pull_action)
w1_po_row1.pack(fill="x", pady=(0, 3))
tk.Label(w1_po_row1, text="Pull Reason:").pack(side="left", padx=(5, 2))
pull_reason_filter_var = tk.StringVar()
pull_reason_filter_entry = ttk.Combobox(w1_po_row1, textvariable=pull_reason_filter_var, width=20)
pull_reason_filter_entry.pack(side="left", padx=(0, 8))
w1_pull_date_from_var = tk.StringVar()
w1_pull_date_to_var   = tk.StringVar()
tip(tk.Button(w1_po_row1, text="↻",  command=reset_pull_out,      width=3),  "Clear pull-out fields and reset the view.").pack(side="left", padx=2)
tip(tk.Button(w1_po_row1, text="WAREHOUSE PULL", command=pull_item, width=16), "Pull the selected item out of Warehouse 1. A pull reason is required.").pack(side="left", padx=(10, 3))

# Status bar W1
w1_status_bar = tk.Frame(shelf_mid_frame)
w1_status_bar.pack(fill="x")
w1_full_label   = tk.Label(w1_status_bar, text="FULL Shelves: None", fg="red");  w1_full_label.pack(side="left", padx=10)
w1_search_label = tk.Label(w1_status_bar, text="", fg="blue");                   w1_search_label.pack(side="left", padx=10)
w1_status_label = tk.Label(w1_status_bar, text="", fg="green");                  w1_status_label.pack(side="left", padx=10)

# ── W1 Table toolbar (Select All + Stored QR) ──────────────
w1_table_toolbar = tk.Frame(shelf_mid_frame)
w1_table_toolbar.pack(fill="x", padx=5, pady=(2, 0))

def w1_toggle_select_all():
    all_iids = list(tree_warehouse.get_children())
    checked  = [iid for iid in all_iids if w1_row_checks.get(iid)]
    # If pull history is visible, toggle that instead
    if tree_pullouts.winfo_ismapped():
        all_iids = list(tree_pullouts.get_children())
        checked  = [iid for iid in all_iids if w1_pull_row_checks.get(iid)]
        new_state = len(checked) < len(all_iids)
        for iid in all_iids:
            w1_pull_row_checks[iid] = new_state
            tree_pullouts.set(iid, "CP0", "☑" if new_state else "☐")
        return
    new_state = len(checked) < len(all_iids)
    for iid in all_iids:
        w1_row_checks[iid] = new_state
        hostname = tree_warehouse.item(iid, "values")[2]
        if new_state:
            w1_persistent_checks.add(hostname)
        else:
            w1_persistent_checks.discard(hostname)
        tree_warehouse.set(iid, "C0", "☑" if new_state else "☐")
    _w1_refresh_select_all_label()

w1_select_all_btn = tk.Button(w1_table_toolbar, text="SELECT ALL",
    command=w1_toggle_select_all, width=12)
w1_select_all_btn.pack(side="left", padx=(0, 6))
tip(w1_select_all_btn, "Select or deselect all visible rows in the active table.")
tip(tk.Button(w1_table_toolbar, text="GENERATE FILES", command=w1_generate_stored_qr,
          bg="#6c3483", fg="white", width=15),
    "Generate QR PNGs, PDF labels and Excel export for selected items.").pack(side="left", padx=(0, 6))
tip(tk.Button(w1_table_toolbar, text="VIEW QR", command=w1_view_stored_qr,
          bg="#1a5276", fg="white", width=10),
    "View all existing QR codes without regenerating.").pack(side="left", padx=(0, 6))
w1_back_to_stage_btn = tip(tk.Button(w1_table_toolbar, text="BACK TO STAGE", command=unstage_from_warehouse,
          bg="#117a65", fg="white", width=15),
    "Move checked warehouse items back to the Item Management staging list.")
w1_back_to_stage_btn.pack(side="left", padx=(0, 6))
w1_back_to_wh_btn = tip(tk.Button(w1_table_toolbar, text="BACK TO WAREHOUSE", command=undo_pull,
          bg="#922b21", fg="white", width=18),
    "Restore checked pull history items back to Warehouse 1.")
# Hidden by default; shown only when Pull History view is active

# Tables W1
w1_table_frame = tk.Frame(shelf_mid_frame)
w1_table_frame.pack(fill="both", expand=True, pady=5)

tree_warehouse = ttk.Treeview(w1_table_frame, columns=("C0","C1","C2","C3","C4","C5","C6","C7"), show='headings')
for col, text, width in zip(("C0","C1","C2","C3","C4","C5","C6","C7"),
    ("✔","QR","Hostname","Checked By","Shelf","Status","Remarks","Date"),
    (30,180,150,115,130,95,145,145)):
    tree_warehouse.heading(col, text=text); tree_warehouse.column(col, width=width)
tree_warehouse.column("C0", anchor="center", stretch=False)
tree_warehouse.bind("<<TreeviewSelect>>", select_item)

tree_available = ttk.Treeview(w1_table_frame, columns=("C1","C2","C3","C4"), show='headings')
for col, text, width in zip(("C1","C2","C3","C4"), ("Shelf","Status","Items Stored","Date Set Full"), (250,140,110,200)):
    tree_available.heading(col, text=text); tree_available.column(col, width=width)

tree_pullouts = ttk.Treeview(w1_table_frame, columns=("CP0","C1","C2","C3","C4","C5","C6"), show='headings')
for col, text, width in zip(("CP0","C1","C2","C3","C4","C5","C6"),
    ("✔","Hostname","Shelf","Status","Remarks","Pull Reason","Date"), (30,145,125,90,155,205,150)):
    tree_pullouts.heading(col, text=text); tree_pullouts.column(col, width=width)
tree_pullouts.column("CP0", anchor="center", stretch=False)
tree_pullouts.bind("<<TreeviewSelect>>", select_pull_item)

tree_qr = ttk.Treeview(w1_table_frame, columns=("C1","C2","C3"), show='headings')
for col, text, width in zip(("C1","C2","C3"),
    ("Hostname","QR UUID String","File Status (PNG)"), (200,400,150)):
    tree_qr.heading(col, text=text); tree_qr.column(col, width=width)

# ══════════════════════════════════════════════════════════
#  WAREHOUSE 2 TAB
# ══════════════════════════════════════════════════════════

w2_main = tk.Frame(tab2)
w2_main.pack(fill="both", expand=True, padx=8, pady=8)

w2_row1 = tk.Frame(w2_main)
w2_row1.pack(fill="both", expand=True)

# Equipment selection + staging panel
w2_input_frame = tk.LabelFrame(w2_row1, text="Set Staging", padx=10, pady=5)
w2_input_frame.pack(side="left", fill="both", padx=5)
tip(w2_input_frame, "Build equipment sets (Monitor, Keyboard, Mouse, Headset) and stage them before committing to Warehouse 2.")

tip(tk.Label(w2_input_frame, text="Select Equipment:", font=("Helvetica", 9, "bold")),
    "Check the equipment types to include in this set.").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 4))

w2_equip_vars = {}
for i, eq in enumerate(EQUIPMENT_TYPES):
    var = tk.BooleanVar()
    w2_equip_vars[eq] = var
    tk.Checkbutton(w2_input_frame, text=eq, variable=var, width=10, anchor="w").grid(
        row=1 + i // 2, column=i % 2, sticky="w", padx=4)

tip(tk.Button(w2_input_frame, text="BUILD SET", command=w2_build_set,
          bg="#2980b9", fg="white", width=28),
    "Open a form to enter details for each selected equipment type, then stage the set.").grid(row=3, column=0, columnspan=2, pady=(8, 4))

tip(tk.Label(w2_input_frame, text="Staged Sets", fg="green", font=("Helvetica", 9, "bold")),
    "Sets waiting to be committed to Warehouse 2.").grid(row=4, column=0, columnspan=2, sticky="w")
w2_staged_listbox = tk.Listbox(w2_input_frame, width=34)
w2_staged_listbox.grid(row=5, column=0, columnspan=2, sticky="nswe", pady=3)
tip(w2_staged_listbox, "Click a set to select it for editing. ⚠ means a required equipment type is missing.")
w2_input_frame.rowconfigure(5, weight=1)
w2_input_frame.columnconfigure(0, weight=1)
w2_input_frame.columnconfigure(1, weight=1)

w2_stage_btns = tk.Frame(w2_input_frame)
w2_stage_btns.grid(row=6, column=0, columnspan=2, pady=3)
tip(tk.Button(w2_stage_btns, text="CLEAR SETS",    command=w2_remove_staged_set, width=13),
    "Remove the selected staged set (or all sets if none selected).").pack(side="left", padx=2)
tip(tk.Button(w2_stage_btns, text="↻", command=lambda: w2_staged_listbox.selection_clear(0, tk.END), width=3),
    "Deselect the current staged set so CLEAR SETS will clear all.").pack(side="left", padx=2)
tip(tk.Button(w2_stage_btns, text="PUT WAREHOUSE", command=w2_put_warehouse,     width=13),
    "Commit all staged sets to Warehouse 2 and generate QR codes. Use GENERATE FILES to create PDF labels.").pack(side="left", padx=2)

w2_import_btn_frame = tk.Frame(w2_input_frame)
w2_import_btn_frame.grid(row=7, column=0, columnspan=2, pady=(0, 4))
tip(tk.Button(w2_import_btn_frame, text="📥 IMPORT EXCEL", command=w2_import_excel_to_staging,
          bg="#1e8449", fg="white", width=20),
    "Import items from an Excel file directly into the W2 staging list.").pack()
w2_item_mgmt_frame = tk.LabelFrame(w2_input_frame, text="Item Management", padx=6, pady=4)
w2_item_mgmt_frame.grid(row=8, column=0, columnspan=2, pady=(6, 0), sticky="we")
tip(w2_item_mgmt_frame, "Edit a staged set's details before committing it to the warehouse.")
tk.Label(w2_item_mgmt_frame,
         text="Select a row in the staged sets table,\nthen use buttons below:",
         font=("Helvetica", 8), fg="gray", justify="left").pack(anchor="w", pady=(0, 4))
w2_item_action_btns = tk.Frame(w2_item_mgmt_frame)
w2_item_action_btns.pack()
tip(tk.Button(w2_item_action_btns, text="UPDATE ITEM", command=w2_update_item,
          bg="#1a5276", fg="white", width=13),
    "Edit the items in the selected staged set.").pack(side="left", padx=2)

# Shelf Controls W2
w2_shelf_mid = tk.Frame(w2_row1)
w2_shelf_mid.pack(side="left", fill="both", expand=True, padx=5)

# Shelf + View sub-row (side by side inside w2_shelf_mid)
w2_shelf_view_row = tk.Frame(w2_shelf_mid)
w2_shelf_view_row.pack(fill="x")

w2_shelf_ctrl_frame = tk.LabelFrame(w2_shelf_view_row, text="Shelf Control & Management", padx=10, pady=5)
w2_shelf_ctrl_frame.pack(side="left", fill="both", expand=True)
tip(w2_shelf_ctrl_frame, "Manage shelf availability and add/remove shelves for Warehouse 2.")

w2_status_ctrl = tk.LabelFrame(w2_shelf_ctrl_frame, text="Status Control", padx=8, pady=5)
w2_status_ctrl.pack(fill="x", pady=(0, 5))
tip(w2_status_ctrl, "Mark a shelf as FULL (no more items) or AVAILABLE.")
w2_shelf_control_var = tk.StringVar()
w2_shelf_control_dropdown = ttk.Combobox(w2_status_ctrl, textvariable=w2_shelf_control_var, width=22, state="readonly")
w2_shelf_control_dropdown.pack(side="left", padx=5)
tip(w2_shelf_control_dropdown, "Select the shelf whose status you want to change.")
tip(tk.Button(w2_status_ctrl, text="SET FULL",      command=lambda: w2_set_shelf_status("FULL"),      width=10),
    "Mark selected shelf as FULL — prevents new items from being placed there.").pack(side="left", padx=3)
tip(tk.Button(w2_status_ctrl, text="SET AVAILABLE", command=lambda: w2_set_shelf_status("AVAILABLE"), width=12),
    "Mark selected shelf as AVAILABLE — allows new items to be placed.").pack(side="left", padx=3)
tip(tk.Button(w2_status_ctrl, text="↻",             command=w2_reset_shelf_control,                   width=3),
    "Clear the shelf selection.").pack(side="left", padx=3)

w2_add_remove = tk.LabelFrame(w2_shelf_ctrl_frame, text="Add / Remove", padx=8, pady=5)
w2_add_remove.pack(fill="x")
tip(w2_add_remove, "Add a new shelf by typing its name, or remove an existing empty shelf.")
w2_remove_shelf_var = tk.StringVar()
w2_remove_shelf_dropdown = ttk.Combobox(w2_add_remove, textvariable=w2_remove_shelf_var, width=22)
w2_remove_shelf_dropdown.pack(side="left", padx=5)
tip(w2_remove_shelf_dropdown, "Type a new shelf name to add, or select an existing one to remove.")
tip(tk.Button(w2_add_remove, text="ADD",    command=w2_add_shelf   ), "Add the typed shelf name as a new shelf.").pack(side="left", padx=3)
tip(tk.Button(w2_add_remove, text="REMOVE", command=w2_remove_shelf ), "Remove the selected shelf (only if it has no items).").pack(side="left", padx=3)
tip(tk.Button(w2_add_remove, text="↻",      command=w2_reset_shelf_addition, width=3), "Clear the shelf name field.").pack(side="left", padx=3)

# View W2 — sits to the right of Shelf Control inside w2_shelf_mid
w2_view_frame = tk.LabelFrame(w2_shelf_view_row, text="View", padx=10, pady=5)
w2_view_frame.pack(side="left", fill="y", padx=(5, 0))
tip(w2_view_frame, "Switch between different table views for Warehouse 2.")
for text, cmd, tooltip in [
    ("SHOW WAREHOUSE", w2_show_warehouse, "View all items currently stored in Warehouse 2."),
    ("SHELF STATUS",   w2_show_available, "View each shelf's status and how many items it holds."),
    ("PULL HISTORY",   w2_show_pullouts,  "View items that have been pulled out of Warehouse 2."),
    ("QR LABELS",      lambda: open_label_manager(warehouse=2), "Open and manage printable QR label PDF files."),
    ("VIEW EXCEL",     w2_view_excel,     "Open the last Excel file generated by GENERATE FILES for Warehouse 2."),
]:
    tip(tk.Button(w2_view_frame, text=text, command=cmd, width=15), tooltip).pack(anchor="w", pady=3)

# Search & Filter W2
w2_pullout_frame = tk.LabelFrame(w2_shelf_mid, text="Warehouse 2", padx=10, pady=8)
w2_pullout_frame.pack(fill="x", pady=5)

w2_search_filter = tk.LabelFrame(w2_pullout_frame, text="Search & Filter", padx=8, pady=5)
w2_search_filter.pack(fill="x", pady=(0, 5))

# Row 1: search + shelf + type + single 🔍 button + clear
w2_sf_row1 = tk.Frame(w2_search_filter)
w2_sf_row1.pack(fill="x", pady=(0, 3))
tk.Label(w2_sf_row1, text="Search:").pack(side="left", padx=(5, 2))
w2_search_entry = tk.Entry(w2_sf_row1, width=18); w2_search_entry.pack(side="left", padx=(0, 2))
w2_search_entry.bind("<Return>", lambda e: w2_search_item())
w2_search_entry.bind("<KeyRelease>", w2_pull_search_live)

tk.Label(w2_sf_row1, text="Shelf:").pack(side="left", padx=(5, 2))
w2_pull_shelf_var = tk.StringVar()
w2_pull_shelf_dropdown = ttk.Combobox(w2_sf_row1, textvariable=w2_pull_shelf_var, width=16, state="readonly")
w2_pull_shelf_dropdown.pack(side="left", padx=(0, 10))

tk.Label(w2_sf_row1, text="Type:").pack(side="left", padx=(5, 2))
w2_type_filter_var = tk.StringVar()
ttk.Combobox(w2_sf_row1, textvariable=w2_type_filter_var,
             values=[""] + EQUIPMENT_TYPES, width=12, state="readonly").pack(side="left", padx=(0, 8))

tip(tk.Button(w2_sf_row1, text="🔍", command=w2_search_item,  width=3), "Search and filter the Warehouse 2 table.").pack(side="left", padx=3)
tip(tk.Button(w2_sf_row1, text="↻",  command=w2_clear_filters, width=3), "Clear all search and filter fields.").pack(side="left", padx=3)

# Row 2: date range (calendar pickers)
w2_sf_row2 = tk.Frame(w2_search_filter)
w2_sf_row2.pack(fill="x", pady=(0, 2))
w2_date_from_var = tk.StringVar()
w2_date_to_var   = tk.StringVar()
_date_picker_widget(w2_sf_row2, w2_date_from_var, "From:").pack(side="left", padx=(5, 8))
_date_picker_widget(w2_sf_row2, w2_date_to_var,   "To:").pack(side="left", padx=(0, 5))

w2_pull_item_entry = w2_search_entry

# Pull Out W2
w2_pull_action = tk.LabelFrame(w2_pullout_frame, text="Pull Out", padx=8, pady=5)
w2_pull_action.pack(fill="x", pady=(5, 0))

# Pull Out Row 1: pull reason dropdown + redo + warehouse pull buttons only
w2_po_row1 = tk.Frame(w2_pull_action)
w2_po_row1.pack(fill="x", pady=(0, 3))
tk.Label(w2_po_row1, text="Pull Reason:").pack(side="left", padx=(5, 2))
w2_pull_reason_filter_var = tk.StringVar()
w2_pull_reason_filter_entry = ttk.Combobox(w2_po_row1, textvariable=w2_pull_reason_filter_var, width=20)
w2_pull_reason_filter_entry.pack(side="left", padx=(0, 8))
w2_pull_date_from_var = tk.StringVar()
w2_pull_date_to_var   = tk.StringVar()
tip(tk.Button(w2_po_row1, text="↻",  command=w2_reset_pull_out,      width=3),  "Clear pull-out fields and reset the view.").pack(side="left", padx=2)
tip(tk.Button(w2_po_row1, text="WAREHOUSE PULL", command=w2_pull_item, width=16), "Pull the selected item out of Warehouse 2. A pull reason is required.").pack(side="left", padx=(10, 3))

# Status bar W2
w2_status_bar = tk.Frame(w2_shelf_mid)
w2_status_bar.pack(fill="x")
w2_full_label   = tk.Label(w2_status_bar, text="FULL Shelves: None", fg="red");  w2_full_label.pack(side="left", padx=10)
w2_search_label = tk.Label(w2_status_bar, text="", fg="blue");                   w2_search_label.pack(side="left", padx=10)
w2_status_label = tk.Label(w2_status_bar, text="", fg="green");                  w2_status_label.pack(side="left", padx=10)

# ── W2 Table toolbar (Select All + Stored QR) ──────────────
w2_table_toolbar = tk.Frame(w2_shelf_mid)
w2_table_toolbar.pack(fill="x", padx=5, pady=(2, 0))

def w2_toggle_select_all():
    # If pull history is visible, toggle that instead
    if tree_w2_pullouts.winfo_ismapped():
        all_iids = list(tree_w2_pullouts.get_children())
        checked  = [iid for iid in all_iids if w2_pull_row_checks.get(iid)]
        new_state = len(checked) < len(all_iids)
        for iid in all_iids:
            w2_pull_row_checks[iid] = new_state
            tree_w2_pullouts.set(iid, "CP0", "☑" if new_state else "☐")
        return
    all_iids = list(tree_w2_warehouse.get_children())
    checked  = [iid for iid in all_iids if w2_row_checks.get(iid)]
    new_state = len(checked) < len(all_iids)
    for iid in all_iids:
        w2_row_checks[iid] = new_state
        key = (tree_w2_warehouse.item(iid, "values")[2], tree_w2_warehouse.item(iid, "values")[4])
        if new_state:
            w2_persistent_checks.add(key)
        else:
            w2_persistent_checks.discard(key)
        tree_w2_warehouse.set(iid, "C0", "☑" if new_state else "☐")
    _w2_refresh_select_all_label()

w2_select_all_btn = tk.Button(w2_table_toolbar, text="SELECT ALL",
    command=w2_toggle_select_all, width=12)
w2_select_all_btn.pack(side="left", padx=(0, 6))
tip(w2_select_all_btn, "Select or deselect all visible rows in the active table.")

tip(tk.Button(w2_table_toolbar, text="GENERATE FILES", command=w2_generate_stored_qr,
          bg="#6c3483", fg="white", width=15),
    "Generate QR PNGs, PDF labels and Excel export for selected items.").pack(side="left", padx=(0, 6))
tip(tk.Button(w2_table_toolbar, text="VIEW QR", command=w2_view_stored_qr,
          bg="#1a5276", fg="white", width=10),
    "View all existing QR codes without regenerating.").pack(side="left", padx=(0, 6))
w2_back_to_stage_btn = tip(tk.Button(w2_table_toolbar, text="BACK TO STAGE", command=w2_unstage_from_warehouse,
          bg="#117a65", fg="white", width=15),
    "Move checked warehouse items back to the Set Staging list.")
w2_back_to_stage_btn.pack(side="left", padx=(0, 6))
w2_back_to_wh_btn = tip(tk.Button(w2_table_toolbar, text="BACK TO WAREHOUSE", command=w2_undo_pull,
          bg="#922b21", fg="white", width=18),
    "Restore checked pull history items back to Warehouse 2.")
# Hidden by default; shown only when Pull History view is active

# Tables W2
w2_table_frame = tk.Frame(w2_shelf_mid)
w2_table_frame.pack(fill="both", expand=True, pady=5)

tree_w2_warehouse = ttk.Treeview(w2_table_frame,
    columns=("C0","C1","C2","C3","C4","C5","C6","C7","C8","C9","C10"), show='headings')
for col, text, width in zip(
    ("C0","C1","C2","C3","C4","C5","C6","C7","C8","C9","C10"),
    ("✔","QR","Set ID","Hostname","Equipment Type","Serial Number","Checked By","Shelf","Status","Remarks","Date"),
    (30,170,90,120,115,115,115,125,90,150,135)):
    tree_w2_warehouse.heading(col, text=text); tree_w2_warehouse.column(col, width=width)
tree_w2_warehouse.column("C0", anchor="center", stretch=False)
tree_w2_warehouse.bind("<<TreeviewSelect>>", w2_select_item)

tree_w2_available = ttk.Treeview(w2_table_frame, columns=("C1","C2","C3","C4"), show='headings')
for col, text, width in zip(("C1","C2","C3","C4"), ("Shelf","Status","Items Stored","Date Set Full"), (250,140,110,200)):
    tree_w2_available.heading(col, text=text); tree_w2_available.column(col, width=width)

tree_w2_pullouts = ttk.Treeview(w2_table_frame,
    columns=("CP0","C1","C2","C3","C4","C5","C6","C7","C8","C9"), show='headings')
for col, text, width in zip(
    ("CP0","C1","C2","C3","C4","C5","C6","C7","C8","C9"),
    ("✔","Set ID","Hostname","Equipment Type","Serial Number","Shelf","Status","Remarks","Pull Reason","Date"),
    (30,85,115,115,115,110,80,150,180,135)):
    tree_w2_pullouts.heading(col, text=text); tree_w2_pullouts.column(col, width=width)
tree_w2_pullouts.column("CP0", anchor="center", stretch=False)
tree_w2_pullouts.bind("<<TreeviewSelect>>", w2_select_pull_item)

tree_w2_qr = ttk.Treeview(w2_table_frame, columns=("C1","C2","C3","C4"), show='headings')
for col, text, width in zip(("C1","C2","C3","C4"),
    ("Set ID","Equipment Type","QR UUID String","File Status (PNG)"), (100,120,380,130)):
    tree_w2_qr.heading(col, text=text); tree_w2_qr.column(col, width=width)
# ══════════════════════════════════════════════════════════
#  YUBIKEY TAB
# ══════════════════════════════════════════════════════════

yk_main = tk.Frame(tab3)
yk_main.pack(fill="both", expand=True, padx=8, pady=8)

yk_row1 = tk.Frame(yk_main)
yk_row1.pack(fill="both", expand=True)

# ── Item Staging Panel (left side) ────────────────────────
yk_input_frame = tk.LabelFrame(yk_row1, text="Item Staging", padx=10, pady=5)
yk_input_frame.pack(side="left", fill="both", padx=5)
tip(yk_input_frame, "Enter Yubikey details and stage them before committing to the warehouse.")

fields_frame = tk.Frame(yk_input_frame)
fields_frame.pack(fill="x", pady=(4, 2))

tk.Label(fields_frame, text="Hostname:",      anchor="w", width=13).grid(row=0, column=0, sticky="w", pady=3, padx=(0,4))
yk_hostname_entry = tk.Entry(fields_frame, width=22)
yk_hostname_entry.grid(row=0, column=1, pady=3)
tip(yk_hostname_entry, "Enter the hostname for this Yubikey.")

tk.Label(fields_frame, text="Serial Number:", anchor="w", width=13).grid(row=1, column=0, sticky="w", pady=3, padx=(0,4))
yk_serial_entry = tk.Entry(fields_frame, width=22)
yk_serial_entry.grid(row=1, column=1, pady=3)
tip(yk_serial_entry, "Enter the serial number for this Yubikey.")

tk.Label(fields_frame, text="Checked By:",   anchor="w", width=13).grid(row=2, column=0, sticky="w", pady=3, padx=(0,4))
yk_checked_by_entry = tk.Entry(fields_frame, width=22)
yk_checked_by_entry.grid(row=2, column=1, pady=3)
tip(yk_checked_by_entry, "Enter the name of the person who checked this item.")

tk.Label(fields_frame, text="Shelf:",         anchor="w", width=13).grid(row=3, column=0, sticky="w", pady=3, padx=(0,4))
yk_shelf_var = tk.StringVar()
yk_shelf_dropdown = ttk.Combobox(fields_frame, textvariable=yk_shelf_var, width=20, state="readonly")
yk_shelf_dropdown.grid(row=3, column=1, pady=3)
tip(yk_shelf_dropdown, "Select the shelf where this Yubikey will be stored.")

tk.Label(fields_frame, text="Status:",        anchor="w", width=13).grid(row=4, column=0, sticky="w", pady=3, padx=(0,4))
yk_status_var = tk.StringVar()
ttk.Combobox(fields_frame, textvariable=yk_status_var, values=STATUS_CHOICES,
             width=20, state="readonly").grid(row=4, column=1, pady=3)

tk.Label(fields_frame, text="Remarks:",       anchor="w", width=13).grid(row=5, column=0, sticky="w", pady=3, padx=(0,4))
yk_remarks_var = tk.StringVar()
ttk.Combobox(fields_frame, textvariable=yk_remarks_var,
             values=[""] + STATUS_CHOICES, width=20, state="readonly").grid(row=5, column=1, pady=3)

yk_stage_btn_row = tk.Frame(yk_input_frame)
yk_stage_btn_row.pack(pady=(4, 2))
tip(tk.Button(yk_stage_btn_row, text="STAGE ITEM", command=yk_put_item,
          bg="#2980b9", fg="white", width=14),
    "Add this item to the staging queue.").pack(side="left", padx=3)
tip(tk.Button(yk_stage_btn_row, text="UPDATE ITEM", command=yk_update_item,
          bg="#1a5276", fg="white", width=13),
    "Save edits to the selected staged item.").pack(side="left", padx=3)
tip(tk.Button(yk_stage_btn_row, text="↻", command=_yk_clear_fields, width=3),
    "Clear all input fields.").pack(side="left", padx=3)

yk_import_btn_row = tk.Frame(yk_input_frame)
yk_import_btn_row.pack(pady=(0, 2))
tip(tk.Button(yk_import_btn_row, text="📥 IMPORT EXCEL", command=yk_import_excel_to_staging,
          bg="#1e8449", fg="white", width=20),
    "Import Yubikey items from an Excel file directly into the staging list.").pack()

tip(tk.Label(yk_input_frame, text="Staged Items", fg="green", font=("Helvetica", 9, "bold")),
    "Items waiting to be committed to the Yubikey warehouse.").pack(anchor="w", pady=(6, 0))
yk_staged_listbox = tk.Listbox(yk_input_frame, width=36)
yk_staged_listbox.pack(fill="both", expand=True, pady=3)
yk_staged_listbox.bind("<<ListboxSelect>>", select_yk_staged_item)
tip(yk_staged_listbox, "Click an item to load it for editing. Select then click UPDATE to modify.")

yk_stage_action_row = tk.Frame(yk_input_frame)
yk_stage_action_row.pack(pady=3)
tip(tk.Button(yk_stage_action_row, text="CLEAR ITEMS",   command=yk_remove_from_staging, width=13),
    "Remove selected staged item, or all items if none selected.").pack(side="left", padx=2)
tip(tk.Button(yk_stage_action_row, text="↻", command=lambda: yk_staged_listbox.selection_clear(0, tk.END), width=3),
    "Deselect the current staged item so CLEAR ITEMS will clear all.").pack(side="left", padx=2)
tip(tk.Button(yk_stage_action_row, text="PUT WAREHOUSE", command=yk_put_warehouse,       width=13),
    "Commit all staged items to the Yubikey warehouse and generate QR codes.").pack(side="left", padx=2)

# ── Middle: Shelf Control + View + Search + Pull + Tables ─
yk_shelf_mid = tk.Frame(yk_row1)
yk_shelf_mid.pack(side="left", fill="both", expand=True, padx=5)

# ── Shelf + View sub-row ───────────────────────────────────
yk_shelf_view_row = tk.Frame(yk_shelf_mid)
yk_shelf_view_row.pack(fill="x")

yk_shelf_ctrl_frame = tk.LabelFrame(yk_shelf_view_row, text="Shelf Control & Management", padx=10, pady=5)
yk_shelf_ctrl_frame.pack(side="left", fill="both", expand=True)
tip(yk_shelf_ctrl_frame, "Manage shelf availability and add/remove shelves for Yubikey.")

yk_status_ctrl = tk.LabelFrame(yk_shelf_ctrl_frame, text="Status Control", padx=8, pady=5)
yk_status_ctrl.pack(fill="x", pady=(0, 5))
tip(yk_status_ctrl, "Mark a shelf as FULL (no more items) or AVAILABLE.")
yk_shelf_control_var = tk.StringVar()
yk_shelf_control_dropdown = ttk.Combobox(yk_status_ctrl, textvariable=yk_shelf_control_var, width=22, state="readonly")
yk_shelf_control_dropdown.pack(side="left", padx=5)
tip(yk_shelf_control_dropdown, "Select the shelf whose status you want to change.")
tip(tk.Button(yk_status_ctrl, text="SET FULL",      command=lambda: yk_set_shelf_status("FULL"),      width=10),
    "Mark selected shelf as FULL — prevents new items from being placed there.").pack(side="left", padx=3)
tip(tk.Button(yk_status_ctrl, text="SET AVAILABLE", command=lambda: yk_set_shelf_status("AVAILABLE"), width=12),
    "Mark selected shelf as AVAILABLE — allows new items to be placed.").pack(side="left", padx=3)
tip(tk.Button(yk_status_ctrl, text="↻",             command=yk_reset_shelf_control,                   width=3),
    "Clear the shelf selection.").pack(side="left", padx=3)

yk_add_remove = tk.LabelFrame(yk_shelf_ctrl_frame, text="Add / Remove", padx=8, pady=5)
yk_add_remove.pack(fill="x")
tip(yk_add_remove, "Add a new shelf by typing its name, or remove an existing empty shelf.")
yk_remove_shelf_var = tk.StringVar()
yk_remove_shelf_dropdown = ttk.Combobox(yk_add_remove, textvariable=yk_remove_shelf_var, width=22)
yk_remove_shelf_dropdown.pack(side="left", padx=5)
tip(yk_remove_shelf_dropdown, "Type a new shelf name to add, or select an existing one to remove.")
tip(tk.Button(yk_add_remove, text="ADD",    command=yk_add_shelf    ), "Add the typed shelf name as a new shelf.").pack(side="left", padx=3)
tip(tk.Button(yk_add_remove, text="REMOVE", command=yk_remove_shelf  ), "Remove the selected shelf (only if it has no items).").pack(side="left", padx=3)
tip(tk.Button(yk_add_remove, text="↻",      command=yk_reset_shelf_addition, width=3), "Clear the shelf name field.").pack(side="left", padx=3)

# View Panel — sits to the right of Shelf Control
yk_view_frame = tk.LabelFrame(yk_shelf_view_row, text="View", padx=10, pady=5)
yk_view_frame.pack(side="left", fill="y", padx=(5, 0))
tip(yk_view_frame, "Switch between different table views for Yubikey.")
for text, cmd, tooltip in [
    ("SHOW WAREHOUSE", yk_show_warehouse,     "View all Yubikeys currently in the warehouse."),
    ("PULL HISTORY",   yk_show_pullouts,      "View Yubikeys that have been pulled out."),
    ("SHELF STATUS",   yk_show_available,     "View each shelf's status and how many items it holds."),
    ("QR LABELS",      yk_open_label_manager, "Open and manage printable QR label PDF files."),
    ("VIEW EXCEL",     yk_view_excel,         "Open the last Excel file generated by GENERATE FILES."),
]:
    tip(tk.Button(yk_view_frame, text=text, command=cmd, width=15), tooltip).pack(anchor="w", pady=3)

# Search & Filter Panel
yk_pullout_frame = tk.LabelFrame(yk_shelf_mid, text="Yubikey Warehouse", padx=10, pady=8)
yk_pullout_frame.pack(fill="x", pady=5)

yk_search_filter = tk.LabelFrame(yk_pullout_frame, text="Search & Filter", padx=8, pady=5)
yk_search_filter.pack(fill="x", pady=(0, 5))

yk_sf_row1 = tk.Frame(yk_search_filter)
yk_sf_row1.pack(fill="x", pady=(0, 3))
tk.Label(yk_sf_row1, text="Search:").pack(side="left", padx=(5, 2))
yk_search_entry = tk.Entry(yk_sf_row1, width=20)
yk_search_entry.pack(side="left", padx=(0, 2))
yk_search_entry.bind("<Return>",     lambda e: yk_search_item())
yk_search_entry.bind("<KeyRelease>", yk_pull_search_live)

tk.Label(yk_sf_row1, text="Shelf:").pack(side="left", padx=(5, 2))
yk_shelf_filter_var = tk.StringVar()
yk_shelf_filter_dropdown = ttk.Combobox(yk_sf_row1, textvariable=yk_shelf_filter_var, width=16, state="readonly")
yk_shelf_filter_dropdown.pack(side="left", padx=(0, 8))

tk.Label(yk_sf_row1, text="Status:").pack(side="left", padx=(5, 2))
yk_status_filter_var = tk.StringVar()
ttk.Combobox(yk_sf_row1, textvariable=yk_status_filter_var,
             values=[""] + STATUS_CHOICES, width=14, state="readonly").pack(side="left", padx=(0, 8))

tip(tk.Button(yk_sf_row1, text="🔍", command=yk_search_item,  width=3),
    "Search and filter the Yubikey warehouse table.").pack(side="left", padx=3)
tip(tk.Button(yk_sf_row1, text="↻",  command=yk_clear_filters, width=3),
    "Clear all search and filter fields.").pack(side="left", padx=3)

yk_sf_row2 = tk.Frame(yk_search_filter)
yk_sf_row2.pack(fill="x", pady=(0, 2))
yk_date_from_var = tk.StringVar()
yk_date_to_var   = tk.StringVar()
_date_picker_widget(yk_sf_row2, yk_date_from_var, "From:").pack(side="left", padx=(5, 8))
_date_picker_widget(yk_sf_row2, yk_date_to_var,   "To:").pack(side="left", padx=(0, 5))

# Pull Out Panel
yk_pull_action = tk.LabelFrame(yk_pullout_frame, text="Pull Out", padx=8, pady=5)
yk_pull_action.pack(fill="x", pady=(5, 0))

yk_po_row1 = tk.Frame(yk_pull_action)
yk_po_row1.pack(fill="x", pady=(0, 3))
tk.Label(yk_po_row1, text="Pull Reason:").pack(side="left", padx=(5, 2))
yk_pull_reason_var = tk.StringVar()
yk_pull_reason_entry = ttk.Combobox(yk_po_row1, textvariable=yk_pull_reason_var, width=22)
yk_pull_reason_entry.pack(side="left", padx=(0, 8))
yk_pull_date_from_var = tk.StringVar()
yk_pull_date_to_var   = tk.StringVar()
tip(tk.Button(yk_po_row1, text="↻",  command=yk_reset_pull_out,  width=3),
    "Clear pull-out fields and reset the view.").pack(side="left", padx=2)
tip(tk.Button(yk_po_row1, text="WAREHOUSE PULL", command=yk_pull_item, width=16),
    "Pull the selected Yubikey out of the warehouse. A pull reason is required.").pack(side="left", padx=(10, 3))

# Status bar
yk_status_bar = tk.Frame(yk_shelf_mid)
yk_status_bar.pack(fill="x")
yk_full_label   = tk.Label(yk_status_bar, text="FULL Shelves: None", fg="red");  yk_full_label.pack(side="left", padx=10)
yk_search_label = tk.Label(yk_status_bar, text="", fg="blue");  yk_search_label.pack(side="left", padx=10)
yk_status_label = tk.Label(yk_status_bar, text="", fg="green"); yk_status_label.pack(side="left", padx=10)

# Table toolbar
yk_table_toolbar = tk.Frame(yk_shelf_mid)
yk_table_toolbar.pack(fill="x", padx=5, pady=(2, 0))

def yk_toggle_select_all():
    if tree_yk_pullouts.winfo_ismapped():
        all_iids = list(tree_yk_pullouts.get_children())
        checked  = [iid for iid in all_iids if yk_pull_row_checks.get(iid)]
        new_state = len(checked) < len(all_iids)
        for iid in all_iids:
            yk_pull_row_checks[iid] = new_state
            tree_yk_pullouts.set(iid, "CP0", "☑" if new_state else "☐")
        return
    all_iids  = list(tree_yk_warehouse.get_children())
    checked   = [iid for iid in all_iids if yk_row_checks.get(iid)]
    new_state = len(checked) < len(all_iids)
    for iid in all_iids:
        yk_row_checks[iid] = new_state
        hostname = tree_yk_warehouse.item(iid, "values")[2]
        if new_state: yk_persistent_checks.add(hostname)
        else:         yk_persistent_checks.discard(hostname)
        tree_yk_warehouse.set(iid, "C0", "☑" if new_state else "☐")
    _yk_refresh_select_all_label()

yk_select_all_btn = tk.Button(yk_table_toolbar, text="SELECT ALL",
    command=yk_toggle_select_all, width=12)
yk_select_all_btn.pack(side="left", padx=(0, 6))
tip(yk_select_all_btn, "Select or deselect all visible rows in the active table.")

tip(tk.Button(yk_table_toolbar, text="GENERATE FILES", command=yk_generate_stored_qr,
          bg="#6c3483", fg="white", width=15),
    "Generate QR PNGs, PDF labels and Excel export for selected items.").pack(side="left", padx=(0, 6))
tip(tk.Button(yk_table_toolbar, text="VIEW QR", command=yk_view_stored_qr,
          bg="#1a5276", fg="white", width=10),
    "View all existing QR codes without regenerating.").pack(side="left", padx=(0, 6))

yk_back_to_stage_btn = tip(tk.Button(yk_table_toolbar, text="BACK TO STAGE",
          command=yk_unstage_from_warehouse, bg="#117a65", fg="white", width=15),
    "Move checked warehouse items back to the Item Staging list.")
yk_back_to_stage_btn.pack(side="left", padx=(0, 6))

yk_back_to_wh_btn = tip(tk.Button(yk_table_toolbar, text="BACK TO WAREHOUSE",
          command=yk_undo_pull, bg="#922b21", fg="white", width=18),
    "Restore checked pull history items back to the Yubikey warehouse.")
# Hidden by default; shown only when Pull History view is active

# Tables
yk_table_frame = tk.Frame(yk_shelf_mid)
yk_table_frame.pack(fill="both", expand=True, pady=5)

tree_yk_warehouse = ttk.Treeview(yk_table_frame,
    columns=("C0","C1","C2","C3","C4","C5","C6","C7","C8"), show='headings')
for col, text, width in zip(
    ("C0","C1","C2","C3","C4","C5","C6","C7","C8"),
    ("✔","QR","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Date"),
    (30, 180, 150, 130, 115, 125, 90, 140, 140)):
    tree_yk_warehouse.heading(col, text=text)
    tree_yk_warehouse.column(col, width=width)
tree_yk_warehouse.column("C0", anchor="center", stretch=False)
tree_yk_warehouse.bind("<<TreeviewSelect>>", yk_select_item)

tree_yk_pullouts = ttk.Treeview(yk_table_frame,
    columns=("CP0","C1","C2","C3","C4","C5","C6","C7","C8"), show='headings')
for col, text, width in zip(
    ("CP0","C1","C2","C3","C4","C5","C6","C7","C8"),
    ("✔","Hostname","Serial Number","Checked By","Shelf","Status","Remarks","Pull Reason","Date"),
    (30, 150, 130, 115, 125, 90, 140, 195, 140)):
    tree_yk_pullouts.heading(col, text=text)
    tree_yk_pullouts.column(col, width=width)
tree_yk_pullouts.column("CP0", anchor="center", stretch=False)
tree_yk_pullouts.bind("<<TreeviewSelect>>", yk_select_pull_item)

tree_yk_available = ttk.Treeview(yk_table_frame, columns=("C1","C2","C3","C4"), show='headings')
for col, text, width in zip(
        ("C1","C2","C3","C4"),
        ("Shelf","Status","Items Stored","Date Set Full"),
        (250, 140, 110, 200)):
    tree_yk_available.heading(col, text=text)
    tree_yk_available.column(col, width=width)

# Scrollbars for YK tables
yk_vsb = ttk.Scrollbar(yk_table_frame, orient="vertical")
yk_vsb.pack(side="right", fill="y")
yk_hsb = ttk.Scrollbar(yk_table_frame, orient="horizontal")
yk_hsb.pack(side="bottom", fill="x")
tree_yk_warehouse.configure(yscrollcommand=yk_vsb.set, xscrollcommand=yk_hsb.set)
tree_yk_pullouts.configure(yscrollcommand=yk_vsb.set,  xscrollcommand=yk_hsb.set)
tree_yk_available.configure(yscrollcommand=yk_vsb.set, xscrollcommand=yk_hsb.set)
tree_yk_warehouse.pack(fill="both", expand=True)


initialize_file()
update_all_shelf_dropdowns()
update_staged_display()
update_w2_staged_display()
update_yk_staged_display()
show_warehouse()
w2_show_warehouse()
yk_show_warehouse()

# Pre-load pull reason dropdowns from saved history
try:
    all_reasons_w1 = sorted(load_pullouts()["Pull Reason"].dropna().unique().tolist())
    pull_reason_filter_entry["values"] = [""] + all_reasons_w1
except Exception:
    pass
try:
    all_reasons_w2 = sorted(load_pullouts_w2()["Pull Reason"].dropna().unique().tolist())
    w2_pull_reason_filter_entry["values"] = [""] + all_reasons_w2
except Exception:
    pass
try:
    all_reasons_yk = sorted(load_pullouts_yk()["Pull Reason"].dropna().unique().tolist())
    yk_pull_reason_entry["values"] = [""] + all_reasons_yk
except Exception:
    pass

# Attach click-to-sort on all main treeviews
for _t in (tree_warehouse, tree_available, tree_pullouts, tree_qr,
           tree_w2_warehouse, tree_w2_available, tree_w2_pullouts, tree_w2_qr,
           tree_yk_warehouse, tree_yk_pullouts, tree_yk_available):
    attach_sort_headers(_t)

def on_main_close():
    if messagebox.askyesno("Exit", f"Log out '{current_user}' and exit the system?"):
        save_log("LOGOUT", f"Session ended for '{current_user}'")
        root.destroy()

root.protocol("WM_DELETE_WINDOW", on_main_close)
root.mainloop()